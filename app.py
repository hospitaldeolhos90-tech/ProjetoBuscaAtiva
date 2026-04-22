import streamlit as st
import json
import os
import streamlit.components.v1 as components
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import io
import re
import math
import zipfile
import base64
import csv
from datetime import datetime
import xml.etree.ElementTree as ET

# --- 1. CONFIGURAÇÃO MASTER ---
try:
    st.set_page_config(
        page_title="Hova | Master Intelligence",
        page_icon="💠",
        layout="wide",
        initial_sidebar_state="auto"
    )
except:
    pass

# --- 2. INICIALIZAÇÃO DE MEMÓRIA E DICIONÁRIOS ---
if 'lista_pendencias' not in st.session_state:
    st.session_state['lista_pendencias'] = []
if 'movimento_gravado' not in st.session_state:
    st.session_state['movimento_gravado'] = False
if 'base_deduplicada' not in st.session_state:
    st.session_state['base_deduplicada'] = None

# --- MEGA DICIONÁRIO HOVA ---
if 'dicionario_tipos' not in st.session_state:
    st.session_state['dicionario_tipos'] = pd.DataFrame([
        {"SIGLA": "C", "TRADUCAO": "Consulta Especializada"},
        {"SIGLA": "R", "TRADUCAO": "Retorno"},
        {"SIGLA": "RC", "TRADUCAO": "Retorno de Cirurgia"},
        {"SIGLA": "RT", "TRADUCAO": "Reconsulta"},
        {"SIGLA": "RL", "TRADUCAO": "Retorno de Lente"},
        {"SIGLA": "RO", "TRADUCAO": "Revisão de Óculos"},
        {"SIGLA": "CRA", "TRADUCAO": "Consulta Reavaliação Anual"},
        {"SIGLA": "CDI", "TRADUCAO": "Consulta Diagnóstico Inicial"},
        {"SIGLA": "C_INTER", "TRADUCAO": "Consulta de Intercorrência"},
        {"SIGLA": "PO", "TRADUCAO": "Pós Operatório"},
        {"SIGLA": "AC", "TRADUCAO": "Avaliação de Cirurgia"},
        {"SIGLA": "AGL", "TRADUCAO": "Avaliação de Glaucoma"},
        {"SIGLA": "AR", "TRADUCAO": "Avaliação com Retinólogo"},
        {"SIGLA": "PREOP", "TRADUCAO": "Pré-Operatório"},
        {"SIGLA": "NAN", "TRADUCAO": "CÉLULA VAZIA NO DOCTORS"},
        {"SIGLA": "NONE", "TRADUCAO": "CÉLULA VAZIA NO DOCTORS"},
        {"SIGLA": "TESTEORTOP.", "TRADUCAO": "Teste Ortóptico"},
        {"SIGLA": "TESTE ORTOP", "TRADUCAO": "Teste Ortóptico"},
        {"SIGLA": "ORTOPTICO", "TRADUCAO": "Teste Ortóptico"},
        {"SIGLA": "ORTÓPTICO", "TRADUCAO": "Teste Ortóptico"},
        {"SIGLA": "1 CAT", "TRADUCAO": "1ª Avaliação"},
        {"SIGLA": "2 CAT", "TRADUCAO": "2ª Avaliação"},
        {"SIGLA": "3 CAT", "TRADUCAO": "3ª Avaliação"},
        {"SIGLA": "1 CCAT", "TRADUCAO": "1ª Avaliação"},
        {"SIGLA": "2 CCAT", "TRADUCAO": "2ª Avaliação"},
        {"SIGLA": "3 CCAT", "TRADUCAO": "3ª Avaliação"},
        {"SIGLA": "1ªCAT", "TRADUCAO": "1ª Avaliação"},
        {"SIGLA": "2ªCAT", "TRADUCAO": "2ª Avaliação"},
        {"SIGLA": "3ªCAT", "TRADUCAO": "3ª Avaliação"},
        {"SIGLA": "AGF", "TRADUCAO": "Angiofluoresceinografia"},
        {"SIGLA": "RF", "TRADUCAO": "Retinografia Fluorescente"},
        {"SIGLA": "RD", "TRADUCAO": "Retinografia Digital"},
        {"SIGLA": "TOPO", "TRADUCAO": "Topografia"},
        {"SIGLA": "CVC", "TRADUCAO": "Campo Visual Computadorizado"},
        {"SIGLA": "IOL", "TRADUCAO": "Biometria Óptica (IOL Master)"},
        {"SIGLA": "BIO", "TRADUCAO": "Biometria Ultrassônica"},
        {"SIGLA": "OCT", "TRADUCAO": "Tomografia de Coerência Óptica"},
        {"SIGLA": "ANGIO-OCT", "TRADUCAO": "Angio-OCT"},
        {"SIGLA": "PTC", "TRADUCAO": "Pentacam"},
        {"SIGLA": "MBG", "TRADUCAO": "Meibomiografia"},
        {"SIGLA": "MEIBO", "TRADUCAO": "Meibomiografia"},
        {"SIGLA": "TSC", "TRADUCAO": "Teste de Schirmer"},
        {"SIGLA": "TSH", "TRADUCAO": "Teste de Sobrecarga Hídrica"},
        {"SIGLA": "MR", "TRADUCAO": "Mapeamento de Retina"},
        {"SIGLA": "US", "TRADUCAO": "Ultrassom (Ecografia)"},
        {"SIGLA": "CDPO", "TRADUCAO": "Curva Diária de Pressão"},
        {"SIGLA": "GONIO", "TRADUCAO": "Gonioscopia"},
        {"SIGLA": "PAQ", "TRADUCAO": "Paquimetria"},
        {"SIGLA": "MICRO", "TRADUCAO": "Microscopia Especular"},
        {"SIGLA": "AV", "TRADUCAO": "Acuidade Visual"},
        {"SIGLA": "EX", "TRADUCAO": "Tonometria"},
        {"SIGLA": "TO", "TRADUCAO": "Teste do Olhinho"},
        {"SIGLA": "ISHIHARA", "TRADUCAO": "Teste de Ishihara"},
        {"SIGLA": "VVPP", "TRADUCAO": "Vitrectomia Posterior"},
        {"SIGLA": "ESTRABISMO", "TRADUCAO": "Cirurgia de Estrabismo"},
        {"SIGLA": "CICLOFOTO", "TRADUCAO": "Ciclofotocoagulação"},
        {"SIGLA": "TRAB", "TRADUCAO": "Trabeculectomia"},
        {"SIGLA": "ANTIVEG", "TRADUCAO": "Aplicação Intravítrea"},
        {"SIGLA": "TL", "TRADUCAO": "Teste de Lente"},
        {"SIGLA": "TLR", "TRADUCAO": "Teste de Lente Rígida"},
        {"SIGLA": "TLESCL", "TRADUCAO": "Teste de Lente Escleral"},
        {"SIGLA": "BL", "TRADUCAO": "Busca de Lente"},
        {"SIGLA": "CLGN", "TRADUCAO": "Compra Lente Gelatinosa"},
        {"SIGLA": "CLGP", "TRADUCAO": "Compra Lente Gelatinosa Positiva"},
        {"SIGLA": "CLIO", "TRADUCAO": "Compra Lente Intra Ocular"},
        {"SIGLA": "CLMF", "TRADUCAO": "Compra Lente Multifocal"},
        {"SIGLA": "CLR", "TRADUCAO": "Compra Lente Rígida"},
        {"SIGLA": "CLT", "TRADUCAO": "Compra Lente Tórica"},
        {"SIGLA": "CLTERAPEUTICA", "TRADUCAO": "Compra Lente Terapêutica"},
        {"SIGLA": "YAG", "TRADUCAO": "Capsulotomia Yag Laser"},
        {"SIGLA": "FOTO", "TRADUCAO": "Fotocoagulação a Laser"},
        {"SIGLA": "IRI", "TRADUCAO": "Iridectomia"},
        {"SIGLA": "FOTOTRAB", "TRADUCAO": "Fototrabeculoplastia"},
        {"SIGLA": "ILIO", "TRADUCAO": "Procedimento ILIO"},
        {"SIGLA": "LP", "TRADUCAO": "Luz Pulsada"},
        {"SIGLA": "IC", "TRADUCAO": "Instrução Cirúrgica"},
        {"SIGLA": "IC FACO", "TRADUCAO": "Instrução de Cirurgia de Faco"},
        {"SIGLA": "IC_FACO", "TRADUCAO": "Instrução de Cirurgia de Faco"},
        {"SIGLA": "IC VITRE", "TRADUCAO": "Instrução de Vitrectomia"},
        {"SIGLA": "IC_VITRE", "TRADUCAO": "Instrução de Vitrectomia"},
        {"SIGLA": "IC PTERIGIO", "TRADUCAO": "Instrução de Pterígio"},
        {"SIGLA": "IC_PTERIGIO", "TRADUCAO": "Instrução de Pterígio"},
        {"SIGLA": "IC CICLO", "TRADUCAO": "Instrução de Cirurgia de Ciclofoto"},
        {"SIGLA": "IC_CICLO", "TRADUCAO": "Instrução de Cirurgia de Ciclofoto"},
        {"SIGLA": "IC CALAZIO", "TRADUCAO": "Instrução de Calázio"},
        {"SIGLA": "IC_CALAZIO", "TRADUCAO": "Instrução de Calázio"},
        {"SIGLA": "RIC", "TRADUCAO": "Risco Cirúrgico"},
        {"SIGLA": "RET.CIR", "TRADUCAO": "Retoque de Cirurgia"},
        {"SIGLA": "CO", "TRADUCAO": "Conversa"},
        {"SIGLA": "EEX", "TRADUCAO": "Entrega de Exames"},
        {"SIGLA": "TFD", "TRADUCAO": "Entrega de TFD"},
        {"SIGLA": "PDC", "TRADUCAO": "Pendência de Marcação"},
        {"SIGLA": "MD", "TRADUCAO": "Medicação"},
        {"SIGLA": "CE", "TRADUCAO": "Compra de Estojo"},
        {"SIGLA": "RP", "TRADUCAO": "Retirar Ponto"},
        {"SIGLA": "RPG", "TRADUCAO": "Restante Pagamento"},
        {"SIGLA": "PAG.A", "TRADUCAO": "Pagamento Antecipado"},
        {"SIGLA": "TX", "TRADUCAO": "Taxas Administrativas"},
        {"SIGLA": "EXLAB", "TRADUCAO": "Exames Laboratoriais"},
        {"SIGLA": "RE", "TRADUCAO": "2ª Via de Exame"},
        {"SIGLA": "REP.EX", "TRADUCAO": "Repetir Exame"}
    ])

# --- REGRAS DE OURO DA CLÍNICA ---
tipos_ignorados = ['R', 'CRA', 'EEX', '1', '2', '3', '1ª', '2ª', '3ª', 'CAT', 'CCAT',
                   '1CAT', '2CAT', '3CAT', '1CCAT', '2CCAT', '3CCAT', '1ªCAT', '2ªCAT', '3ªCAT']

agendas_enviar = [
    'Altair Rosa', 'Denise Matos', 'Felipe Ferreira', 'Francesca de As',
    'Gabriel Conde', 'Gabriel Lemos', 'Gustavo Sampaio', 'Mariluci Mendes',
    'Luis Antonio', 'Rodrigo Costa', 'Vera Lucia',
    'Centro Cirúrgico', 'Exames', 'Enfermagem', 'Catarata',
    'Mateus Barbosa', 'Maxwell dos Reis', 'Lucas', 'Victor'
]

exames_dilatam = ['AGF', 'RF', 'RD', 'MR', 'OCT', 'YAG', 'FOTO', 'ANGIO-OCT', 'IRI', 'FOTOTRAB']

def get_base64(bin_file):
    try:
        with open(bin_file, 'rb') as f: return base64.b64encode(f.read()).decode()
    except: return None

def colorir_status(row):
    status = str(row.get('STATUS DO PACIENTE', ''))
    if '✅ CONFIRMADO' in status:
        return ['background-color: #d4edda; color: #155724'] * len(row)
    elif '❌ CANCELADO' in status:
        return ['background-color: #f8d7da; color: #721c24'] * len(row)
    elif '💬 MENSAGEM RECEBIDA' in status:
        return ['background-color: #cce5ff; color: #004085'] * len(row)
    elif '⏳ SEM RESPOSTA' in status:
        return ['background-color: #fff3cd; color: #856404'] * len(row)
    return [''] * len(row)

# --- 3. DESIGN "HYBRID ELITE" ---
st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap');
    @keyframes fadeInUp {{ from {{ opacity: 0; transform: translateY(20px); }} to {{ opacity: 1; transform: translateY(0); }} }}
    @keyframes auroraBG {{ 0% {{ background-position: 0% 50%; }} 50% {{ background-position: 100% 50%; }} 100% {{ background-position: 0% 50%; }} }}
    .stApp {{ background: linear-gradient(-45deg, #f0f7f4, #e0efeb, #ffffff, #cde0dc); background-size: 400% 400% !important; animation: auroraBG 15s ease infinite !important; font-family: 'Outfit', sans-serif; }}
    [data-testid="stSidebar"] {{ background-color: #1e3d3a !important; border-right: 2px solid rgba(0, 255, 204, 0.2); }}
    [data-testid="stSidebar"] .stMarkdown p, [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {{ color: #ffffff !important; font-family: 'Outfit', sans-serif; }}
    .premium-title {{ color: #1e3d3a; font-weight: 600 !important; font-size: 2.2rem; text-align: center; text-transform: uppercase; margin-top: 50px !important; margin-bottom: 40px; display: flex; align-items: center; justify-content: center; gap: 15px; animation: fadeInUp 0.6s ease-out; }}
    @media (max-width: 768px) {{ .premium-title {{ font-size: 1.4rem !important; margin-top: 20px !important; }} .master-card {{ padding: 25px 15px !important; border-radius: 25px !important; }} div.stButton > button {{ height: 60px !important; font-size: 1.1rem !important; }} .stTabs [data-baseweb="tab"] {{ height: 60px !important; }} .stTabs [data-baseweb="tab"] p {{ font-size: 0.8rem !important; }} }}
    .stTabs [data-baseweb="tab-list"] {{ display: flex !important; width: 100% !important; gap: 8px !important; background-color: transparent !important; padding: 5px; }}
    .stTabs [data-baseweb="tab"] {{ flex-grow: 1 !important; height: 85px !important; background-color: #1e3d3a !important; border-radius: 15px 15px 0 0 !important; transition: all 0.5s cubic-bezier(0.175, 0.885, 0.32, 1.275) !important; }}
    .stTabs [data-baseweb="tab"] p {{ font-size: 1.1rem !important; font-weight: 700 !important; color: #FFFFFF !important; }}
    .stTabs [data-baseweb="tab"]:hover {{ transform: translateY(-10px) !important; background-color: #2f6c68 !important; }}
    .stTabs [aria-selected="true"] {{ background-color: #2f6c68 !important; border-bottom: 6px solid #00ffcc !important; }}
    .master-card {{ background: rgba(255, 255, 255, 0.7); backdrop-filter: blur(20px); border-radius: 40px; padding: 60px 80px; box-shadow: 0 40px 80px rgba(0,0,0,0.08); text-align: center; border: 1px solid rgba(255, 255, 255, 0.4); animation: fadeInUp 1s ease-out; }}
    div.stButton > button, div.stDownloadButton > button {{ background: linear-gradient(135deg, #1e3d3a 0%, #2f6c68 100%) !important; color: white !important; border-radius: 60px !important; height: 60px !important; width: 100% !important; font-size: 1.1rem !important; font-weight: 700 !important; text-transform: uppercase !important; border: none !important; transition: 0.4s all cubic-bezier(0.175, 0.885, 0.32, 1.275) !important; letter-spacing: 1px; }}
    div.stButton > button:hover, div.stDownloadButton > button:hover {{ transform: scale(1.07) translateY(-5px) !important; box-shadow: 0 20px 45px rgba(30, 61, 58, 0.35) !important; }}
    .response-box {{ background: white; padding: 25px; border-radius: 22px; border-top: 8px solid #1e3d3a; box-shadow: 0 10px 30px rgba(0,0,0,0.04); margin-bottom: 25px; text-align: left; animation: fadeInUp 0.8s ease-out; }}
    .penc-row {{ background: #f8fafc; padding: 15px; border-radius: 12px; margin-bottom: 10px; font-size: 0.9rem; color: #1e3d3a; font-weight: 700; border-left: 5px solid #1e3d3a; border: 1px solid #e2e8f0; }}
    .footer-master {{ text-align: center; color: #1e3d3a; font-weight: 600; letter-spacing: 10px; margin-top: 80px; margin-bottom: 30px; font-size: 1rem; opacity: 0.8; animation: fadeInUp 1.2s ease-out; }}
    .led-green {{ color: #00ffcc; text-shadow: 0 0 10px #00ffcc; }}
    </style>
""", unsafe_allow_html=True)

# --- 4. FUNÇÕES DE LÓGICA E DEDUPLICAÇÃO ---
def limpar_num(num): return "".join(filter(str.isdigit, str(num)))[-8:]

def eh_celular(num):
    n = "".join(filter(str.isdigit, str(num)))
    while n.startswith('0'): n = n[1:]
    if len(n) >= 10: return n[-9] == '9' or n[-8] in ['7', '8', '9']
    elif len(n) >= 8: return n[-8] in ['7', '8', '9']
    return False

def eh_numero_falso(num):
    n = str(num).strip()
    if len(n) < 8: return True
    if len(set(n)) == 1: return True
    if '12345678' in n: return True
    return False

def formatar_telefone_real(num):
    txt = str(num).strip().upper()
    if txt in ['NAN', 'NONE', '<NA>', '']: return ""
    if "(000)" in txt or "___" in txt: return ""
    digitos = re.sub(r'\D', '', txt)
    if not digitos: return ""
    if len(set(digitos)) == 1 and digitos[0] == '0': return ""
    while digitos.startswith('0') and len(digitos) > 9: digitos = digitos[1:]
    return digitos

def pegar_nome_curto(nome_completo):
    try:
        if not nome_completo or str(nome_completo).strip() == "": return ""
        nome_limpo = re.sub(r'[^A-Za-zÀ-ÖØ-öø-ÿ\s]', '', str(nome_completo)).strip()
        partes = nome_limpo.split()
        if len(partes) == 0: return ""
        if len(partes) == 1: return partes[0].capitalize()
        p1 = partes[0].capitalize()
        if partes[1].upper() in ['DE', 'DA', 'DO', 'DOS', 'DAS', 'E']:
            p2 = partes[1].lower()
            if len(partes) > 2: return f"{p1} {p2} {partes[2].capitalize()}"
            else: return f"{p1} {p2}"
        else: return f"{p1} {partes[1].capitalize()}"
    except: return ""

def traduzir_tipo(tipo_raw, dict_df):
    tipo = str(tipo_raw).strip().upper()
    if tipo.startswith("CI_") or tipo.startswith("CI ") or tipo.startswith("CI.") or tipo.startswith("IC ") or tipo.startswith("IC_"):
        medico_cirurgia = tipo[3:].replace(".", " ").replace("_", " ").title()
        return f"Cirurgia Dr(a). {medico_cirurgia}"
    if "/" in tipo:
        partes = tipo.split("/")
        traducoes = []
        for p in partes:
            p = p.strip()
            match = dict_df[dict_df['SIGLA'].str.upper() == p]
            if not match.empty: traducoes.append(match.iloc[0]['TRADUCAO'])
            else: traducoes.append(p)
        if len(traducoes) > 1: return ", ".join(traducoes[:-1]) + " e " + traducoes[-1]
        return traducoes[0] if traducoes else "DESCONHECIDO"
    match = dict_df[dict_df['SIGLA'].str.upper() == tipo]
    return match.iloc[0]['TRADUCAO'] if not match.empty else "DESCONHECIDO"

def resolver_nome_completo(nome_arquivo):
    n = str(nome_arquivo).upper()
    if "GABRIELL" in n: return "Gabriel Lemos"
    if "GABRIEL" in n: return "Gabriel Conde"
    if "ALTAIR" in n: return "Altair Rosa"
    if "DENISE" in n: return "Denise Matos"
    if "FELIPE" in n: return "Felipe Ferreira"
    if "FRANCESCA" in n or "FRAN" in n: return "Francesca de As"
    if "GUSTAVO" in n: return "Gustavo Sampaio"
    if "MARILUCI" in n: return "Mariluci Mendes"
    if "LUIS" in n: return "Luis Antonio"
    if "RODRIGO" in n: return "Rodrigo Costa"
    if "VERA" in n: return "Vera Lucia"
    if "CIRURGIA" in n: return "Centro Cirúrgico"
    if "LENTE" in n: return "Setor de Lentes de Contato"
    if "ENFERMAG" in n: return "Enfermagem"
    if "EXAME" in n: return "Exames"
    if "GLAUCOMA" in n: return "Glaucoma"
    if "LAUDOS" in n: return "Laudos"
    if "VAGNER" in n: return "Vagner Diniz"
    if "LUCAS" in n: return "Lucas"
    if "CATARATA" in n: return "Catarata"
    if "FARMACIA" in n: return "Farmacia"
    if "MATEUS" in n: return "Mateus Barbosa"
    if "MAXWELL" in n: return "Maxwell dos Reis"
    if "VICTOR" in n: return "Victor"
    partes = n.rsplit('.', 1)[0].replace(" ", "_").split("_")
    return partes[-1].title()

def formatar_brasileiro_sem_hora(txt):
    txt = str(txt).strip()
    if txt.lower() in ['nan', 'none', 'nat', '<na>', '']: return ""
    if re.match(r'^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}', txt): txt = txt.split(' ')[0]
    if re.match(r'^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}', txt): txt = txt.split(' ')[0]
    match_iso = re.fullmatch(r'(\d{4})-(\d{2})-(\d{2})', txt)
    if match_iso: return f"{match_iso.group(3)}/{match_iso.group(2)}/{match_iso.group(1)}"
    return txt

def organizar_mensagens_lista(lista_msgs):
    todas_msgs = set()
    for val in lista_msgs:
        val_str = str(val).strip()
        if val_str and val_str.lower() not in ['nan', 'none', 'nat', '<na>']:
            for l in val_str.split('\n'):
                l = l.strip()
                if l:
                    l = l.replace(' 00:00:00', '')
                    l = re.sub(r'\s*\b\d{2}:\d{2}(:\d{2})?\b', '', l).strip()
                    todas_msgs.add(l)
    def date_key(txt):
        match = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4})', txt)
        if match:
            try: return pd.to_datetime(match.group(1), format='%d/%m/%Y', dayfirst=True)
            except: return pd.Timestamp.min
        return pd.Timestamp.min
    return "\n".join(sorted(list(todas_msgs), key=date_key))

# --- FUNÇÃO DE PRÉ-VISUALIZAÇÃO ---
def load_excel_with_ui(file_obj, key_prefix):
    if file_obj.name.endswith('.csv'):
        try:
            df = pd.read_csv(file_obj, sep=';', dtype=str)
            if len(df.columns) < 2:
                file_obj.seek(0)
                df = pd.read_csv(file_obj, sep=',', dtype=str)
        except:
            file_obj.seek(0)
            df = pd.read_csv(file_obj, sep=',', dtype=str)
        st.markdown(f"**Pré-visualização ({file_obj.name}):**")
        st.dataframe(df.head(5), use_container_width=True)
        return df
    xl = pd.ExcelFile(file_obj)
    if len(xl.sheet_names) > 1:
        sheet = st.selectbox(f"Qual aba processar? ({file_obj.name})", xl.sheet_names, key=f"sheet_{key_prefix}")
    else:
        sheet = xl.sheet_names[0]
    df = pd.read_excel(file_obj, sheet_name=sheet, dtype=str)
    st.markdown(f"**Pré-visualização da aba '{sheet}':**")
    st.dataframe(df.head(5), use_container_width=True)
    return df

# --- MOTOR DE TELEFONES ---
def processar_telefones_avancado(row, cols_telefone):
    def _eh_lixo(d):
        if not d or len(d) < 8: return True
        if len(set(d)) == 1: return True
        if '12345678' in d: return True
        return False

    def _formatar(t):
        txt = str(t).strip().upper()
        if not txt or txt in ['NAN', 'NONE', '<NA>', '-', '']: return ""
        if re.match(r'^[\s_\-\(\)\.]+$', txt): return ""
        digitos = re.sub(r'\D', '', txt)
        if digitos.startswith('55') and len(digitos) >= 12:
            digitos = digitos[2:]
        if _eh_lixo(digitos): return ""
        n = len(digitos)
        if n == 11: return digitos
        if n == 10:
            corpo = digitos[2:]
            if corpo[0] in ['2', '3', '4']: return f"{digitos} FIXO"
            if corpo[0] in ['7', '8', '9']: return digitos
            return ""
        if n == 8:
            if digitos[0] in ['7', '8', '9']: return digitos
            if digitos[0] in ['2', '3', '4']: return f"{digitos} FIXO"
            return ""
        if n == 9: return digitos if digitos[0] == '9' else ""
        if n >= 12: return _formatar(digitos[-11:])
        return ""

    todos_brutos = []
    for c in cols_telefone:
        val = str(row.get(c, '')).strip()
        if val and val.lower() not in ['nan', 'none', '<na>', '-', '']:
            todos_brutos.append(val.upper())

    if not todos_brutos: return ["", ""]

    formatados = []
    for bruto in todos_brutos:
        tem_w = 'W' in bruto
        fmt = _formatar(bruto)
        if fmt: formatados.append((fmt, tem_w))

    if not formatados: return ["", ""]

    whatsapps = [f for f, w in formatados if w]
    celulares  = [f for f, w in formatados if not w and 'FIXO' not in f]
    fixos      = [f for f, w in formatados if not w and 'FIXO' in f]

    if whatsapps:
        principal = whatsapps[0]
        resto = whatsapps[1:] + celulares + fixos
    elif celulares:
        principal = celulares[0]
        resto = celulares[1:] + fixos
    else:
        principal = fixos[0]
        resto = fixos[1:]

    vistos = {principal}
    adicionais = []
    for num in resto:
        if num not in vistos:
            vistos.add(num)
            adicionais.append(num)

    return [principal, " / ".join(adicionais)]

# --- MOTOR DE CONDUTAS ---
def formatar_conduta(c):
    c = str(c).upper().strip()
    if not c or c.lower() in ['nan', 'none', '<na>']: return ""
    c = re.sub(r'[+-]?\d+[,.]\d+\s*(OD|OE|AO)?', '', c)
    c = re.sub(r'(PERTO|LONGE)\s*[+-]?\d*[,.]?\d*', '', c)
    c = re.sub(r'(MANTER|USAR|CONTINUAR)\s*[ÓO]CULOS', '', c)
    c = re.sub(r'ENTREGUE\s*RECEITA', '', c)
    c = re.sub(r'RX\s*[ÓO]CULOS', '', c)
    c = c.strip()
    if not c: return "CONSULTA ANUAL"
    c = re.sub(r'\+?\s*LAUDO\b', '', c)
    c = re.sub(r'\bREPETIR\s+\w+\b', '', c)
    if re.fullmatch(r'C\.?', c): return "CONSULTA ANUAL"
    if re.search(r'\b(C/US/MR|US/MR/C|C/MR/US|MR/C|C/MR)\b', c): return "CONSULTA ANUAL"
    c = re.sub(r'\b(RETORNO|VER|CONTROLE|SEGUIMENTO)\s+EM\b', 'CONSULTA EM', c)
    c = re.sub(r'\bRETORNO\b', 'CONSULTA', c)
    c = re.sub(r'\b12\s*MESES?\b|\b1\s*ANO\b', 'ANUAL', c)
    c = re.sub(r'\b6\s*MESES?\b', 'SEMESTRAL', c)
    c = re.sub(r'\b3\s*MESES?\b', 'TRIMESTRAL', c)
    c = re.sub(r'\b2\s*MESES?\b', 'BIMESTRAL', c)
    c = re.sub(r'\b4\s*MESES?\b', '4 MESES', c)
    c = re.sub(r'\b5\s*MESES?\b', '5 MESES', c)
    for termo, nome in [('ANUAL', 'CONSULTA ANUAL'), ('SEMESTRAL', 'CONSULTA SEMESTRAL'),
                        ('TRIMESTRAL', 'CONSULTA TRIMESTRAL'), ('BIMESTRAL', 'CONSULTA BIMESTRAL')]:
        if re.search(rf'\b{termo}\b', c) and 'CONSULTA' not in c:
            c = re.sub(rf'\b{termo}\b', nome, c)
    c = re.sub(r'\s+', ' ', c).strip().strip('+').strip()
    return c if c else "CONSULTA ANUAL"

# --- LIMPEZA DE COLUNA PROXIMA ---
def limpar_proxima(c):
    """Remove RX OCULOS e variações da coluna PROXIMA, mantendo só o prazo clínico."""
    if not c or str(c).strip() == '': return ''
    c = str(c).strip()
    c = re.sub(r'(?i)([\+\s]*RX\s*[ÓO]CULOS[^\n]*)', '', c)
    c = re.sub(r'(?i)([\+\s]*MANTER\s*[ÓÒO]CULOS[^\n]*)', '', c)
    c = re.sub(r'(?i)([\+\s]*[ÓO]CULOS\s*A\s*PEDIDO[^\n]*)', '', c)
    c = re.sub(r'(?i)([\+\s]*LENTES\s*DE\s*CONTATO[^\n]*)', '', c)
    linhas = [l.strip() for l in c.split('\n') if l.strip()]
    # Só filtra por prazo se NÃO houver termos clínicos importantes
    termos_clinicos = ['FACO','LIO','CATARATA','VVPP','IIV','ANTIVEGF','FOTOCOAGULAÇÃO',
                       'FOTOCOAGULACAO','VITRECTOMIA','CICLOFOTO','TRABECULECT','PTERIGIO',
                       'PTERÍGIO','CALAZIO','CALÁZIO','TRANSPLANTE','ANEL','YAG','CAPSULOTOMIA']
    tem_clinico = any(any(t in l.upper() for t in termos_clinicos) for l in linhas)
    if tem_clinico:
        resultado = '\n'.join(linhas)  # mantém tudo quando tem termo clínico
    else:
        prioridade = [l for l in linhas if any(p in l.upper() for p in
            ['CONSULTA', 'SEGUIMENTO', 'TRIMESTRAL', 'SEMESTRAL', 'ANUAL', 'BIMESTRAL', 'MESES', 'DIAS', 'SEMANA'])]
        resultado = prioridade[0] if prioridade else (linhas[0] if linhas else '')
    return re.sub(r'\s+', ' ', resultado).strip().strip('+').strip()

def rank_conduta(c):
    c = str(c).upper()
    m_dias = re.search(r'(\d+)\s*DIAS?', c)
    if m_dias: return int(m_dias.group(1))
    m_meses = re.search(r'(\d+)\s*MES', c)
    if m_meses: return int(m_meses.group(1)) * 30
    if 'BIMESTRAL' in c: return 60
    if 'TRIMESTRAL' in c: return 90
    if 'SEMESTRAL' in c: return 180
    if 'ANUAL' in c: return 365
    desc_bonus = 0 if len(c) > 10 else 50
    return 999 + desc_bonus

def calcular_proxima_data(data_visita_str, conduta_str):
    try:
        data_visita = pd.to_datetime(str(data_visita_str).strip(), format='%d/%m/%Y', dayfirst=True, errors='coerce')
        if pd.isna(data_visita): return ""
        c = str(conduta_str).upper().strip()
        if not c or c in ['NAN', 'NONE', '']: return ""
        dias = 30
        m_dias = re.search(r'(\d+)\s*DIAS?', c)
        m_meses = re.search(r'(\d+)\s*MES', c)
        if m_dias: dias = int(m_dias.group(1))
        elif m_meses: dias = int(m_meses.group(1)) * 30
        elif 'BIMESTRAL' in c: dias = 60
        elif 'TRIMESTRAL' in c: dias = 90
        elif 'SEMESTRAL' in c: dias = 180
        elif 'ANUAL' in c: dias = 365
        proxima = data_visita + pd.Timedelta(days=dias)
        return proxima.strftime('%d/%m/%Y')
    except:
        return ""

# --- SIDEBAR ---
with st.sidebar:
    st.markdown("<div style='text-align: center; padding-bottom: 20px;'><h2 style='color: white; letter-spacing: 2px;'>OPERACIONAL</h2></div>", unsafe_allow_html=True)
    st.markdown("---")
    qtd_ajustes = len(st.session_state['lista_pendencias'])
    st.markdown(f"""
        <div style='background: rgba(255,255,255,0.05); padding: 15px; border-radius: 15px; border: 1px solid rgba(255,255,255,0.1);'>
            <p style='margin:0; font-size: 0.9rem;'><span class='led-green'>●</span> SERVIDOR: <b>ESTÁVEL</b></p>
            <p style='margin:0; font-size: 0.9rem;'><span class='led-green'>●</span> CONEXÃO REDE: <b>OK</b></p>
            <p style='margin:0; font-size: 0.9rem;'>AJUSTES LOGADOS: <b>{qtd_ajustes}</b></p>
            <p style='margin-top:10px; font-size: 0.75rem; opacity: 0.7;'>Início: {datetime.now().strftime('%H:%M')}</p>
        </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    if st.button("REINICIAR APP"):
        st.session_state['lista_pendencias'] = []
        st.session_state['movimento_gravado'] = False
        st.session_state['base_deduplicada'] = None
        if 'nome_arquivo_aba3' in st.session_state: del st.session_state['nome_arquivo_aba3']
        if 'abas_planilha' in st.session_state: del st.session_state['abas_planilha']
        st.rerun()
        
# --- HEADER LOGO ---
logo = get_base64('logo.png')
if logo:
    st.markdown(
        f"""
        <div style="text-align:center; padding:20px 0px;">
            <img src="data:image/png;base64,{logo}" width="380" style="filter: drop-shadow(0px 4px 6px rgba(0,0,0,0.1));">
        </div>
        """,
        unsafe_allow_html=True
    )

# --- ABAS OPERACIONAIS ---
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12 = st.tabs([
    "Triagem", "Divisor", "Conciliador", "Rependentes", "Busca Ativa", "Gravador",
    "Confirmação Agenda", "Salva-Vidas", "Central de Limpeza", "Agentes IA", "Solicitados x Atendidos", "Conduta"
])


with tab1:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">TRIAGEM DE CONTATOS</div>""", unsafe_allow_html=True)
    f_tri = st.file_uploader("Suba a Planilha Mestre", type=["xlsx", "csv"], key="tri_file")
    if f_tri:
        df = load_excel_with_ui(f_tri, "triagem")
        if st.button("INICIAR TRIAGEM", type="primary"):
            col = next((c for c in df.columns if 'TELEFONE' in str(c).upper()), df.columns[2])
            df_vazio = df[df[col].isna() | (df[col].astype(str).str.strip() == '') | (df[col].astype(str).str.lower() == 'nan')].copy()
            df_resto = df[~df.index.isin(df_vazio.index)].copy()
            df_c = df_resto[df_resto[col].apply(eh_celular)].copy()
            df_f = df_resto[~df_resto[col].apply(eh_celular)].copy()
            c1, c2, c3 = st.columns(3)
            buf_c = io.BytesIO(); df_c.to_excel(buf_c, index=False); c1.download_button("BAIXAR CELULARES", buf_c.getvalue(), "1_CELULARES.xlsx", type="primary", use_container_width=True)
            buf_f = io.BytesIO(); df_f.to_excel(buf_f, index=False); c2.download_button("BAIXAR FIXOS", buf_f.getvalue(), "2_FIXOS.xlsx", use_container_width=True)
            buf_v = io.BytesIO(); df_vazio.to_excel(buf_v, index=False); c3.download_button("BAIXAR S/ CONTATO", buf_v.getvalue(), "3_EM_BRANCO.xlsx", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

with tab2:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">DIVISOR DE LOTES</div>""", unsafe_allow_html=True)
    c_d1, c_d2 = st.columns([3, 1])
    f_div = c_d1.file_uploader("Planilha de Celulares", type=["xlsx", "csv"], key="div_file")
    tam = c_d2.number_input("Qtd p/ Lote", min_value=1, value=40)
    if f_div:
        df_l = load_excel_with_ui(f_div, "divisor")
        if st.button("GERAR LOTES CSV", type="primary"):
            try:
                lotes = math.ceil(len(df_l)/tam)
                b_zip = io.BytesIO()
                with zipfile.ZipFile(b_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                    for i in range(lotes):
                        fatia = df_l.iloc[i*tam : (i+1)*tam].copy()
                        csv_buf = io.StringIO()
                        fatia.to_csv(csv_buf, index=False, sep=';', quoting=csv.QUOTE_MINIMAL)
                        zf.writestr(f"LOTE_PT{i+1}.csv", csv_buf.getvalue().encode('utf-8-sig'))
                b_zip.seek(0)
                st.success(f"Dividido com sucesso em {lotes} lotes.")
                st.download_button("BAIXAR PACOTE DE LOTES (ZIP)", b_zip.getvalue(), "lotes_hova_csv.zip", mime="application/zip", type="primary")
            except Exception as e:
                st.error(f"Erro ao dividir a planilha: {e}")
    st.markdown('</div>', unsafe_allow_html=True)

with tab3:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">CONCILIADOR DE RESPOSTAS</div>""", unsafe_allow_html=True)
    c_op1, c_op2 = st.columns(2)
    op = c_op1.text_input("NOME DO OPERADOR:", value="ESTER").upper()
    col_destino = c_op2.text_input("NOME DA COLUNA (Ex: MSG2026):", value="MSG2026").upper()
    st.markdown("---")
    c_c1, c_c2 = st.columns(2)
    b_geral = c_c1.file_uploader("Base Geral (Planilha Doctor's)", type=["xlsx"], key="base_geral")
    r_zap = c_c2.file_uploader("Relatórios ZapRocket (CSV)", type=["csv"], accept_multiple_files=True)
    selected_sheet = None
    if b_geral:
        cache_key = f"abas_{b_geral.name}_{b_geral.size}"
        if cache_key not in st.session_state:
            try:
                b_geral.seek(0)
                with zipfile.ZipFile(b_geral, 'r') as z:
                    xml_content = z.read('xl/workbook.xml')
                    root = ET.fromstring(xml_content)
                    ns = {'xmlns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    sheets = root.findall('.//xmlns:sheet', ns)
                    sheet_names = [s.get('name') for s in sheets]
                    st.session_state[cache_key] = sheet_names
            except Exception as e:
                b_geral.seek(0)
                xls = pd.ExcelFile(b_geral)
                st.session_state[cache_key] = xls.sheet_names
        selected_sheet = st.selectbox("Escolha qual Aba (Mês) você quer atualizar:", st.session_state[cache_key])
        b_geral.seek(0)
        df_prev_concil = pd.read_excel(b_geral, sheet_name=selected_sheet, dtype=str)
        st.markdown(f"**Pré-visualização da Aba '{selected_sheet}':**")
        st.dataframe(df_prev_concil.head(3), use_container_width=True)
    if b_geral and r_zap and selected_sheet:
        if st.button("CONCILIAR DADOS", type="primary"):
            with st.spinner(f"Injetando respostas na aba '{selected_sheet}'..."):
                try:
                    b_geral.seek(0)
                    wb = openpyxl.load_workbook(b_geral)
                    ws = wb[selected_sheet]
                    header = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]
                    alvo_limpo = col_destino.replace(" ", "")
                    idx_msg = next((i for i, h in enumerate(header) if alvo_limpo in h.replace(" ", "")), None)
                    idx_tel = next((i for i, h in enumerate(header) if "TELEFONE" in h.replace(" ", "")), None)
                    if idx_msg is None: st.error(f"Não achei a coluna '{col_destino}'.")
                    elif idx_tel is None: st.error(f"Não achei a coluna 'TELEFONE'.")
                    else:
                        idx_msg += 1; idx_tel += 1; s_map = {}
                        for r in r_zap:
                            r.seek(0)
                            texto_csv = r.read().decode('utf-8', errors='ignore')
                            separador = ';' if ';' in texto_csv.splitlines()[0] else ','
                            for ln in texto_csv.splitlines():
                                p = ln.split(separador)
                                if len(p) > 1:
                                    num = limpar_num(p[0])
                                    linha_inteira = str(p).lower()
                                    dt = re.search(r'(\d{2}/\d{2}/\d{4})', linha_inteira)
                                    data_final = dt.group(1) if dt else datetime.now().strftime('%d/%m/%Y')
                                    if "sent" in linha_inteira or "enviado" in linha_inteira:
                                        s_map[num] = f"{data_final} ENVIEI ZAP ROCKET {op}"
                                    else:
                                        s_map[num] = f"SEM WHATS {op} {data_final}"
                        max_col_needed = max(idx_tel, idx_msg)
                        for row_cells in ws.iter_rows(min_row=2, max_col=max_col_needed):
                            cell_tel = row_cells[idx_tel - 1]
                            n_l = limpar_num(str(cell_tel.value))
                            if n_l in s_map:
                                cell_msg = row_cells[idx_msg - 1]
                                m_n = s_map[n_l]
                                m_a = str(cell_msg.value or "")
                                if m_n not in m_a:
                                    cell_msg.value = m_n if m_a in ["","None","nan"] else f"{m_a}\n{m_n}"
                        out_f = io.BytesIO(); wb.save(out_f)
                        st.session_state['base_conciliada'] = out_f.getvalue()
                        st.success(f"Aba '{selected_sheet}' conciliada com sucesso.")
                except Exception as e: st.error(f"Erro inesperado: {e}")
    if 'base_conciliada' in st.session_state:
        st.download_button("BAIXAR PLANILHA CONCILIADA", st.session_state['base_conciliada'], "Base_Final_Hova.xlsx", type="primary")
    st.markdown('</div>', unsafe_allow_html=True)

with tab4:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">FILTRO DE REENVIO</div>""", unsafe_allow_html=True)
    cr1, cr2 = st.columns(2)
    f_ori = cr1.file_uploader("Lote Original", type=["xlsx", "csv"], key="original_lote")
    f_ret = cr2.file_uploader("Retorno Zap", type=["csv"], accept_multiple_files=True, key="retorno_lote")
    if f_ori and f_ret:
        df_o = load_excel_with_ui(f_ori, "reenvio")
        if st.button("FILTRAR PENDENTES", type="primary"):
            try:
                col_t = next((c for c in df_o.columns if 'TELEFONE' in str(c).upper()), df_o.columns[0])
                enviados = set()
                for r in f_ret:
                    r.seek(0)
                    linhas = r.read().decode('utf-8', errors='ignore').splitlines()
                    for ln in linhas:
                        p = ln.split(',')
                        enviados.add(limpar_num(p[0]))
                df_pt2 = df_o[~df_o[col_t].apply(limpar_num).isin(enviados)].copy()
                csv_buf = io.StringIO()
                df_pt2.to_csv(csv_buf, index=False, sep=';', quoting=csv.QUOTE_MINIMAL)
                st.success(f"Filtro aplicado! Restaram {len(df_pt2)} pacientes.")
                st.download_button("BAIXAR LISTA DE REENVIO", csv_buf.getvalue().encode('utf-8-sig'), "REENVIO_FALTANTES.csv", mime="text/csv", type="primary")
            except Exception as e:
                st.error(f"Erro ao processar: {e}")
    st.markdown('</div>', unsafe_allow_html=True)

with tab5:
    st.markdown('<div class="master-card" style="padding: 40px;">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">BUSCA ATIVA E RESPOSTAS RÁPIDAS</div>""", unsafe_allow_html=True)
    col_ia, col_nome = st.columns([1, 2])
    ias_nomes = ["Ester", "Clara", "Iris", "Piter", "Theia", "Lumina", "Aurora", "Verônica", "Glauco"]
    ia_sel = col_ia.selectbox("IA ATIVA:", options=ias_nomes, index=0)
    nome_paciente_global = col_nome.text_input("NOME DO PACIENTE:", placeholder="Digite o nome para preencher todas as mensagens...")
    nome_display = nome_paciente_global.strip().title() if nome_paciente_global else "[NOME DO PACIENTE]"
    st.markdown("---")
    banco = {
        "Aviso de Áudio 🎧": f"Olá, *{nome_display}*! 👁️\nAqui é a *{ia_sel}*, do Hospital de Olhos Vale do Aço.\nVou te enviar um áudio explicativo com informações importantes sobre seu atendimento.\nPor favor, ouça com atenção. 🎧\nQualquer dúvida, estou à disposição. 🤍",
        "Contato Incorreto / Cadastro Antigo 🚫": f"Olá! 👁️\nPedimos desculpas pelo contato anterior. 🙏\nSeu número estava vinculado a um cadastro antigo — já corrigimos para que não aconteça novamente.\nAgradecemos o aviso!\nSe você ou sua família precisarem de cuidado com a visão, entre em contato pelo nosso canal oficial:\n👉 https://wa.me/553138011800\nEstamos à disposição. 🤍\nHospital de Olhos Vale do Aço",
        "Lembrete Indevido (Já Consultou) 😊": f"Olá, *{nome_display}*! 👁️\nPedimos desculpas pelo envio. 🙏\nComo nosso sistema de lembretes é automático, acabou enviando a mensagem mesmo você já tendo consultado recentemente.\nAgradecemos por avisar! Já atualizamos aqui para que você não receba novos lembretes por enquanto. 😊\nEstamos à disposição. 🤍\nHospital de Olhos Vale do Aço",
        "Redirecionamento para Canal Oficial 🤖": f"Olá! Obrigado pelo contato. 👁️\nEste número é apenas para lembretes automáticos e não recebe mensagens.\nPara falar com nossa equipe, clique aqui:\n👉 https://wa.me/553138011800\n\nNossa equipe pode ajudar com:\n✔️ Agendamentos\n✔️ Reagendamentos\n✔️ Dúvidas sobre exames\n✔️ Informações sobre consultas\n✔️ Qualquer outra necessidade\n\nEstamos à disposição! 🤍\nHospital de Olhos Vale do Aço",
        "Óbito (Acolhimento Familiar) 🕊️": f"Olá, 🕊️\nAqui é a *{ia_sel}*, do Hospital de Olhos Vale do Aço.\nRecebemos a notícia do falecimento de *{nome_display}* e queremos expressar nossos sentimentos à família.\nPedimos desculpas pela mensagem automática que foi enviada — não sabíamos o que havia acontecido.\nJá corrigimos nosso cadastro e não enviaremos mais mensagens.\nDesejamos que Deus conforte o coração de toda a família neste momento difícil.\nNossos sentimentos,\n*{ia_sel}*\nHospital de Olhos Vale do Aço\n🕊️ ❤️",
        "Retorno Preventivo 📅": f"Olá, *{nome_display}*! Tudo bem? 👁️\n\nAqui é a *{ia_sel}*, do *Hospital de Olhos Vale do Aço*.\n\nEstava revisando seu histórico e percebi que já faz um tempo desde sua última consulta. A saúde dos seus olhinhos merece atenção regular — está na hora de realizar uma nova avaliação preventiva.\n\n📲 *Para agendar:*\n1. Toque no link azul abaixo\n2. O WhatsApp abre automaticamente\n3. Nossa equipe vai te atender\n\n👇 *Clique aqui para agendar:*\nhttps://wa.me/553138011800\n\n🛑 *Aviso:* Este número só envia lembretes. Para agendar, use o link acima.\n\n💡 Já consultou recentemente? Pode desconsiderar esta mensagem.\n\nPodemos te esperar? 🤍\n\n*{ia_sel} — Hospital de Olhos Vale do Aço*",
        "Pendência de Documentos 📁": f"Olá, *{nome_display}*. Aqui é o(a) {ia_sel}. Notamos que você tem um agendamento, mas ainda constam pendências de exames laboratoriais ou risco cirúrgico em seu prontuário. Por favor, envie uma foto dos resultados por aqui para validarmos sua vaga.",
        "Cancelar ou Remarcar ❌": f"Olá,\n\nPara cancelar ou remarcar, não pode ser por aqui.\n\nToque no link azul abaixo para falar com a equipe:\n👉 https://wa.me/553138011800\n\n(É só tocar que abre sozinho)\n\nA equipe vai te ajudar a cancelar ou remarcar! 🤍\n\n{ia_sel} — Hospital de Olhos Vale do Aço"
    }
    c_msgs, c_log = st.columns([1.8, 1])
    with c_msgs:
        with st.expander("📂 VER TODAS AS MENSAGENS PRONTAS", expanded=True):
            for titulo, texto in banco.items():
                st.markdown(f'<div class="response-box" style="margin-bottom: 10px;"><b>{titulo}</b>', unsafe_allow_html=True)
                st.code(texto, language="text")
                if st.button(f"Salvar no Log", key=f"btn_log_{titulo}"):
                    if nome_paciente_global:
                        st.session_state['lista_pendencias'].append({
                            "Data": datetime.now().strftime("%d/%m"),
                            "Paciente": nome_paciente_global.upper(),
                            "Motivo": titulo,
                            "IA": ia_sel
                        })
                        st.toast(f"Log: {nome_paciente_global.upper()} registrado!")
                    else:
                        st.warning("⚠️ Digite o Nome do Paciente lá em cima para salvar no LOG.")
                st.markdown('</div>', unsafe_allow_html=True)
    with c_log:
        st.markdown("<div style='background:#ffffff; padding:25px; border-radius:25px; border:1px solid #e2e8f0;'>", unsafe_allow_html=True)
        st.markdown("<h4 style='color:#1e3d3a; margin-top:0;'>📝 LOG DA EQUIPE</h4>", unsafe_allow_html=True)
        if st.session_state['lista_pendencias']:
            for it in st.session_state['lista_pendencias']:
                st.markdown(f'<div class="penc-row">{it["Paciente"]}<br><small>{it["Motivo"]} | {it["IA"]} | {it["Data"]}</small></div>', unsafe_allow_html=True)
            df_aj = pd.DataFrame(st.session_state['lista_pendencias']); b_aj = io.BytesIO(); df_aj.to_excel(b_aj, index=False)
            st.download_button("📥 EXPORTAR LOGS (EXCEL)", b_aj.getvalue(), "ajustes_hova.xlsx")
            if st.button("LIMPAR LOGS"): st.session_state['lista_pendencias'] = []; st.rerun()
        else: st.info("Vazio.")
        st.markdown('</div>', unsafe_allow_html=True)

with tab6:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">🤖 APRENDER E REPETIR MOVIMENTO</div>""", unsafe_allow_html=True)
    col_learn, col_apply = st.columns(2)
    with col_learn:
        st.subheader("1. Gravar Movimento")
        f_orig = st.file_uploader("Planilha ANTES", type=["xlsx"], key="orig")
        f_done = st.file_uploader("Planilha DEPOIS", type=["xlsx"], key="done")
        if f_orig and f_done and st.button("🔴 GRAVAR MEU MOVIMENTO"):
            st.session_state['movimento_gravado'] = True
            st.success("✅ Movimento lido! Já aprendi a blindar seus números.")
    with col_apply:
        st.subheader("2. Repetir Sequência")
        f_batch = st.file_uploader("Suba as outras planilhas", type=["xlsx", "html", "htm"], accept_multiple_files=True, key="batch")
        if f_batch and st.session_state['movimento_gravado'] and st.button("🤖 REPETIR MOVIMENTO EM TODAS"):
            b_zip = io.BytesIO()
            with zipfile.ZipFile(b_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for idx, f in enumerate(f_batch):
                    f.seek(0)
                   if f.name.lower().endswith(('.html', '.htm')):
                       conteudo = f.read()
                       if isinstance(conteudo, bytes):
                           conteudo = conteudo.decode('latin-1', errors='replace')
                       df = pd.read_html(io.StringIO(conteudo))[0].astype(str)
                    else:
                        df = pd.read_excel(f, dtype=str)
                    for col in df.columns:
                        df[col] = df[col].apply(lambda x: f'="{x}"' if pd.notnull(x) else x)
                    nome_original = f.name.rsplit('.', 1)[0].upper().replace(" ", "_")
                    csv_buf = io.StringIO(); csv_buf.write("sep=;\n")
                    df.to_csv(csv_buf, index=False, sep=';', quoting=csv.QUOTE_NONE, escapechar="\\", encoding='utf-8-sig')
                    zf.writestr(f"{nome_original}.csv", csv_buf.getvalue())
            b_zip.seek(0)
            st.session_state['zip_tab6'] = b_zip.getvalue()
            st.success("🤖 Sequência repetida com sucesso!")
        if 'zip_tab6' in st.session_state:
            st.download_button("📥 BAIXAR RESULTADO", st.session_state['zip_tab6'], "hova_movimento_concluido.zip", mime="application/zip")
    st.markdown('</div>', unsafe_allow_html=True)

with tab7:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title"> CONFIRMAÇÃO MASTER (MÓDULO CIRÚRGICO ATIVO)</div>""", unsafe_allow_html=True)
    with st.expander("💬 NOVO MODELO ESTRATÉGICO ZAPROCKET (VARIÁVEIS AUTOMÁTICAS)"):
        st.info("💡 **DICA:** Copie o texto abaixo e cole direto no ZapRocket.")
        st.code("""Olá, {{column_4}}! Tudo bem? 👁️

Aqui é a PITER, sua assistente do Hospital de Olhos Vale do Aço.

📌 *Lembrete automático do seu atendimento:*

📋 *Atendimento:* {{column_6}}
👨‍⚕️ *Médico(a):* {{column_7}} 
📅 *Data:* {{column_1}}
⏰ *Chegada:* {{column_2}}

🚨 *PREPARO OBRIGATÓRIO:* {{column_8}}

📁 *O QUE TRAZER:* {{column_9}}

⚠️ *CONFIRMAÇÃO:*

✅ *Confirmar:* Responda *SIM* _(Sistema registra automaticamente)_

❌ *Cancelar/Remarcar:* Não responda aqui  
Temos fila de espera. Clique no link:  
👉 https://wa.me/553138011800

🤖 Mensagem automática. Sistema em atualização.  
Dúvidas? Use o link acima.

🤍 *PITER — Hospital de Olhos Vale do Aço*""", language="text")

    f_agendas = st.file_uploader("Suba as planilhas do Doctors (Consultas, Exames, Cirurgias)", type=["html", "htm", "xls", "xlsx"], accept_multiple_files=True, key="agenda_batch")
    if f_agendas:
        if st.button("🚀 PROCESSAR COM INTELIGÊNCIA CIRÚRGICA"):
            siglas_nao_encontradas = set()
            lista_consolidada = []
            audit_log = []
            for f in f_agendas:
                try:
                    f.seek(0)
                    data_extraida = ""
                    match_data = re.search(r'(\d{2}[-_\.]\d{2}(?:[-_\.]\d{2,4})?)', f.name)
                    if match_data:
                        data_extraida = match_data.group(1).replace('_', '/').replace('.', '/').replace('-', '/')
                    if f.name.lower().endswith(('.html', '.htm')):
                        conteudo = f.read()
                        if isinstance(conteudo, bytes):
                            conteudo = conteudo.decode('latin-1', errors='replace')
                        dfs = pd.read_html(io.StringIO(conteudo), header=0)
                        df_conf = max(dfs, key=len).astype(str)
                    else:
                        df_conf = pd.read_excel(f, dtype=str)
                    df_conf.columns = [str(c).upper().strip() for c in df_conf.columns]
                    if 'PACIENTE' not in df_conf.columns and len(df_conf) > 0:
                        row_str = " ".join([str(x).upper() for x in df_conf.iloc[0].values])
                        if 'PACIENTE' in row_str:
                            df_conf.columns = [str(x).upper().strip() for x in df_conf.iloc[0].values]
                            df_conf = df_conf[1:]
                    if 'PACIENTE' in df_conf.columns:
                        df_conf = df_conf.rename(columns={'PACIENTE': 'NOME'})
                    colunas_obrigatorias = ['NOME', 'TELEFONE', 'HORA']
                    colunas_faltantes = [c for c in colunas_obrigatorias if c not in df_conf.columns]
                    if colunas_faltantes:
                        st.error(f"⚠️ OPA! A planilha '{f.name}' está sem a(s) coluna(s): {', '.join(colunas_faltantes)}.")
                        continue
                    nome_medico_completo = resolver_nome_completo(f.name)
                    if 'NOME' in df_conf.columns:
                        termo_analise = r'(?i)(EM AN[AÁ]LISE|EM ANALISE)'
                        mask_analise_nome = df_conf['NOME'].astype(str).str.contains(termo_analise, na=False, regex=True)
                        col_obs = next((c for c in df_conf.columns if 'OBS' in str(c).upper()), None)
                        mask_analise_obs = pd.Series(False, index=df_conf.index)
                        if col_obs:
                            mask_analise_obs = df_conf[col_obs].astype(str).str.contains(termo_analise, na=False, regex=True)
                        mask_analise_total = mask_analise_nome | mask_analise_obs
                        df_em_analise = df_conf[mask_analise_total]
                        if not df_em_analise.empty:
                            for _, row_analise in df_em_analise.iterrows():
                                nome_analise = str(row_analise['NOME']).strip().upper()
                                st.session_state['lista_pendencias'].append({"Data": datetime.now().strftime("%d/%m"), "Paciente": nome_analise, "Motivo": "⚠️ GUIA EM ANÁLISE", "IA": "Sistema"})
                                audit_log.append({"DATA": data_extraida, "NOME": nome_analise, "TELEFONE": row_analise.get('TELEFONE', ''), "MÉDICO DA AGENDA": nome_medico_completo, "MOTIVO DO CORTE": "⚠️ GUIA EM ANÁLISE"})
                            st.toast(f"⚠️ {len(df_em_analise)} paciente(s) com Guia em Análise interceptado(s)!", icon="🚨")
                        termos_cancel = r'(?i)(CANCELADO|CANCELADA|DESMARCOU|DESMARCADO|DESISTIU|FALTOU)'
                        mask_cancel_nome = df_conf['NOME'].astype(str).str.contains(termos_cancel, na=False, regex=True)
                        df_cancelados = df_conf[mask_cancel_nome]
                        for _, r_canc in df_cancelados.iterrows():
                            audit_log.append({"DATA": data_extraida, "NOME": str(r_canc['NOME']).strip().upper(), "TELEFONE": r_canc.get('TELEFONE', ''), "MÉDICO DA AGENDA": nome_medico_completo, "MOTIVO DO CORTE": "❌ CANCELADO/FALTOU NO DOCTOR'S"})
                        df_conf = df_conf[~(mask_analise_total | mask_cancel_nome)]
                        df_conf = df_conf.dropna(subset=['NOME'])
                        df_conf['OBS_EXTRAIDA'] = df_conf['NOME'].astype(str)
                        padrao_corte = r'(?i)(#|-|,|/|\(|MOTIVO|IDADE|CPF|\bTEM\b|\bCONFIRMAD[OA]S?\b|\bGUIA\b|\bSALV[OA]\b|\bREMARCAD[OA]\b|\bAV\b|\bOD\b|\bOE\b|\bAO\b|\bJUSTIFICATIVA\b|\bPEGAR\b|\bPEDIDO\b|\bDOC\b|\bPACIENTE\b|\bSOLICITADO\b|\bJUNT[OA] COM\b|Nº|N°|\d|\?|<|>|\bJ[AÁ] EST[AÁ]\b|\bEST[AÁ] PAGO\b|\bAPRESENTAR\b|\bA MESMA\b|\bESPOSA\b|\bMARIDO\b|\bPAI\b|\bM[AÃ]E\b|\bFILH[OA]S?\b|\bIRM[AÃ]O?\b|\bNET[OA]S?\b|\bTIO\b|\bTIA\b|\bAV[OÓÔ]\b|\bSOGRA\b|\bSOGRO\b|\bACOMPANHANTE\b|\bLIGAÇ[AÃ]O\b|\bATEND\.?|\bTBM\b|\bTAMB[EÉ]M\b|\bRECADO\b|\bREALIZA\b|\bHEMODI[AÁ]LISE\b|\bESTAVAM?\b|\bCADASTRO\b|\bRELATA\b|\bVEIO\b|\bJUSTIFICAR\b|\bC/\b|\bCOM\b|\bOU\b|\bOBS\b|\bDRA?\b)'
                        df_conf['NOME'] = df_conf['NOME'].apply(lambda x: re.split(padrao_corte, str(x).replace('*', ''))[0].strip() if pd.notnull(x) else "")
                    if 'TELEFONE' in df_conf.columns:
                        df_conf = df_conf[df_conf['TELEFONE'].str.strip() != '-']
                    col_tipo = next((c for c in df_conf.columns if 'TIPO' in c), None)
                    df_conf['TIPO_RAW'] = df_conf[col_tipo] if col_tipo else 'C'
                    for index, row in df_conf.iterrows():
                        tipo_raw = str(row.get('TIPO_RAW', '')).strip().upper()
                        traducao = traduzir_tipo(tipo_raw, st.session_state['dicionario_tipos'])
                        if traducao == "DESCONHECIDO" and tipo_raw != "":
                            siglas_nao_encontradas.add(tipo_raw)
                        df_conf.at[index, 'TIPO_FINAL'] = traducao
                    df_conf['MEDICO_ORIGINAL'] = nome_medico_completo
                    df_conf['DATA'] = data_extraida
                    lista_consolidada.append(df_conf)
                except Exception as e:
                    st.error(f"Erro no arquivo {f.name}: {e}")

            if siglas_nao_encontradas:
                st.error(f"🛑 ATENÇÃO! Siglas não encontradas no Dicionário: **{', '.join(siglas_nao_encontradas)}**")
                st.warning("👉 Adicione essas siglas no Dicionário e clique em Processar de novo.")
            elif lista_consolidada:
                df_full = pd.concat(lista_consolidada, ignore_index=True)
                df_full['NOME'] = df_full['NOME'].str.strip().str.upper()
                df_full['TELEFONE_LIMPO'] = df_full['TELEFONE'].apply(limpar_num)
                mask_bad_phone = (df_full['TELEFONE_LIMPO'] == '') | df_full['TELEFONE_LIMPO'].apply(eh_numero_falso)
                df_bad_phone = df_full[mask_bad_phone]
                for _, r_bad in df_bad_phone.iterrows():
                    audit_log.append({"DATA": r_bad.get('DATA', ''), "NOME": r_bad['NOME'], "TELEFONE": r_bad.get('TELEFONE', ''), "MÉDICO DA AGENDA": r_bad.get('MEDICO_ORIGINAL', ''), "MOTIVO DO CORTE": "🚫 SEM TELEFONE OU NÚMERO FALSO"})
                df_full = df_full[~mask_bad_phone]

                # ── FILTRO HORÁRIO CRÍTICO: 18:00 → auditoria ──────────
                mask_1800 = df_full['HORA'].astype(str).str.strip() == '18:00'
                df_1800 = df_full[mask_1800]
                for _, r_18 in df_1800.iterrows():
                    audit_log.append({"DATA": r_18.get('DATA', ''), "NOME": r_18['NOME'], "TELEFONE": r_18.get('TELEFONE', ''), "MÉDICO DA AGENDA": r_18.get('MEDICO_ORIGINAL', ''), "MOTIVO DO CORTE": "🚫 HORÁRIO DE CANCELAMENTO (18:00)"})
                df_full = df_full[~mask_1800]
                if mask_1800.sum() > 0:
                    st.toast(f"🚫 {mask_1800.sum()} paciente(s) com horário 18:00 removido(s) da confirmação.", icon="⏰")
                mask_dupes = df_full.duplicated(subset=['TELEFONE_LIMPO', 'NOME', 'DATA', 'HORA', 'TIPO_FINAL', 'MEDICO_ORIGINAL'], keep='first')
                df_dupes = df_full[mask_dupes]
                for _, r_dup in df_dupes.iterrows():
                    audit_log.append({"DATA": r_dup.get('DATA', ''), "NOME": r_dup['NOME'], "TELEFONE": r_dup.get('TELEFONE', ''), "MÉDICO DA AGENDA": r_dup.get('MEDICO_ORIGINAL', ''), "MOTIVO DO CORTE": "♻️ DUPLICIDADE EXATA"})
                df_full = df_full[~mask_dupes]
                resultados_finais = []
                if 'HORA' in df_full.columns and 'NOME' in df_full.columns:
                    for (telefone, nome_pct, data_pct), group in df_full.groupby(['TELEFONE_LIMPO', 'NOME', 'DATA']):
                        group['HORA_TEMP'] = pd.to_datetime(group['HORA'], format='%H:%M', errors='coerce')
                        group = group.sort_values(by='HORA_TEMP')
                        menor_hora_do_dia = group.iloc[0]['HORA']
                        group_medicos_validos = group[group['MEDICO_ORIGINAL'].isin(agendas_enviar)]
                        if group_medicos_validos.empty:
                            for _, r_med in group.iterrows():
                                audit_log.append({"DATA": data_pct, "NOME": nome_pct, "TELEFONE": r_med.get('TELEFONE',''), "MÉDICO DA AGENDA": r_med.get('MEDICO_ORIGINAL',''), "MOTIVO DO CORTE": "👨‍⚕️ MÉDICO/SETOR NÃO HABILITADO"})
                            continue
                        f_acompanhante_grupo1 = False; f_acompanhante_grupo2 = False
                        f_jejum_absoluto = False; f_jejum_observacao = False; f_isento_jejum = False
                        f_colirio_24h = False; f_colirio_48h = False; f_colirio_mydriacyl_1h = False
                        f_traz_colirio = False; f_lente_24h = False; f_lente_72h = False; f_lente_7d = False
                        f_traz_oculos = False; f_traz_estojo = False; f_banho_refra = False
                        f_duracao_agf = False; f_duracao_cdpo = False; f_docs_cirurgia = False
                        f_is_ic = False; f_ic_com_exames = False; f_ilio_oculto = False
                        f_pre_op_catarata = False; f_teste_ortoptico = False
                        tem_valido = False; tem_cra = False; tem_cirurgia = False
                        tipo_cirurgia = "GERAL"; medico_cirurgia = ""; tipos_para_nome = set()
                        col_obs_nome = 'OBSERVAÇÃO' if 'OBSERVAÇÃO' in group.columns else ('OBS' if 'OBS' in group.columns else None)
                        observacao_do_dia = " ".join(group[col_obs_nome].astype(str)).upper() if col_obs_nome else " "
                        for _, row in group_medicos_validos.iterrows():
                            tipo_original = str(row['TIPO_RAW']).upper().replace('_', ' ')
                            tipo_final = str(row['TIPO_FINAL'])
                            medico = str(row['MEDICO_ORIGINAL']).upper()
                            obs_isolada = str(row.get('OBSERVAÇÃO', row.get('OBSERVACAO', row.get('OBS', '')))).upper()
                            obs_escondida = str(row.get('OBS_EXTRAIDA', '')).upper()
                            if medico == 'CENTRO CIRÚRGICO' or 'CIRURGIA' in tipo_original or 'CI ' in tipo_original or 'CI_' in tipo_original or medico == 'ENFERMAGEM':
                                tipo_raw = f"{tipo_original} {obs_isolada} {obs_escondida}"
                            else:
                                tipo_raw = tipo_original
                            if medico == 'CATARATA' or 'PRED P' in tipo_raw:
                                f_pre_op_catarata = True
                            partes = tipo_raw.replace('/', ' ').split()
                            if tipo_raw == 'ILIO':
                                f_ilio_oculto = True
                                continue
                            is_ic_current_row = False
                            if medico == 'ENFERMAGEM' or tipo_raw == 'IC' or tipo_raw.startswith(('IC ', 'IC_', 'IC.')):
                                f_is_ic = True
                                is_ic_current_row = True
                                if any(x in tipo_raw or x in observacao_do_dia for x in ['ANTIVEG', 'ANTIVEGF', 'TTO', 'INTRAVITREA']):
                                    f_isento_jejum = True; tipo_cirurgia = 'INTRAVITREA'
                            traducoes_fantasmas = ["Retorno", "Consulta Reavaliação Anual", "Entrega de Exames", "1ª Avaliação", "2ª Avaliação", "3ª Avaliação", "Avaliação de Glaucoma"]
                            is_fantasma = tipo_final in traducoes_fantasmas
                            if not is_fantasma:
                                tem_valido = True
                                tipos_para_nome.add(tipo_final)
                            if tipo_final == "Consulta Reavaliação Anual" or "Retorno" in tipo_final:
                                tem_cra = True
                            if not is_ic_current_row:
                                if any(p in ['AGF', 'RF', 'FOTO', 'YAG'] for p in partes): f_acompanhante_grupo1 = True
                                if any(p in ['MR', 'RD', 'OCT', 'ANGIO-OCT'] for p in partes) or any(p in exames_dilatam for p in partes):
                                    f_acompanhante_grupo2 = True; f_traz_estojo = True
                                if any(p in ['TOPO', 'PTC', 'PAQ'] for p in partes): f_lente_72h = True
                                if any(p in ['AV', 'CVC'] for p in partes): f_traz_oculos = True
                                if any(p in ['AV', 'MR'] for p in partes): f_traz_estojo = True
                                if 'AGF' in partes: f_duracao_agf = True
                                if 'CDPO' in partes: f_duracao_cdpo = True
                                if 'RC' in partes: f_traz_colirio = True
                                if 'TESTEORTOP' in tipo_raw or 'TESTE ORTOP' in tipo_raw: f_teste_ortoptico = True
                                if 'CIRURGIA' in tipo_original or 'CI ' in tipo_original or 'CI_' in tipo_original or 'CI.' in tipo_original or medico == 'CENTRO CIRÚRGICO':
                                    tem_cirurgia = True; f_acompanhante_grupo1 = True; f_docs_cirurgia = True
                                    if "Cirurgia Dr(a)." in tipo_final: medico_cirurgia = tipo_final.replace("Cirurgia ", "").strip()
                                    else: medico_cirurgia = medico
                                    if 'FACO' in tipo_raw or 'CATARATA' in tipo_raw:
                                        tipo_cirurgia = 'FACO'; f_colirio_24h = True
                                        if "JEJUM" in observacao_do_dia: f_jejum_observacao = True
                                    elif 'VITRE' in tipo_raw or 'RETINA' in tipo_raw or 'VVPP' in tipo_raw:
                                        tipo_cirurgia = 'RETINA'; f_jejum_absoluto = True; f_colirio_mydriacyl_1h = True
                                    elif 'CICLOFOTO' in tipo_raw:
                                        tipo_cirurgia = 'CICLOFOTO'; f_jejum_absoluto = True
                                    elif 'TRAB' in tipo_raw or 'GLAUCOMA' in tipo_raw:
                                        tipo_cirurgia = 'GLAUCOMA'; f_jejum_absoluto = True
                                    elif 'ESTRABISMO' in tipo_raw:
                                        tipo_cirurgia = 'ESTRABISMO'; f_jejum_absoluto = True
                                    elif 'ANTIVEG' in tipo_raw or 'ANTIVGF' in tipo_raw or 'INTRAVITREA' in tipo_raw:
                                        tipo_cirurgia = 'INTRAVITREA'; f_isento_jejum = True
                                    elif 'ANEL' in tipo_raw:
                                        tipo_cirurgia = 'ANEL'; f_colirio_48h = True; f_lente_72h = True
                                        if "JEJUM" in observacao_do_dia: f_jejum_observacao = True
                                    elif 'PRK' in tipo_raw or 'LASIK' in tipo_raw or 'REFRATIVA' in tipo_raw:
                                        tipo_cirurgia = 'REFRATIVA'; f_banho_refra = True; f_lente_7d = True
                                        if "JEJUM" in observacao_do_dia: f_jejum_observacao = True
                                    elif 'TRANSPLANTE' in tipo_raw or 'CORNEA' in tipo_raw:
                                        tipo_cirurgia = 'TRANSPLANTE'; f_jejum_absoluto = True; f_lente_24h = True
                                    elif 'PTERIGIO' in tipo_raw or 'CALAZIO' in tipo_raw or 'TUMOR' in tipo_raw or 'LP' in tipo_raw or 'TTO' in tipo_raw:
                                        tipo_cirurgia = 'SUPERFICIAL'
                                        if "JEJUM" in observacao_do_dia: f_jejum_observacao = True
                                    if tipo_cirurgia == 'GERAL':
                                        if 'MARILUCE' in medico_cirurgia.upper():
                                            tipo_cirurgia = 'TRANSPLANTE'; f_jejum_absoluto = True; f_lente_24h = True
                                        else:
                                            if "JEJUM" in observacao_do_dia: f_jejum_observacao = True
                                    if 'GABRIEL L' in medico_cirurgia.upper():
                                        if not f_isento_jejum and tipo_cirurgia != 'INTRAVITREA':
                                            f_jejum_absoluto = True; f_jejum_observacao = False
                        if not tem_valido:
                            motivo_corte = "👻 PROCEDIMENTO ISOLADO (Apenas ILIO)." if f_ilio_oculto else "👻 PROCEDIMENTO INVISÍVEL (Ex: Apenas Retorno ou CAT)"
                            audit_log.append({"DATA": data_pct, "NOME": nome_pct, "TELEFONE": telefone, "MÉDICO DA AGENDA": group_medicos_validos.iloc[0]['MEDICO_ORIGINAL'], "MOTIVO DO CORTE": motivo_corte})
                            continue
                        preparo_lista = []
                        if f_teste_ortoptico and not preparo_lista:
                            preparo_lista.append("✔️ Para crianças, a presença dos pais ou responsável é obrigatória.")
                            preparo_lista.append("✔️ Chegue com 15 minutos de antecedência.")
                        if f_pre_op_catarata: categoria_lote = "4_PRE_OP_CATARATA"
                        elif f_is_ic: categoria_lote = "1_ENFERMAGEM_E_IC"
                        elif tem_cirurgia: categoria_lote = "1_CIRURGIAS"
                        elif f_acompanhante_grupo2 or f_duracao_agf: categoria_lote = "2_EXAMES_QUE_DILATAM"
                        else: categoria_lote = "3_CONSULTAS_NORMAIS"
                        if f_is_ic:
                            preparo_lista.append("⚠️ ATENÇÃO: O paciente não precisa comparecer presencialmente para esta instrução, pode ser algum familiar ou responsável pelo mesmo.")
                            preparo_lista.append("💰 SOBRE O PAGAMENTO: Se o seu atendimento for PARTICULAR, CONVÊNIO DE PAGAMENTO À VISTA ou NÃO AUTORIZADO PELO PLANO, o pagamento deverá ser feito obrigatoriamente nesta data de atendimento.")
                        if f_isento_jejum:
                            preparo_lista.append("✔️ Alimentação: NÃO é necessário fazer jejum. Pode se alimentar normalmente.")
                        elif f_duracao_agf:
                            preparo_lista.append("✔️ Alimentação: Não há necessidade de jejum absoluto. É recomendável apenas uma alimentação mais leve 2h antes do exame.")
                        elif (f_jejum_absoluto or f_jejum_observacao):
                            preparo_lista.append("✔️ Jejum Absoluto: É necessário fazer jejum de 08 horas. Não coma nem beba nada (incluindo água) nas 8 horas antes do procedimento.")
                        if f_duracao_agf:
                            preparo_lista.append("✔️ Medicamentos: Pacientes em uso de medicamentos para pressão alta devem tomar a medicação normalmente.")
                            preparo_lista.append("✔️ Cuidados (Diabéticos): Se você é diabético, lembre-se de trazer seu lanche para não ocorrer hipoglicemia.")
                            preparo_lista.append("✔️ Duração: Exame demorado. Venha com disponibilidade de horário.")
                            preparo_lista.append("✔️ Acompanhante: Presença OBRIGATÓRIA de acompanhante adulto.")
                        else:
                            if f_acompanhante_grupo1:
                                preparo_lista.append("✔️ Acompanhante: Presença OBRIGATÓRIA de 1 acompanhante adulto. Você não poderá dirigir após o procedimento.")
                            elif f_acompanhante_grupo2:
                                preparo_lista.append("✔️ Dilatação da Pupila: Seu atendimento exige a dilatação da pupila, o que causará embaçamento visual temporário.")
                                preparo_lista.append("✔️ Acompanhante: A presença de acompanhante é OPCIONAL caso você se sinta inseguro para retornar. ATENÇÃO: Para pacientes menores de 18 anos, o acompanhante é OBRIGATÓRIO.")
                        if f_colirio_24h: preparo_lista.append("✔️ Colírio: Iniciar o uso do colírio prescrito 1 dia antes da cirurgia.")
                        if f_colirio_48h: preparo_lista.append("✔️ Colírio: Iniciar o uso 48h antes da cirurgia, conforme a receita.")
                        if f_colirio_mydriacyl_1h: preparo_lista.append("✔️ Colírio: Iniciar o uso do colírio Mydriacyl 1 hora antes da cirurgia.")
                        if f_traz_colirio: preparo_lista.append("✔️ Atenção: Por favor, traga todos os colírios que você está usando no tratamento.")
                        if f_lente_24h: preparo_lista.append("✔️ Lentes: Usuários de Lentes de Contato devem suspender o uso 1 dia (24h) antes do procedimento.")
                        if f_lente_72h: preparo_lista.append("✔️ Lentes: Usuários de Lentes de Contato devem suspender o uso 72h (3 dias) antes do procedimento.")
                        if f_lente_7d: preparo_lista.append("✔️ Lentes: Usuários de Lentes de Contato devem suspender o uso 7 dias antes do procedimento.")
                        if f_traz_oculos: preparo_lista.append("✔️ Óculos: É OBRIGATÓRIO trazer os óculos e a receita médica no dia do exame.")
                        if f_traz_estojo: preparo_lista.append("✔️ Lentes: Se você usa Lentes de Contato, traga o seu estojo para retirar as lentes aqui no hospital.")
                        if f_banho_refra: preparo_lista.append("✔️ Higiene: Tomar banho e lavar muito bem a cabeça e o rosto. Não usar maquiagem, cremes, perfume ou produtos no cabelo.")
                        if f_duracao_cdpo:
                            preparo_lista.append("✔️ Atenção à Duração: Você ficará no hospital o dia todo. Medições da pressão ocular às 08:00, 11:00 e 13:30.")
                            preparo_lista.append("✔️ Preparo: A Curva de Pressão não exige dilatação da pupila.")
                        if tem_cirurgia and not f_banho_refra:
                            preparo_lista.append("✔️ Cuidados: Não use maquiagem ou cremes no rosto. Deixe objetos de valor em casa. Vista roupas confortáveis.")
                        if f_ilio_oculto:
                            preparo_lista.append("✔️ Atenção: Siga rigorosamente as orientações de preparo repassadas pela nossa equipe no momento do agendamento.")
                        if not preparo_lista:
                            preparo_lista.append("✔️ Preparo: A dilatação da pupila será realizada conforme necessidade e protocolo médico.")
                            preparo_lista.append("✔️ Chegada: Por favor, chegue com 15 minutos de antecedência para a realização do seu cadastro.")
                        preparo = "\n".join(preparo_lista)
                        nome_paciente = str(group_medicos_validos.iloc[0]['NOME']).upper()
                        tel_valido = group_medicos_validos.iloc[0]['TELEFONE']
                        data_final = group_medicos_validos.iloc[0].get('DATA', '')
                        medico_base = group_medicos_validos.iloc[-1]['MEDICO_ORIGINAL'] if tem_cirurgia else group_medicos_validos.iloc[0]['MEDICO_ORIGINAL']
                        if medico_base not in ['Centro Cirúrgico', 'Setor de Lentes de Contato', 'Exames', 'Glaucoma', 'Enfermagem']:
                            medico_formatado = f"Dr(a). {medico_base}"
                        else:
                            medico_formatado = medico_base
                        lista_nomes_reais = sorted(list(tipos_para_nome))
                        tipo_final_saida = lista_nomes_reais[0] if lista_nomes_reais else "Atendimento Especializado"
                        nomes_cirurgia = {'FACO': 'Cirurgia de Catarata (FACO)', 'ANEL': 'Implante de Anel', 'RETINA': 'Cirurgia de Retina', 'INTRAVITREA': 'Aplicação Intravítrea', 'REFRATIVA': 'Cirurgia Refrativa', 'TRANSPLANTE': 'Transplante de Córnea', 'SUPERFICIAL': 'Procedimento Cirúrgico (Pterígio/Calázio/Tumor)', 'GLAUCOMA': 'Cirurgia de Glaucoma', 'CICLOFOTO': 'Ciclofotocoagulação', 'ESTRABISMO': 'Cirurgia de Estrabismo', 'GERAL': 'Procedimento Cirúrgico'}

                        # ✅ IC sempre aparece como Instrução — nunca confunde com cirurgia
                        if f_is_ic:
                            if tipo_cirurgia == 'INTRAVITREA':
                                tipo_final_saida = "Instrução — Aplicação Intravítrea"
                            elif tipo_cirurgia == 'FACO':
                                tipo_final_saida = "Instrução — Cirurgia de Catarata"
                            elif tipo_cirurgia == 'RETINA':
                                tipo_final_saida = "Instrução — Cirurgia de Retina"
                            elif tipo_cirurgia == 'GLAUCOMA':
                                tipo_final_saida = "Instrução — Cirurgia de Glaucoma"
                            elif tipo_cirurgia == 'CICLOFOTO':
                                tipo_final_saida = "Instrução — Ciclofotocoagulação"
                            elif tipo_cirurgia == 'SUPERFICIAL':
                                tipo_final_saida = "Instrução — Pterígio / Calázio"
                                tipo_final_saida = "Instrução — Pterígio / Calázio"
                            elif tipo_cirurgia == 'REFRATIVA':
                                tipo_final_saida = "Instrução — Cirurgia Refrativa"
                            elif tipo_cirurgia == 'TRANSPLANTE':
                                tipo_final_saida = "Instrução — Transplante de Córnea"
                            else:
                                tipo_final_saida = "Instrução Cirúrgica (Pré-Operatório)"

                        if not f_is_ic:  # IC já teve seu tipo definido acima — não sobrescrever
                            if len(group_medicos_validos) > 1:
                                if tem_cirurgia:
                                    base_nome = nomes_cirurgia.get(tipo_cirurgia, 'Procedimento Cirúrgico')
                                    tipo_final_saida = "Exames Prévios + " + base_nome
                                elif tem_cra:
                                    outros_tipos = [t for t in lista_nomes_reais if "Reavaliação" not in t and "Retorno" not in t]
                                    tipo_final_saida = "Reavaliação Anual + " + " + ".join(outros_tipos) if outros_tipos else "Reavaliação Anual + Exames"
                                else:
                                    tipo_final_saida = " + ".join(lista_nomes_reais)
                            elif tem_cirurgia:
                                tipo_final_saida = nomes_cirurgia.get(tipo_cirurgia, 'Procedimento Cirúrgico')
                        dr_escondido = None
                        if tem_cirurgia:
                            for t in tipos_para_nome:
                                if "Cirurgia Dr(a)." in str(t):
                                    dr_escondido = str(t).replace("Cirurgia ", "").strip()
                                    break
                        if tem_cirurgia and dr_escondido:
                            medico_formatado = dr_escondido
                        docs = ["✔️ Documento de Identidade com foto (RG, CNH ou outro).", "✔️ Carteirinha do convênio (física ou aplicativo no celular)."]
                        if f_is_ic:
                            if f_ic_com_exames:
                                docs.append("✔️ Resultados de Exames de Laboratório, Eletrocardiograma (ECG) e Laudo de Risco Cirúrgico.")
                        elif tem_cirurgia:
                            docs.append("✔️ Pedido Médico.")
                            if f_docs_cirurgia:
                                docs.append("✔️ Termo de Consentimento assinado por extenso em TODAS as folhas.")
                        else:
                            docs.append("✔️ Pedido Médico (se houver).")
                        documentos = "\n".join(docs)
                        if f_pre_op_catarata:
                            tipo_final_saida = "Pré-Operatório para Cirurgia de Catarata"
                            preparo = "⏱️ IMPORTANTE: Venha com disponibilidade de tempo — o pré-operatório costuma ser demorado."
                            documentos = "✔️ Cartão do SUS\n✔️ Documento de Identificação com foto"
                            medico_formatado = "Equipe de Catarata"
                        resultados_finais.append({'CATEGORIA': categoria_lote, 'DATA': data_final, 'HORA': menor_hora_do_dia, 'NOME': nome_paciente, 'NOME_CURTO': pegar_nome_curto(nome_paciente), 'TELEFONE': tel_valido, 'TIPO': tipo_final_saida, 'MEDICO': medico_formatado, 'PREPARO': preparo, 'DOCUMENTOS': documentos})
                df_limpo = pd.DataFrame(resultados_finais)
                if audit_log:
                    df_audit = pd.DataFrame(audit_log)
                    b_audit = io.BytesIO()
                    with pd.ExcelWriter(b_audit, engine='openpyxl') as writer:
                        df_audit.to_excel(writer, index=False, sheet_name='Pacientes Cortados')
                        ws_audit = writer.sheets['Pacientes Cortados']
                        for col in ws_audit.columns:
                            ws_audit.column_dimensions[col[0].column_letter].width = 30
                    st.session_state['audit_report'] = b_audit.getvalue()
                    st.session_state['audit_count'] = len(df_audit)
                if not df_limpo.empty:
                    for col in df_limpo.columns:
                        df_limpo[col] = df_limpo[col].astype(str).replace(r'[;\n]', ' ', regex=True).str.strip()
                        df_limpo[col] = df_limpo[col].apply(lambda x: "" if x.lower() == "nan" else x)
                    df_limpo['PREPARO'] = df_limpo['PREPARO'].str.replace('✔️', '\n✔️').str.strip()
                    df_limpo['DOCUMENTOS'] = df_limpo['DOCUMENTOS'].str.replace('✔️', '\n✔️').str.strip()
                    b_zip = io.BytesIO()
                    with zipfile.ZipFile(b_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                        categorias = df_limpo['CATEGORIA'].unique()
                        for cat in sorted(categorias):
                            df_cat = df_limpo[df_limpo['CATEGORIA'] == cat].drop(columns=['CATEGORIA'])
                            qtd_lotes = math.ceil(len(df_cat) / 45)
                            for i in range(qtd_lotes):
                                df_lote = df_cat.iloc[i * 45 : (i + 1) * 45].copy()
                                if not df_lote.empty:
                                    df_lote = pd.concat([df_lote.iloc[[0]], df_lote], ignore_index=True)
                                sufixo = f"_PT{i+1}" if qtd_lotes > 1 else ""
                                csv_agenda = io.StringIO()
                                df_lote.to_csv(csv_agenda, index=False, sep=';', quoting=csv.QUOTE_MINIMAL)
                                zf.writestr(f"{cat}{sufixo}.csv", csv_agenda.getvalue().encode('utf-8-sig'))
                    b_zip.seek(0)
                    st.session_state['zip_tab7'] = b_zip.getvalue()
                    st.success(f"✅ Master System Atualizado! Todas as Regras e Blindagens Operando com Sucesso.")
                else:
                    st.warning("⚠️ Não sobrou nenhum paciente válido após aplicar os filtros.")
        col_dw1, col_dw2 = st.columns(2)
        if 'zip_tab7' in st.session_state:
            col_dw1.download_button("📥 1. BAIXAR PACOTES (ZAPROCKET)", st.session_state['zip_tab7'], "agendas_hova_master.zip", mime="application/zip", use_container_width=True)
        if 'audit_report' in st.session_state:
            col_dw2.download_button(f"🚨 2. VER {st.session_state['audit_count']} PACIENTES CORTADOS", st.session_state['audit_report'], "Relatorio_Auditoria_Cortados.xlsx", type="primary", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

with tab8:
    st.markdown('<div class="master-card" style="border: 2px solid #ff4b4b;">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title" style="color: #ff4b4b;"> SALVA-VIDAS (CORRETOR DE ERROS)</div>""", unsafe_allow_html=True)
    st.markdown("Use esta aba para comparar as planilhas que você **já enviou (incompletas)** com a planilha **correta (gerada na Aba 7)**.")
    col_err1, col_err2 = st.columns(2)
    f_errada = col_err1.file_uploader("1. Planilhas INCOMPLETAS", type=["csv"], accept_multiple_files=True, key="salva_errada")
    f_correta = col_err2.file_uploader("2. Planilha CORRETA", type=["csv"], accept_multiple_files=True, key="salva_correta")
    if f_errada and f_correta:
        if st.button("🚨 CRUZAR DADOS E DESCOBRIR ERROS", type="primary"):
            try:
                lista_err = []
                for f in f_errada:
                    f.seek(0)
                    df_t = pd.read_csv(f, sep=';', dtype=str)
                    if len(df_t.columns) < 2: f.seek(0); df_t = pd.read_csv(f, sep=',', dtype=str)
                    lista_err.append(df_t)
                df_env = pd.concat(lista_err, ignore_index=True)
                df_env.columns = [str(c).upper().strip() for c in df_env.columns]
                lista_cor = []
                for f in f_correta:
                    f.seek(0)
                    df_c = pd.read_csv(f, sep=';', dtype=str)
                    if len(df_c.columns) < 2: f.seek(0); df_c = pd.read_csv(f, sep=',', dtype=str)
                    lista_cor.append(df_c)
                df_cor = pd.concat(lista_cor, ignore_index=True)
                df_cor.columns = [str(c).upper().strip() for c in df_cor.columns]
                if 'NOME' not in df_env.columns or 'NOME' not in df_cor.columns:
                    st.error("As planilhas precisam ter a coluna NOME para o cruzamento funcionar.")
                else:
                    df_env['NOME_CHAVE'] = df_env['NOME'].astype(str).str.strip().str.upper()
                    df_cor['NOME_CHAVE'] = df_cor['NOME'].astype(str).str.strip().str.upper()
                    nomes_enviados = set(df_env['NOME_CHAVE'].dropna().tolist())
                    df_faltaram = df_cor[~df_cor['NOME_CHAVE'].isin(nomes_enviados)].copy()
                    if 'NOME_CHAVE' in df_faltaram.columns: df_faltaram = df_faltaram.drop(columns=['NOME_CHAVE'])
                    if 'HORA' in df_cor.columns and 'HORA' in df_env.columns:
                        df_intersection = pd.merge(df_cor, df_env[['NOME_CHAVE', 'HORA']], on='NOME_CHAVE', suffixes=('_CORRETA', '_ENVIADA'))
                        df_mudaram = df_intersection[df_intersection['HORA_CORRETA'] != df_intersection['HORA_ENVIADA']].copy()
                        df_mudaram = df_mudaram.rename(columns={'HORA_CORRETA': 'HORA'})
                        cols_to_keep = [c for c in df_cor.columns if c != 'NOME_CHAVE']
                        df_mudaram = df_mudaram[cols_to_keep]
                    else:
                        df_mudaram = pd.DataFrame()
                    b_zip_salva = io.BytesIO()
                    with zipfile.ZipFile(b_zip_salva, "w", zipfile.ZIP_DEFLATED) as zf:
                        if not df_faltaram.empty:
                            qtd_lotes_f = math.ceil(len(df_faltaram) / 45)
                            for i in range(qtd_lotes_f):
                                df_lote_f = df_faltaram.iloc[i * 45 : (i + 1) * 45].copy()
                                sufixo = f"_PT{i+1}" if qtd_lotes_f > 1 else ""
                                csv_f = io.StringIO()
                                df_lote_f.to_csv(csv_f, index=False, sep=';', quoting=csv.QUOTE_MINIMAL)
                                zf.writestr(f"1_FALTARAM_ENVIAR{sufixo}.csv", csv_f.getvalue().encode('utf-8-sig'))
                        if not df_mudaram.empty:
                            qtd_lotes_m = math.ceil(len(df_mudaram) / 45)
                            for i in range(qtd_lotes_m):
                                df_lote_m = df_mudaram.iloc[i * 45 : (i + 1) * 45].copy()
                                sufixo = f"_PT{i+1}" if qtd_lotes_m > 1 else ""
                                csv_m = io.StringIO()
                                df_lote_m.to_csv(csv_m, index=False, sep=';', quoting=csv.QUOTE_MINIMAL)
                                zf.writestr(f"2_CORRECAO_DE_HORARIO{sufixo}.csv", csv_m.getvalue().encode('utf-8-sig'))
                    b_zip_salva.seek(0)
                    st.success("✅ Varredura concluída com sucesso!")
                    st.warning(f"**Identificamos:**\n- {len(df_faltaram)} pacientes que NÃO receberam mensagem.\n- {len(df_mudaram)} pacientes que receberam o HORÁRIO ERRADO.")
                    st.session_state['base_salva_vidas'] = b_zip_salva.getvalue()
            except Exception as e:
                st.error(f"Erro ao cruzar as planilhas: {e}")
    if 'base_salva_vidas' in st.session_state:
        st.download_button("🚑 BAIXAR PLANILHAS DE CORREÇÃO", st.session_state['base_salva_vidas'], "HOVA_SALVA_VIDAS.zip", mime="application/zip", type="primary")
    st.markdown('</div>', unsafe_allow_html=True)

with tab9:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title"> CENTRAL DE LIMPEZA E UNIFICAÇÃO</div>""", unsafe_allow_html=True)
    st.markdown("### ⚙️ ESCOLHA O MOTOR DE PROCESSAMENTO:")
    modo_aba10 = st.radio(
        "Selecione o tipo de limpeza:",
        ["🧹 MODO 1: Limpeza Padrão WhatsApp (Ímã de Dados, Lavar Mensagens, Formatar Telefones)",
         "🤝 MODO 2: Unificador Mensal (Fundir linhas vazias, mantendo Data mais recente e Telefones limpos)"],
        label_visibility="collapsed"
    )
    st.markdown("---")

    if "MODO 2" in modo_aba10:
        st.info("💡 **Aviso:** Neste modo, o robô vai juntar o Paciente X que está duplicado, aplicar o **Novo Motor de Telefonia** (coluna TEL. ADIC.), formatar a conduta e usar o **Desempate de Prazo**!")
        col_n2, col_name2, col_d2 = st.columns(3)
        nm_num2 = col_n2.text_input("Nome da coluna NUM/Prontuário:", "NUM", key="m2_n")
        nm_nome2 = col_name2.text_input("Nome da coluna NOME:", "PACIENTE", key="m2_name")
        nm_data2 = col_d2.text_input("Nome da coluna DATA:", "DATA VISITA", key="m2_d")
        f_geral2 = st.file_uploader("Suba o Relatório Mensal Duplicado (Excel ou CSV)", type=["xlsx", "csv"], key="dup_file_m2")
        if f_geral2:
            df_ba = load_excel_with_ui(f_geral2, "m2_view")
            if st.button("🚀 INICIAR UNIFICAÇÃO MENSAL", type="primary"):
                try:
                    with st.spinner("Buscando a data mais recente, limpando telefones e compactando a planilha..."):
                        colunas_originais = list(df_ba.columns)
                        df_ba['_ORIG_INDEX'] = range(len(df_ba))
                        for c in colunas_originais:
                            df_ba[c] = df_ba[c].apply(formatar_brasileiro_sem_hora)

                        # Limpa PROXIMA na entrada (Modo 2)
                        cols_proxima_m2 = [c for c in colunas_originais
                                           if 'PROXIMA' in str(c).upper() or 'PRÓXIMA' in str(c).upper()]
                        for c in cols_proxima_m2:
                            df_ba[c] = df_ba[c].apply(limpar_proxima)

                        cols_upper = [str(c).upper().strip() for c in colunas_originais]
                        idx_num = next((i for i, c in enumerate(cols_upper) if nm_num2.upper() in c), None)
                        idx_nome = next((i for i, c in enumerate(cols_upper) if nm_nome2.upper() in c), None)
                        idx_data = next((i for i, c in enumerate(cols_upper) if nm_data2.upper() in c), None)
                        idx_conduta = next((i for i, c in enumerate(cols_upper) if 'CONDUTA' in c), None)
                        if None in (idx_num, idx_nome, idx_data):
                            st.error("❌ MODO 2 requer as colunas NUM, NOME e DATA VISITA. Verifique os nomes nas caixas!")
                        else:
                            real_num = colunas_originais[idx_num]
                            real_nome = colunas_originais[idx_nome]
                            real_data = colunas_originais[idx_data]
                            real_conduta = colunas_originais[idx_conduta] if idx_conduta is not None else None
                            df_ba[real_nome] = df_ba[real_nome].astype(str).str.strip().str.upper()
                            # ── FUSÃO TEL. ADIC. + TEL. ADICIONAL ──────────────
                            col_tel_adic_orig  = next((c for c in colunas_originais if str(c).strip().upper() == 'TEL. ADIC.'), None)
                            col_tel_adicional  = next((c for c in colunas_originais if 'ADICIONAL' in str(c).upper()), None)

                            def fundir_tel_adic_m2(v1, v2):
                                def limpa(v):
                                    v = str(v).strip()
                                    if v.lower() in ['nan','none','<na>','']: return []
                                    partes = re.split(r'[|/]', v)
                                    resultado = []
                                    for p in partes:
                                        p = p.strip()
                                        d = re.sub(r'\D','', p)
                                        if len(d) >= 8 and len(set(d)) > 1:
                                            resultado.append(p)
                                    return resultado
                                vistos = set(); final = []
                                for num in limpa(v1) + limpa(v2):
                                    chave = re.sub(r'\D','', num)[-8:]
                                    if chave not in vistos:
                                        vistos.add(chave); final.append(num)
                                return '|'.join(final)

                            if col_tel_adic_orig and col_tel_adicional:
                                # Tem as DUAS colunas → funde tudo em TEL. ADICIONAL com |
                                df_ba[col_tel_adicional] = df_ba.apply(
                                    lambda row: fundir_tel_adic_m2(
                                        row.get(col_tel_adicional, ''),
                                        row.get(col_tel_adic_orig, '')
                                    ), axis=1
                                )
                                df_ba[col_tel_adic_orig] = ''

                            if real_conduta:
                                df_ba[real_conduta] = df_ba[real_conduta].apply(formatar_conduta)
                                df_ba['_SCORE_CONDUTA'] = df_ba[real_conduta].apply(rank_conduta)
                            else:
                                df_ba['_SCORE_CONDUTA'] = 999
                            df_clean = df_ba.copy()
                            df_clean = df_clean.replace(r'^\s*$', pd.NA, regex=True)
                            df_clean = df_clean.replace(['nan', 'NaN', 'None', '<NA>', ''], pd.NA)
                            df_clean['_GRP_NUM'] = df_clean[real_num].astype(str).str.strip().replace(r'\.0$', '', regex=True).str.upper()
                            df_clean['_GRP_NOME'] = df_clean[real_nome].astype(str).str.strip().str.upper()
                            df_clean['_TEMP_DATE'] = pd.to_datetime(df_clean[real_data], format='%d/%m/%Y', errors='coerce', dayfirst=True)
                            df_clean = df_clean.sort_values(by=['_GRP_NUM', '_GRP_NOME', '_SCORE_CONDUTA', '_TEMP_DATE'], ascending=[True, True, True, False])
                            total_linhas_antes = len(df_clean)
                            grp_cols = ['_GRP_NUM', '_GRP_NOME']

                            # Sanitiza telefones lixo antes do merge
                            cols_tel_raw_m2 = [c for c in colunas_originais
                                               if 'TEL' in str(c).upper() or 'CEL' in str(c).upper()]
                            for col_t in cols_tel_raw_m2:
                                df_clean[col_t] = df_clean[col_t].apply(
                                    lambda x: pd.NA if (
                                        pd.isna(x) or
                                        re.match(r'^[\s_\-\(\)\.]+$', str(x).strip()) or
                                        len(re.sub(r'\D', '', str(x))) < 8
                                    ) else x
                                )

                            campos_enrichment = [c for c in colunas_originais
                                                 if c not in [real_num, real_nome, real_data]
                                                 and 'TEL' not in c.upper()
                                                 and 'CEL' not in c.upper()
                                                 and 'MSG' not in c.upper()]
                            cols_telefone_m2 = [c for c in colunas_originais
                                                if 'TEL' in str(c).upper() or 'CEL' in str(c).upper()]
                            resultado_merge = []
                            for chave, grupo in df_clean.groupby(grp_cols, dropna=False):
                                grupo = grupo.sort_values('_TEMP_DATE', ascending=False).reset_index(drop=True)
                                linha_mestre = grupo.iloc[0].copy()
                                # 1. Enriquece campos vazios
                                for campo in campos_enrichment:
                                    if campo not in linha_mestre.index: continue
                                    val = str(linha_mestre.get(campo, '')).strip()
                                    if not val or val.lower() in ['nan', '', 'none', '<na>']:
                                        for _, row_antigo in grupo.iloc[1:].iterrows():
                                            val_antigo = str(row_antigo.get(campo, '')).strip()
                                            if val_antigo and val_antigo.lower() not in ['nan', '', 'none', '<na>']:
                                                linha_mestre[campo] = val_antigo
                                                break
                                # 2. Coleta telefones de todos
                                todos_fones = []
                                for _, row in grupo.iterrows():
                                    for col_tel in cols_telefone_m2:
                                        tel = str(row.get(col_tel, '')).strip()
                                        if tel and tel.lower() not in ['nan', '', 'none', '<na>']:
                                            todos_fones.append(tel)
                                fones_unicos = list(dict.fromkeys(todos_fones))
                                tel_principal = str(linha_mestre.get(cols_telefone_m2[0] if cols_telefone_m2 else '', '')).strip()
                                if not tel_principal or tel_principal.lower() in ['nan', '', 'none', '<na>']:
                                    tel_principal = fones_unicos[0] if fones_unicos else ''
                                tel_principal_fmt = processar_telefones_avancado(
                                    pd.Series({cols_telefone_m2[0]: tel_principal} if cols_telefone_m2 else {}),
                                    cols_telefone_m2[:1] if cols_telefone_m2 else []
                                )[0] if cols_telefone_m2 else tel_principal
                                adicionais_fmt = []
                                vistos = {tel_principal_fmt}
                                for fone in fones_unicos:
                                    if fone == tel_principal: continue
                                    fmt = processar_telefones_avancado(
                                        pd.Series({cols_telefone_m2[0]: fone} if cols_telefone_m2 else {}),
                                        cols_telefone_m2[:1] if cols_telefone_m2 else []
                                    )[0] if cols_telefone_m2 else fone
                                    if fmt and fmt not in vistos:
                                        vistos.add(fmt)
                                        adicionais_fmt.append(fmt)
                                if cols_telefone_m2:
                                    linha_mestre[cols_telefone_m2[0]] = tel_principal_fmt
                                # Filtra adicionais lixo
                                adicionais_limpos = [
                                    f for f in adicionais_fmt
                                    if f and len(re.sub(r'\D', '', f)) >= 8
                                    and not re.match(r'^[\s_\-\(\)\.]+$', f)
                                ]
                                linha_mestre['TEL. ADIC.'] = ' / '.join(adicionais_limpos)
                                resultado_merge.append(linha_mestre)

                            df_unified = pd.DataFrame(resultado_merge)
                            if '_ORIG_INDEX' in df_unified.columns:
                                df_unified = df_unified.sort_values(by='_ORIG_INDEX')
                            if 'TEL. ADIC.' not in colunas_originais and 'TEL. ADIC.' in df_unified.columns:
                                colunas_originais.append('TEL. ADIC.')
                            df_unified = df_unified[[c for c in colunas_originais if c in df_unified.columns]]
                            linhas_removidas = total_linhas_antes - len(df_unified)

                            col_prox_calc2 = next((c for c in list(df_unified.columns) if 'PROXIMA' in str(c).upper() and 'CALC' in str(c).upper()), None)
                            col_conduta2 = next((c for c in list(df_unified.columns) if 'CONDUTA' in str(c).upper()), None)
                            col_data2 = next((c for c in list(df_unified.columns) if nm_data2.upper() in str(c).upper()), None)
                            if col_prox_calc2 and col_conduta2 and col_data2:
                                df_unified[col_prox_calc2] = df_unified.apply(
                                    lambda row: calcular_proxima_data(row[col_data2], row[col_conduta2]), axis=1
                                )

                            b_zip_dup = io.BytesIO()
                            with zipfile.ZipFile(b_zip_dup, "w", zipfile.ZIP_DEFLATED) as zf:
                                buf_limpo = io.BytesIO()
                                with pd.ExcelWriter(buf_limpo, engine='openpyxl') as writer:
                                    df_unified.to_excel(writer, index=False, sheet_name='BASE_UNIFICADA')
                                    ws_new = writer.sheets['BASE_UNIFICADA']
                                    if f_geral2.name.endswith('.xlsx'):
                                        try:
                                            f_geral2.seek(0)
                                            wb_orig = openpyxl.load_workbook(f_geral2)
                                            ws_orig = wb_orig.active
                                            for col_letter, col_dim in ws_orig.column_dimensions.items():
                                                ws_new.column_dimensions[col_letter].width = col_dim.width
                                        except:
                                            for col in ws_new.columns: ws_new.column_dimensions[col[0].column_letter].width = 20
                                    else:
                                        for col in ws_new.columns: ws_new.column_dimensions[col[0].column_letter].width = 20
                                    phone_col_indices = []
                                    for cell in ws_new[1]:
                                        cell.font = Font(bold=True)
                                        if cell.value and ('TEL' in str(cell.value).upper() or 'CEL' in str(cell.value).upper()):
                                            phone_col_indices.append(cell.column)
                                    for row in ws_new.iter_rows(min_row=2):
                                        for cell in row:
                                            cell.alignment = Alignment(wrap_text=True, vertical='center')
                                            if cell.column in phone_col_indices and cell.value:
                                                if len(str(cell.value).replace('+55','')) > 0 and len(str(cell.value).replace('+55','')) < 10:
                                                    cell.font = Font(color="FF0000", bold=True)
                                    ws_new.freeze_panes = 'A2'
                                zf.writestr("RELATORIO_MENSAL_UNIFICADO.xlsx", buf_limpo.getvalue())
                            b_zip_dup.seek(0)
                            st.session_state['base_deduplicada_mensal'] = b_zip_dup.getvalue()
                            st.success(f"✅ Fusão Concluída! O robô unificou {linhas_removidas} linhas, ativou o desempate de condutas e formatou os telefones!")
                except Exception as e:
                    st.error(f"Erro ao processar: {e}")
        if st.session_state.get('base_deduplicada_mensal') is not None:
            st.download_button("📥 BAIXAR RELATÓRIO MENSAL UNIFICADO", st.session_state['base_deduplicada_mensal'], "HOVA_Relatorio_Mensal.zip", mime="application/zip", type="primary")

    else:
        st.info("💡 **Aviso:** Neste modo, o robô vai caçar ativamente Emails e CPFs em colunas erradas, juntar históricos de mensagens, aplicar o **Novo Motor de Telefonia** e formatar Condutas.")
        col_n, col_name, col_t, col_d = st.columns(4)
        nm_num = col_n.text_input("Nome da coluna do NUM/Prontuário:", "NUM")
        nm_nome = col_name.text_input("Nome da coluna de NOME:", "PACIENTE")
        nm_tel = col_t.text_input("Nome da coluna de TELEFONE:", "TELEFONE")
        nm_data = col_d.text_input("Nome da coluna de DATA:", "DATA VISITA")
        f_geral = st.file_uploader("Suba a sua Planilha Geral (Excel ou CSV)", type=["xlsx", "csv"], key="dup_file_m1")
        if f_geral:
            df_ba = load_excel_with_ui(f_geral, "m1_view")
            if st.button("🧹 LIMPAR, ASPIRAR E FORMATAR BASE", type="primary"):
                try:
                    with st.spinner("Escaneando dados, ativando novo motor de telefonia e limpando a base..."):
                        df_ba['_ORIG_INDEX'] = range(len(df_ba))
                        colunas_originais = list(df_ba.columns)
                        colunas_originais.remove('_ORIG_INDEX')

                        # ── PASSO 1: formata datas e limpa nan ──────────────────
                        for c in colunas_originais:
                            df_ba[c] = df_ba[c].apply(formatar_brasileiro_sem_hora)
                            df_ba[c] = df_ba[c].replace(['nan', 'NaN', 'None', '<NA>'], '')

                        # ── PASSO 2: limpa PROXIMA (RX OCULOS, MANTER OCULOS…) ─
                        cols_proxima = [c for c in colunas_originais
                                        if 'PROXIMA' in str(c).upper() or 'PRÓXIMA' in str(c).upper()]
                        for c in cols_proxima:
                            df_ba[c] = df_ba[c].apply(limpar_proxima)

                        # ── PASSO 3: sanitiza telefones lixo antes do merge ────
                        cols_telefone_raw = [c for c in colunas_originais
                                             if 'TEL' in str(c).upper() or 'CEL' in str(c).upper()]
                        for col_t in cols_telefone_raw:
                            df_ba[col_t] = df_ba[col_t].apply(
                                lambda x: '' if (
                                    str(x).strip() == '' or
                                    re.match(r'^[\s_\-\(\)\.]+$', str(x).strip()) or
                                    len(re.sub(r'\D', '', str(x))) < 8
                                ) else str(x)
                            )

                        cols_upper = [str(c).upper().strip() for c in colunas_originais]
                        idx_num = next((i for i, c in enumerate(cols_upper) if nm_num.upper() in c), None)
                        idx_nome = next((i for i, c in enumerate(cols_upper) if nm_nome.upper() in c), None)
                        idx_tel = next((i for i, c in enumerate(cols_upper) if nm_tel.upper() in c), None)
                        idx_data = next((i for i, c in enumerate(cols_upper) if nm_data.upper() in c), None)
                        idx_conduta = next((i for i, c in enumerate(cols_upper) if 'CONDUTA' in c), None)
                        if None in (idx_num, idx_nome, idx_tel, idx_data):
                            st.error("❌ Não foi possível encontrar todas as 4 colunas base. Verifique as caixas acima!")
                        else:
                            real_num = colunas_originais[idx_num]
                            real_nome = colunas_originais[idx_nome]
                            real_tel = colunas_originais[idx_tel]
                            real_data = colunas_originais[idx_data]
                            real_conduta = colunas_originais[idx_conduta] if idx_conduta is not None else None
                            df_ba[real_nome] = df_ba[real_nome].astype(str).str.strip().str.upper()

                            # ── PASSO 4: processa telefones e funde TEL. ADIC. + TEL. ADICIONAL ──
                            cols_telefone = [c for c in colunas_originais if 'TEL' in str(c).upper() or 'CEL' in str(c).upper()]
                            col_tel_adicional_orig = next((c for c in colunas_originais if 'ADICIONAL' in str(c).upper()), None)
                            col_tel_adic_curta     = next((c for c in colunas_originais if str(c).strip().upper() == 'TEL. ADIC.'), None)

                            def fundir_tel_adic_m1(v1, v2):
                                def limpa(v):
                                    v = str(v).strip()
                                    if v.lower() in ['nan','none','<na>','']: return []
                                    partes = re.split(r'[|/]', v)
                                    resultado = []
                                    for p in partes:
                                        p = p.strip()
                                        d = re.sub(r'\D','', p)
                                        if len(d) >= 8 and len(set(d)) > 1:
                                            resultado.append(p)
                                    return resultado
                                vistos = set(); final = []
                                for num in limpa(v1) + limpa(v2):
                                    chave = re.sub(r'\D','', num)[-8:]
                                    if chave not in vistos:
                                        vistos.add(chave); final.append(num)
                                return '|'.join(final)

                            if cols_telefone:
                                col_tel_main = cols_telefone[0]
                                tel_res = df_ba.apply(lambda row: processar_telefones_avancado(row, cols_telefone), axis=1)
                                df_ba[col_tel_main] = [r[0] for r in tel_res]
                                adicionais_lista = [r[1] for r in tel_res]

                                if col_tel_adicional_orig and col_tel_adic_curta:
                                    # Tem as DUAS colunas → funde tudo em TEL. ADICIONAL com |
                                    df_ba[col_tel_adicional_orig] = df_ba.apply(
                                        lambda row: fundir_tel_adic_m1(
                                            row.get(col_tel_adicional_orig, ''),
                                            row.get(col_tel_adic_curta, '')
                                        ), axis=1
                                    )
                                    if col_tel_adic_curta in colunas_originais:
                                        colunas_originais.remove(col_tel_adic_curta)
                                elif col_tel_adicional_orig:
                                    # Só tem TEL. ADICIONAL → grava adicionais nela com |
                                    df_ba[col_tel_adicional_orig] = [
                                        r.replace(' / ','|').replace('/','|').strip('|')
                                        for r in adicionais_lista
                                    ]
                                else:
                                    # Nenhuma coluna de adicional → cria TEL. ADIC.
                                    if 'TEL. ADIC.' not in df_ba.columns:
                                        colunas_originais.append('TEL. ADIC.')
                                    df_ba['TEL. ADIC.'] = [
                                        r.replace(' / ','|').replace('/','|').strip('|')
                                        for r in adicionais_lista
                                    ]

                            if real_conduta:
                                df_ba[real_conduta] = df_ba[real_conduta].apply(formatar_conduta)
                                df_ba['_SCORE_CONDUTA'] = df_ba[real_conduta].apply(rank_conduta)
                            else:
                                df_ba['_SCORE_CONDUTA'] = 999

                            msg_cols = [c for c in colunas_originais if 'MSG' in str(c).upper()]
                            col_email = next((c for c in colunas_originais if 'MAIL' in str(c).upper()), None)
                            col_cpf = next((c for c in colunas_originais if 'CPF' in str(c).upper()), None)
                            col_nasc = next((c for c in colunas_originais if 'NASCIMENTO' in str(c).upper()), None)
                            regex_email = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')
                            regex_cpf = re.compile(r'\b\d{3}\.\d{3}\.\d{3}-\d{2}\b')
                            regex_data = re.compile(r'\b\d{2}/\d{2}/\d{4}\b')

                            # ── PASSO 5: caça email/CPF/data em colunas erradas
                            for index, row in df_ba.iterrows():
                                for col in colunas_originais:
                                    val = str(row[col]).strip()
                                    if val and val.lower() not in ['nan', 'none', '<na>']:
                                        if col_email and col != col_email:
                                            achados_email = regex_email.findall(val)
                                            if achados_email:
                                                if not str(df_ba.at[index, col_email]).strip() or str(df_ba.at[index, col_email]).lower() == 'nan':
                                                    df_ba.at[index, col_email] = achados_email[0]
                                                val = val.replace(achados_email[0], '').strip()
                                                df_ba.at[index, col] = val
                                        if col_cpf and col != col_cpf:
                                            achados_cpf = regex_cpf.findall(val)
                                            if achados_cpf:
                                                if not str(df_ba.at[index, col_cpf]).strip() or str(df_ba.at[index, col_cpf]).lower() == 'nan':
                                                    df_ba.at[index, col_cpf] = achados_cpf[0]
                                                val = val.replace(achados_cpf[0], '').strip()
                                                df_ba.at[index, col] = val
                                        if col_nasc and col != col_nasc and col != real_data:
                                            achados_data = regex_data.findall(val)
                                            for d in achados_data:
                                                try:
                                                    ano = int(d.split('/')[-1])
                                                    if ano < 2020:
                                                        if not str(df_ba.at[index, col_nasc]).strip() or str(df_ba.at[index, col_nasc]).lower() == 'nan':
                                                            df_ba.at[index, col_nasc] = d
                                                        val = val.replace(d, '').strip()
                                                        df_ba.at[index, col] = val
                                                except: pass

                            def lavar_mensagens_duplicadas(lista_msgs):
                                msgs_limpas = {}
                                for val_msg in lista_msgs:
                                    val_str = str(val_msg).strip()
                                    if val_str and val_str.lower() not in ['nan', 'none', 'nat', '<na>']:
                                        for linha_msg in val_str.split('\n'):
                                            linha_msg = linha_msg.strip()
                                            if linha_msg:
                                                linha_msg = linha_msg.replace(' 00:00:00', '')
                                                linha_msg = re.sub(r'\s*\b\d{2}:\d{2}(:\d{2})?\b', '', linha_msg).strip()
                                                frase_matematica = re.sub(r'[^A-Z0-9]', '', linha_msg.upper())
                                                if frase_matematica in msgs_limpas:
                                                    if linha_msg.count(' ') > msgs_limpas[frase_matematica].count(' '):
                                                        msgs_limpas[frase_matematica] = linha_msg
                                                else:
                                                    msgs_limpas[frase_matematica] = linha_msg
                                def date_key(txt):
                                    match = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4})', txt)
                                    if match:
                                        try: return pd.to_datetime(match.group(1), format='%d/%m/%Y', dayfirst=True)
                                        except: return pd.Timestamp.min
                                    return pd.Timestamp.min
                                return "\n".join(sorted(list(msgs_limpas.values()), key=date_key))

                            df_ba['_GRP_NUM'] = df_ba[real_num].astype(str).str.strip().replace(r'\.0$', '', regex=True).str.upper()
                            df_ba['_GRP_NOME'] = df_ba[real_nome].astype(str).str.strip().str.upper()
                            df_ba['_GRP_TEL'] = df_ba[real_tel].apply(limpar_num)
                            df_ba['_NOME_NORM'] = df_ba['_GRP_NOME'].apply(lambda n: re.sub(r'[^A-Z0-9]', '', str(n).upper()))
                            df_ba['_TEMP_DATE'] = pd.to_datetime(df_ba[real_data], format='%d/%m/%Y', errors='coerce', dayfirst=True)
                            df_ba = df_ba.sort_values(by=['_GRP_NUM', '_NOME_NORM', '_SCORE_CONDUTA', '_TEMP_DATE'], ascending=[True, True, True, False])
                            grp_cols = ['_GRP_NUM', '_NOME_NORM']

                            # Coleta duplicados para relatório
                            duplicated_mask = df_ba.duplicated(subset=grp_cols, keep='first')
                            df_duplicados = df_ba[duplicated_mask].copy()

                            # ── PASSO 6: MERGE INTELIGENTE com hierarquia de data
                            cols_telefone_m1 = [c for c in colunas_originais
                                                if 'TEL' in str(c).upper() or 'CEL' in str(c).upper()]
                            campos_enrichment_m1 = [c for c in colunas_originais
                                                    if c not in [real_num, real_nome, real_tel, real_data]
                                                    and 'TEL' not in c.upper()
                                                    and 'CEL' not in c.upper()
                                                    and 'MSG' not in c.upper()]
                            resultado_merge_m1 = []
                            for chave, grupo in df_ba.groupby(grp_cols, dropna=False):
                                grupo = grupo.sort_values('_TEMP_DATE', ascending=False).reset_index(drop=True)
                                linha_mestre = grupo.iloc[0].copy()

                                # 1. Enriquece campos vazios com dados de registros antigos
                                for campo in campos_enrichment_m1:
                                    if campo not in linha_mestre.index: continue
                                    val = str(linha_mestre.get(campo, '')).strip()
                                    if not val or val.lower() in ['nan', '', 'none', '<na>']:
                                        for _, row_antigo in grupo.iloc[1:].iterrows():
                                            val_antigo = str(row_antigo.get(campo, '')).strip()
                                            if val_antigo and val_antigo.lower() not in ['nan', '', 'none', '<na>']:
                                                linha_mestre[campo] = val_antigo
                                                break

                                # 2. Unifica mensagens de todos os registros
                                for col_msg in msg_cols:
                                    if col_msg in grupo.columns:
                                        linha_mestre[col_msg] = lavar_mensagens_duplicadas(grupo[col_msg].tolist())

                                # 3. Coleta todos os telefones de todos os registros
                                todos_fones = []
                                for _, row in grupo.iterrows():
                                    for col_tel in cols_telefone_m1:
                                        tel = str(row.get(col_tel, '')).strip()
                                        if tel and tel.lower() not in ['nan', '', 'none', '<na>']:
                                            todos_fones.append(tel)
                                fones_unicos = list(dict.fromkeys(todos_fones))

                                # 4. Principal = telefone do registro mais recente
                                tel_principal = str(linha_mestre.get(real_tel, '')).strip()
                                if not tel_principal or tel_principal.lower() in ['nan', '', 'none', '<na>']:
                                    tel_principal = fones_unicos[0] if fones_unicos else ''
                                tel_principal_fmt = processar_telefones_avancado(
                                    pd.Series({real_tel: tel_principal}), [real_tel]
                                )[0] if tel_principal else ''

                                # 5. Adicionais = demais telefones diferentes do principal
                                adicionais_fmt = []
                                vistos = {tel_principal_fmt}
                                for fone in fones_unicos:
                                    if fone == tel_principal: continue
                                    fmt = processar_telefones_avancado(
                                        pd.Series({real_tel: fone}), [real_tel]
                                    )[0]
                                    if fmt and fmt not in vistos:
                                        vistos.add(fmt)
                                        adicionais_fmt.append(fmt)

                                linha_mestre[real_tel] = tel_principal_fmt
                                # Filtra adicionais lixo (máscara vazia, menos de 8 dígitos)
                                adicionais_limpos = [
                                    f for f in adicionais_fmt
                                    if f and not re.match(r'^[\s_\-\(\)\.]+$', f)
                                    and len(re.sub(r'\D', '', f)) >= 8
                                ]
                                linha_mestre['TEL. ADIC.'] = ' / '.join(adicionais_limpos)
                                resultado_merge_m1.append(linha_mestre)

                            df_limpo = pd.DataFrame(resultado_merge_m1)
                            df_limpo = df_limpo.sort_values(by='_ORIG_INDEX')

                            # Garante TEL. ADIC. nas colunas finais
                            if 'TEL. ADIC.' not in colunas_originais:
                                colunas_originais.append('TEL. ADIC.')

                            _num_series = df_limpo[real_num].astype(str).str.strip().str.upper()
                            _num_counts = _num_series.map(_num_series.value_counts())
                            df_casos_especiais = df_limpo[_num_counts > 1].copy()

                            df_limpo = df_limpo[colunas_originais]
                            df_duplicados = df_duplicados[colunas_originais]
                            if not df_casos_especiais.empty:
                                df_casos_especiais = df_casos_especiais[colunas_originais]

                            col_prox_calc = next((c for c in colunas_originais if 'PROXIMA' in str(c).upper() and 'CALC' in str(c).upper()), None)
                            if col_prox_calc and real_conduta and real_data in df_limpo.columns:
                                df_limpo[col_prox_calc] = df_limpo.apply(
                                    lambda row: calcular_proxima_data(row[real_data], row[real_conduta]), axis=1
                                )

                            b_zip_dup = io.BytesIO()
                            with zipfile.ZipFile(b_zip_dup, "w", zipfile.ZIP_DEFLATED) as zf:
                                buf_limpo = io.BytesIO()
                                with pd.ExcelWriter(buf_limpo, engine='openpyxl') as writer:
                                    df_limpo.to_excel(writer, index=False, sheet_name='BASE_LIMPA')
                                    ws_new = writer.sheets['BASE_LIMPA']
                                    if f_geral.name.endswith('.xlsx'):
                                        try:
                                            f_geral.seek(0)
                                            wb_orig = openpyxl.load_workbook(f_geral)
                                            ws_orig = wb_orig.active
                                            for col_letter, col_dim in ws_orig.column_dimensions.items():
                                                ws_new.column_dimensions[col_letter].width = col_dim.width
                                        except:
                                            for col in ws_new.columns: ws_new.column_dimensions[col[0].column_letter].width = 20
                                    else:
                                        for col in ws_new.columns: ws_new.column_dimensions[col[0].column_letter].width = 20
                                    phone_col_indices = []
                                    for cell in ws_new[1]:
                                        cell.font = Font(bold=True)
                                        if cell.value and ('TEL' in str(cell.value).upper() or 'CEL' in str(cell.value).upper() or 'ADIC' in str(cell.value).upper()):
                                            phone_col_indices.append(cell.column)
                                    for row in ws_new.iter_rows(min_row=2):
                                        for cell in row:
                                            cell.alignment = Alignment(wrap_text=True, vertical='center')
                                            if cell.column in phone_col_indices and cell.value:
                                                if len(str(cell.value).replace('+55','')) > 0 and len(str(cell.value).replace('+55','')) < 10:
                                                    cell.font = Font(color="FF0000", bold=True)
                                    ws_new.freeze_panes = 'A2'
                                zf.writestr("1_BASE_LIMPA_E_ASPIRADA.xlsx", buf_limpo.getvalue())
                                if not df_duplicados.empty:
                                    buf_dup = io.BytesIO()
                                    df_duplicados.to_excel(buf_dup, index=False)
                                    zf.writestr("2_DUPLICADOS_DESCARTADOS.xlsx", buf_dup.getvalue())
                                if not df_casos_especiais.empty:
                                    buf_esp = io.BytesIO()
                                    with pd.ExcelWriter(buf_esp, engine='openpyxl') as writer_esp:
                                        df_casos_especiais.to_excel(writer_esp, index=False, sheet_name='CASOS_ESPECIAIS')
                                        ws_esp = writer_esp.sheets['CASOS_ESPECIAIS']
                                        fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                                        for cell in ws_esp[1]:
                                            cell.font = Font(bold=True)
                                            if real_num.upper() in str(cell.value).upper():
                                                for row_esp in ws_esp.iter_rows(min_row=1, min_col=cell.column, max_col=cell.column):
                                                    for c_esp in row_esp:
                                                        c_esp.fill = fill_amarelo
                                        for col in ws_esp.columns:
                                            ws_esp.column_dimensions[col[0].column_letter].width = 25
                                    zf.writestr("3_CASOS_ESPECIAIS_REVISAR.xlsx", buf_esp.getvalue())
                            b_zip_dup.seek(0)
                            st.session_state['base_deduplicada_padrao'] = b_zip_dup.getvalue()
                            st.session_state['casos_especiais_count'] = len(df_casos_especiais) if not df_casos_especiais.empty else 0
                            st.success(f"✅ Limpeza Master Concluída! {len(df_duplicados)} duplicados pulverizados.")
                            if not df_casos_especiais.empty:
                                st.warning(f"🔍 **RADAR DE CASOS ESPECIAIS:** {len(df_casos_especiais)} pacientes sinalizados em **3_CASOS_ESPECIAIS_REVISAR.xlsx**.")
                except Exception as e:
                    st.error(f"Erro ao processar: {e}")
        if st.session_state.get('base_deduplicada_padrao') is not None:
            st.download_button("📥 BAIXAR BASE LIMPA FORMATADA", st.session_state['base_deduplicada_padrao'], "HOVA_Base_Limpa.zip", mime="application/zip", type="primary")
    st.markdown('</div>', unsafe_allow_html=True)

with tab10:
    st.markdown('<div class="master-card" style="padding: 20px;">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title"> ESQUADRÃO HOVA MASTER INTELLIGENCE</div>""", unsafe_allow_html=True)
    import os
    import base64
    import streamlit.components.v1 as components

    def get_foto_b64(nome):
        primeiro_nome = nome.split(' ')[0].lower()
        caminhos = [f"{primeiro_nome}.jpg", f"{primeiro_nome}.png", f"AGENTES IA/{primeiro_nome}.jpg", f"AGENTES IA/{primeiro_nome}.png"]
        for c in caminhos:
            if os.path.exists(c):
                try:
                    with open(c, "rb") as f:
                        ext = "png" if "png" in c.lower() else "jpeg"
                        return f"data:image/{ext};base64,{base64.b64encode(f.read()).decode()}"
                except: pass
        return f"https://ui-avatars.com/api/?name={primeiro_nome}&background=1e3d3a&color=00ffcc&size=150&rounded=true&bold=true"

    agentes_lista = [
        {"n": "LUMINA ALMEIDA",  "s": "WEB",     "z": "31 9723-6408",  "g": "F", "t": "busca"},
        {"n": "PITER SANTOS",    "s": "SLOT 11",  "z": "31 9528-5492",  "g": "M", "t": "busca"},
        {"n": "CLARA MARTINS",   "s": "SLOT 07",  "z": "31 9743-4631",  "g": "F", "t": "busca"},
        {"n": "PRISMA RAMOS",    "s": "SLOT 20",  "z": "31 7221-8952",  "g": "F", "t": "busca"},
        {"n": "ROGER OLIVEIRA",  "s": "SLOT 14",  "z": "31 9953-2096",  "g": "M", "t": "confirm"},
        {"n": "NATALIA VIANA",   "s": "SLOT 01",  "z": "31 7150-8930",  "g": "F", "t": "catarata"},
        {"n": "STELLA VEIRA",    "s": "SLOT 10",  "z": "31 9670-1479",  "g": "F", "t": "confirm"},
        {"n": "ESTER TEIXEIRA",  "s": "SLOT 16",  "z": "31 97202-3913", "g": "F", "t": "confirm"},
        {"n": "AYLA FREITAS",    "s": "SLOT 04",  "z": "N/A",           "g": "F", "t": "yag"},
        {"n": "OSCAR SIQUEIRA",  "s": "SLOT 13",  "z": "N/A",           "g": "M", "t": "busca"},
        {"n": "THEIA DIAS",      "s": "SLOT 12",  "z": "31 9788-9331",  "g": "F", "t": "busca"}
    ]

    cards_html = ""
    for i, ag in enumerate(agentes_lista):
        nome = ag['n']
        art = "o" if ag['g'] == "M" else "a"
        msg_busca = f"Olá, *{{{{column_2}}}}*! Tudo bem? 👁️\n\nAqui é {art} *{nome}*, {art}ssistente do *Hospital de Olhos Vale do Aço*.\n\nEstava revisando seu histórico e percebi que já faz um tempo desde sua última consulta. A saúde dos seus olhinhos merece atenção regular — está na hora de realizar uma nova avaliação preventiva.\n\n📲 *Para agendar:*\n1. Toque no link azul abaixo\n2. O WhatsApp abre automaticamente\n3. Nossa equipe vai te atender\n\n👇 *Clique aqui para agendar:*\nhttps://wa.me/553138011800\n\n*Aviso:* Este número só envia lembretes. Para agendar, use o link acima.\n\n💡 Já consultou recentemente? Pode desconsiderar esta mensagem.\n\nPodemos te esperar? 🤍\n\n*{nome} — Hospital de Olhos Vale do Aço*"
        msg_conf = f"Olá, {{{{column_4}}}}! Tudo bem? 👁️\n\nAqui é {art} *{nome}*, su{art} assistente do Hospital de Olhos Vale do Aço.\n\n📌 *Lembrete automático do seu atendimento:*\n\n📋 *Atendimento:* {{{{column_6}}}}\n👨‍⚕️ *Médico(a):* {{{{column_7}}}}\n📅 *Data:* {{{{column_1}}}}\n⏰ *Chegada:* {{{{column_2}}}}\n\n🚨 *PREPARO OBRIGATÓRIO:*\n\n{{{{column_8}}}}\n\n📁 *O QUE TRAZER:*\n\n{{{{column_9}}}}\n\n⚠️ *CONFIRMAÇÃO:*\n\n✅ *Confirmar:* Responda *SIM* _(Sistema registra automaticamente)_\n\n❌ *Cancelar/Remarcar:* Não responda aqui\nTemos fila de espera. Clique no link:\n👉 https://wa.me/553138011800\n\n🤖 Mensagem automática. Sistema em atualização.\n\n🤍 *{nome} — Hospital de Olhos Vale do Aço*"
        msg_cat = f"Olá, *{{{{column_2}}}}*! 👁️\n\nAqui é a *NATALIA VIANA*, do Hospital de Olhos Vale do Aço.\n\nVou te enviar um áudio explicativo com informações importantes sobre seu atendimento. Por favor, ouça com atenção. 🎧\n\nQualquer dúvida, estou à disposição. 🤍\n\nCaso queira agendar, entre em contato pelo WhatsApp: https://wa.me/553138011800\n\nObs.: Caso você já esteja se preparando ou já tenha realizado o procedimento, por gentileza, desconsidere esta mensagem."
        msg_pterigio = f"Olá, *{{{{column_2}}}}*! 👁️\nAqui é a *{nome}*, do Hospital de Olhos Vale do Aço.\n\nEstou te enviando uma orientação em áudio da Dra. Mariluci sobre a sua saúde visual. Ela preparou esse recado para acompanhar de perto o seu caso e garantir que você tenha todo o suporte necessário. ✨\n\nPeço que, por gentileza, ouça com atenção. É muito importante para o seu conforto visual e para evitar que a irritação nos olhos aumente.\n\nNossa equipe está à disposição para te orientar com todo o carinho e atenção. 🤍\n\nPara agendar, ligue agora ou mande uma mensagem no número 31 3801-1800.\nObs.: Caso você já tenha realizado o procedimento, por gentileza, desconsidere esta mensagem."
        msg_yag = "Olá, {{{{column_2}}}}! 👁️\nAqui é a Ayla, do Hospital de Olhos Vale do Aço.\nSabemos o quanto a sua visão é valiosa para a sua independência e bem-estar. Por isso, preparei um áudio explicativo muito importante sobre o próximo passo para restaurar a nitidez do seu olhar.\nPor favor, ouça com atenção, pois queremos garantir que você tenha a melhor qualidade de vida possível após a sua cirurgia. 🤍\nPara agendar, ligue agora ou mande uma mensagem no número 3138011800.\nObs.: Caso você já tenha realizado o procedimento, por gentileza, desconsidere esta mensagem."

        img_b64 = get_foto_b64(nome)
        delay = i * 0.05
        txt_busca_js    = msg_busca.replace('\n', '\\n').replace('`', '\\`')
        txt_conf_js     = msg_conf.replace('\n', '\\n').replace('`', '\\`')
        txt_cat_js      = msg_cat.replace('\n', '\\n').replace('`', '\\`')
        txt_pterigio_js = msg_pterigio.replace('\n', '\\n').replace('`', '\\`')
        txt_yag_js      = msg_yag.replace('\n', '\\n').replace('`', '\\`')
        btn_busca_class = "btn-solid" if ag['t'] == 'busca' else "btn-outline"
        btn_conf_class  = "btn-solid" if ag['t'] == 'confirm' else "btn-outline"
        botoes_html = f"""
            <button class="btn-ag {btn_busca_class}" onclick="copiar(this, `{txt_busca_js}`)">📋 BUSCA ATIVA</button>
            <button class="btn-ag {btn_conf_class}" onclick="copiar(this, `{txt_conf_js}`)">📅 CONFIRMAÇÃO</button>
        """
        if ag['t'] == 'catarata':
            botoes_html += f"""<button class="btn-ag btn-solid btn-catarata" onclick="copiar(this, `{txt_cat_js}`)">🎧 ÁUDIO CATARATA</button>"""
        if ag['t'] == 'yag':
            botoes_html += f"""<button class="btn-ag btn-solid btn-catarata" onclick="copiar(this, `{txt_yag_js}`)">💎 ÁUDIO YAG LASER</button>"""
        botoes_html += f"""<button class="btn-ag btn-outline btn-pterigio" onclick="copiar(this, `{txt_pterigio_js}`)">🌿 ÁUDIO PTERÍGIO</button>"""

        cards_html += f"""
        <div class="card-ag" style="animation-delay: {delay}s;">
            <div class="img-wrapper"><img src="{img_b64}"></div>
            <div class="nome-ag">{nome}</div>
            <div class="slot-ag">{ag['s']}</div>
            <div class="zap-ag">📱 {ag['z']}</div>
            <div class="botoes-container">{botoes_html}</div>
        </div>
        """

    html_final = f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap');
        :root {{ --verde-escuro: #1e3d3a; --verde-claro: #2f6c68; --destaque: #00ffcc; }}
        .grid-ag {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(360px, 1fr)); gap: 30px; padding: 20px; font-family: 'Outfit', sans-serif; width: 100%; }}
        @keyframes fadeSlideUp {{ 0% {{ opacity: 0; transform: translateY(40px); }} 100% {{ opacity: 1; transform: translateY(0); }} }}
        .card-ag {{ background: #ffffff; border-radius: 25px; padding: 35px 20px; text-align: center; box-shadow: 0 10px 30px rgba(30,61,58,0.08); border: 1px solid rgba(30,61,58,0.1); opacity: 0; animation: fadeSlideUp 0.6s cubic-bezier(0.16, 1, 0.3, 1) forwards; transition: all 0.3s ease; position: relative; overflow: hidden; display: flex; flex-direction: column; }}
        .card-ag::before {{ content: ''; position: absolute; top: 0; left: 0; right: 0; height: 6px; background: linear-gradient(90deg, var(--verde-escuro), var(--destaque)); transform: scaleX(0); transition: transform 0.4s ease; transform-origin: left; }}
        .card-ag:hover {{ transform: translateY(-10px); box-shadow: 0 20px 40px rgba(30,61,58,0.15); border-color: rgba(30,61,58,0.3); z-index: 10; }}
        .card-ag:hover::before {{ transform: scaleX(1); }}
        .img-wrapper {{ width: 150px; height: 150px; margin: 0 auto 20px; border-radius: 50%; padding: 5px; background: linear-gradient(135deg, var(--verde-escuro), var(--verde-claro)); box-shadow: 0 8px 25px rgba(30,61,58,0.25); transition: 0.3s ease; }}
        .card-ag:hover .img-wrapper {{ transform: scale(1.05) rotate(3deg); }}
        .img-wrapper img {{ width: 100%; height: 100%; object-fit: cover; border-radius: 50%; border: 4px solid #fff; background: white; }}
        .nome-ag {{ font-weight: 800; color: var(--verde-escuro); font-size: 1.4rem; letter-spacing: 0.5px; margin-bottom: 6px; }}
        .slot-ag {{ background: var(--verde-escuro); color: #fff; padding: 5px 15px; border-radius: 20px; font-size: 0.85rem; font-weight: bold; display: inline-block; margin-bottom: 15px; }}
        .zap-ag {{ color: #64748b; font-size: 1rem; font-weight: 700; display: block; margin-bottom: 25px; }}
        .botoes-container {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: auto; }}
        .btn-ag {{ padding: 14px 5px; border-radius: 12px; cursor: pointer; font-weight: 800; transition: 0.2s; font-size: 0.85rem; text-transform: uppercase; display: flex; align-items: center; justify-content: center; font-family: 'Outfit', sans-serif; }}
        .btn-solid {{ background: var(--verde-escuro); color: white; border: none; box-shadow: 0 6px 15px rgba(30,61,58,0.15); }}
        .btn-solid:hover {{ background: var(--verde-claro); transform: translateY(-2px); box-shadow: 0 8px 20px rgba(30,61,58,0.25); }}
        .btn-outline {{ background: transparent; color: var(--verde-escuro); border: 2px solid rgba(30,61,58,0.25); }}
        .btn-outline:hover {{ background: rgba(30,61,58,0.05); border-color: var(--verde-escuro); transform: translateY(-2px); }}
        .btn-catarata {{ grid-column: span 2; background: #e67e22; color: white; border: none; }}
        .btn-catarata:hover {{ background: #d35400; }}
        .btn-pterigio {{ grid-column: span 2; background: transparent; color: #2f6c68; border: 2px solid #2f6c68; }}
        .btn-pterigio:hover {{ background: rgba(47,108,104,0.08); }}
        .btn-ag:active {{ transform: scale(0.95); }}
        .btn-ag.success {{ background: #000 !important; color: var(--destaque) !important; border-color: #000 !important; grid-column: span 2; font-size: 0.95rem; }}
    </style>
    <div class="grid-ag">{cards_html}</div>
    <script>
        function copiar(btn, texto) {{
            const txtCorrigido = texto.replace(/\\n/g, '\\n');
            navigator.clipboard.writeText(txtCorrigido).then(() => {{
                const txtVelho = btn.innerHTML;
                btn.innerHTML = "✅ COPIADO!";
                btn.classList.add('success');
                setTimeout(() => {{ btn.innerHTML = txtVelho; btn.classList.remove('success'); }}, 1500);
            }});
        }}
    </script>
    """
    components.html(html_final, height=1800, scrolling=True)
    st.markdown('</div>', unsafe_allow_html=True)

with tab11:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">SOLICITADOS x ATENDIDOS</div>""", unsafe_allow_html=True)
    MAPA_ATENDIDOS = {
    'NUM': 'NUM', 'NÚM': 'NUM', 'NÚMERO': 'NUM', 'N°': 'NUM', 'Nº': 'NUM',
    'NÚMERO DO PRONTUÁRIO': 'NUM', 'NUMERO': 'NUM',
    'PACIENTE': 'NOME', 'CATEGORIA': 'CONV',
    'DT.VISITA': 'DATA VISITA', 'DT. VISITA': 'DATA VISITA', 'DTVISITA': 'DATA VISITA',
    'DATA': 'DATA VISITA', 'DATA VISITA': 'DATA VISITA',
    'MÉDICO': 'MEDICO', 'MEDICO': 'MEDICO',
    'ATENDIMENTO': 'CONDUTA',
    'INDICAÇÃO': 'PACIENTE GLAUCOMA', 'INDICACAO': 'PACIENTE GLAUCOMA',
    'FONE INDICAÇÃO': 'TEL.ADIC', 'FONE INDICACAO': 'TEL.ADIC',
    'FONE ADICIONAL': 'TEL.ADIC', 'FONE ADICIONAIS': 'TEL.ADIC',
    'TELEFONE': 'TELEFONE',
    'DT. NASCIMENTO': 'DATA NASC', 'DT.NASCIMENTO': 'DATA NASC', 'DTNASCIMENTO': 'DATA NASC',
    'PROFISSÃO': 'PROFISSAO', 'PROFISSAO': 'PROFISSAO', 'CPF': 'CPF',
    'EMAIL': 'EMAIL', 'E-MAIL': 'EMAIL', 'MAIL': 'EMAIL',
}
    MAPA_SOLICITADOS = {
        'NÚM': 'NUM', 'NUM': 'NUM', 'PACIENTE': 'NOME', 'TELEFONE': 'TELEFONE',
        'CATEGORIA': 'CONV', 'DATA': 'DATA VISITA', 'MÉDICO': 'MEDICO', 'MEDICO': 'MEDICO',
        'ATENDIMENTO': 'CONDUTA', 'CONDUTA MÉDICA:': 'PROXIMA', 'CONDUTA MÉDICA': 'PROXIMA',
        'INDICAÇÃO': 'PACIENTE GLAUCOMA', 'INDICACAO': 'PACIENTE GLAUCOMA',
        # FONE INDICAÇÃO e FONE ADICIONAL mantidos com nomes originais
        # para o motor aplicar_telefone capturar ambos via filtro 'FONE'
        'DT. NASCIMENTO': 'DATA NASC', 'DT.NASCIMENTO': 'DATA NASC',
        'CPF': 'CPF', 'EMAIL': 'EMAIL', 'MAIL': 'EMAIL',
    }
    CABECALHO_PADRAO = [
        'NUM', 'NOME', 'TELEFONE', 'TEL.ADIC', 'CONV',
        'DATA VISITA', 'MEDICO', 'CONDUTA', 'PROXIMA',
        'MSG2024', 'MSG2025', 'MSG2026',
        'PACIENTE GLAUCOMA', 'EMAIL', 'CPF', 'PROFISSAO', 'DATA NASC'
    ]
    col_at, col_sol = st.columns(2)
    f_atendidos  = col_at.file_uploader("📂 Base ATENDIDOS",  type=["xlsx","xls","csv"], key="f_atendidos")
    f_solicitados = col_sol.file_uploader("📂 Base SOLICITADOS", type=["xlsx","xls","csv"], key="f_solicitados")
    if f_atendidos and f_solicitados:
        if st.button("🚀 CRUZAR E UNIFICAR BASES", type="primary"):
            try:
                with st.spinner("Cruzando bases, traduzindo, padronizando e organizando..."):
                    def ler_arquivo(f):
                        if f.name.lower().endswith('.csv'):
                            try:
                                df = pd.read_csv(f, sep=';', dtype=str)
                                if len(df.columns) < 2:
                                    f.seek(0); df = pd.read_csv(f, sep=',', dtype=str)
                            except:
                                f.seek(0); df = pd.read_csv(f, sep=',', dtype=str)
                        elif f.name.lower().endswith('.xls'):
                            df = pd.read_excel(f, engine='xlrd', dtype=str, sheet_name=0)
                        else:
                            df = pd.read_excel(f, sheet_name=0, dtype=str, engine='openpyxl')
                        df = pd.DataFrame(df.values, columns=df.columns)
                        df.columns = [str(c).strip().upper() for c in df.columns]
                        return df

                    f_atendidos.seek(0);  df_at  = ler_arquivo(f_atendidos)
                    f_solicitados.seek(0); df_sol = ler_arquivo(f_solicitados)

                    # Formata datas em todas as colunas de ambas as bases
                    cols_data = ['DATA VISITA', 'DATA NASC', 'DT.VISITA', 'DT. VISITA',
                                 'DT. NASCIMENTO', 'DT.NASCIMENTO', 'DTNASCIMENTO']
                    for base in [df_at, df_sol]:
                        for col in base.columns:
                            if any(d in str(col).upper() for d in ['DATA', 'NASC', 'VISITA', 'NASCIMENTO']):
                                base[col] = base[col].apply(formatar_brasileiro_sem_hora)

                    novos_cols = {}
                    for col in df_at.columns:
                        chave = col.strip().upper()
                        novos_cols[col] = MAPA_ATENDIDOS.get(chave, col)
                    df_at = df_at.rename(columns=novos_cols)
                    df_at = df_at.loc[:, ~df_at.columns.duplicated(keep='first')]

                    novos_cols_sol = {}
                    for col in df_sol.columns:
                        chave = col.strip().upper()
                        novos_cols_sol[col] = MAPA_SOLICITADOS.get(chave, col)
                    df_sol = df_sol.rename(columns=novos_cols_sol)
                    df_sol = df_sol.loc[:, ~df_sol.columns.duplicated(keep='first')]

                    # Remove Prefeitura de Ipatinga de AMBAS as bases
                    for base_nome, base_df in [('Atendidos', df_at), ('Solicitados', df_sol)]:
                        if 'CONV' in base_df.columns:
                            mask_pref = base_df['CONV'].astype(str).str.upper().str.contains('PREFEITURA DE IPATINGA', na=False)
                            removidos = mask_pref.sum()
                            if base_nome == 'Atendidos':
                                df_at = base_df[~mask_pref].copy()
                            else:
                                df_sol = base_df[~mask_pref].copy()
                            if removidos > 0:
                                st.info(f"🚫 {removidos} paciente(s) com 'Prefeitura de Ipatinga' removido(s) da base {base_nome}.")

                    def garantir_colunas(df):
                        for col in CABECALHO_PADRAO:
                            if col not in df.columns: df[col] = ""
                        return df

                    df_at  = garantir_colunas(df_at)
                    df_sol = garantir_colunas(df_sol)

                    def aplicar_telefone(df):
                        # Coluna principal de telefone — prioridade WhatsApp (W)
                        col_principal = next((c for c in df.columns
                                              if str(c).upper() == 'TELEFONE'), None)

                        # Todas as colunas com número: TELEFONE, TEL.ADIC, FONE INDICAÇÃO, FONE ADICIONAL
                        cols_todos = [c for c in df.columns if any(p in str(c).upper() for p in
                                        ['TEL', 'CEL', 'FONE', 'PHONE'])
                                        and 'INDICAÇÃO' not in str(c).upper()
                                        and 'INDICACAO' not in str(c).upper()
                                        or str(c).upper() == 'TEL.ADIC']
                        if not cols_todos:
                            return df

                        def processar_linha(row):
                            # 1. Coleta todos os números válidos
                            def _limpa(t):
                                t = str(t).strip().upper()
                                if not t or t in ['NAN','NONE','<NA>','-','']: return ''
                                if re.match(r'^[\s_\-\(\)\.]+$', t): return ''
                                if re.match(r'^\(0+\)', t): return ''   # (000)___
                                d = re.sub(r'\D','', t)
                                if d.startswith('55') and len(d) >= 12: d = d[2:]
                                if len(d) < 8 or len(set(d)) == 1: return ''
                                return t  # devolve original para o motor principal decidir

                            brutos = []
                            for c in cols_todos:
                                v = _limpa(str(row.get(c, '')))
                                if v:
                                    brutos.append((c, v))

                            if not brutos:
                                return ['', '']

                            # 2. Separa WhatsApp dos demais
                            whasapps = [(c, v) for c, v in brutos if 'W' in v.upper()
                                        and str(c).upper() in ['TELEFONE', 'FONE ADICIONAL',
                                                                'FONE INDICAÇÃO', 'FONE INDICACAO']]
                            outros   = [(c, v) for c, v in brutos if (c, v) not in whasapps]

                            # 3. Principal = primeiro WhatsApp; se não tiver, primeiro da col TELEFONE
                            if whasapps:
                                principal_raw = whasapps[0][1]
                                resto_raw = [v for c, v in whasapps[1:]] + [v for c, v in outros]
                            elif col_principal and any(str(c).upper() == 'TELEFONE' for c, v in brutos):
                                principal_raw = next(v for c, v in brutos if str(c).upper() == 'TELEFONE')
                                resto_raw = [v for c, v in brutos if str(c).upper() != 'TELEFONE']
                            else:
                                principal_raw = brutos[0][1]
                                resto_raw = [v for c, v in brutos[1:]]

                            # 4. Formata pelo motor
                            def _fmt(t):
                                res = processar_telefones_avancado(
                                    pd.Series({'__TEL__': t}), ['__TEL__']
                                )
                                return res[0]

                            principal_fmt = _fmt(principal_raw)

                            # 5. Adicionais — todos os outros números diferentes do principal
                            vistos = {principal_fmt}
                            adicionais = []
                            for raw in resto_raw:
                                fmt = _fmt(raw)
                                if fmt and fmt not in vistos:
                                    vistos.add(fmt)
                                    adicionais.append(fmt)

                            return [principal_fmt, '|'.join(adicionais)]

                        resultados = df.apply(processar_linha, axis=1)
                        df['TELEFONE'] = [r[0] for r in resultados]
                        df['TEL.ADIC'] = [r[1] for r in resultados]
                        return df

                    df_at  = aplicar_telefone(df_at)
                    df_sol = aplicar_telefone(df_sol)
                    for _df in [df_at, df_sol]:
                        if 'MEDICO' in _df.columns:
                            _df['MEDICO'] = _df['MEDICO'].astype(str).str.strip().str.upper()
                    # ── CRUZAMENTO INTELIGENTE: NUM + NOME → data mais recente vence ──
                    df_at['_ORIGEM']  = 'ATENDIDO'
                    df_sol['_ORIGEM'] = 'SOLICITADO'

                    # Garante colunas padrão em ambas
                    for _df in [df_at, df_sol]:
                        for col in CABECALHO_PADRAO:
                            if col not in _df.columns: _df[col] = ''

                    # Concatena as duas bases
                    df_concat = pd.concat([df_sol, df_at], ignore_index=True)
                    df_concat = df_concat.fillna('').astype(str)
                    df_concat = df_concat.replace(['nan','NaN','None','<NA>'], '')

                    # Normaliza chaves de cruzamento
                    df_concat['_NUM_KEY']  = df_concat['NUM'].str.strip().str.upper().str.replace(r'\.0$','',regex=True)
                    df_concat['_NOME_KEY'] = df_concat['NOME'].str.strip().str.upper().apply(
                        lambda n: re.sub(r'[^A-Z0-9]','',n)
                    )
                    df_concat['_TEMP_DATE'] = pd.to_datetime(
                        df_concat['DATA VISITA'], format='%d/%m/%Y', dayfirst=True, errors='coerce'
                    )

                    # Ordena: mais recente primeiro
                    df_concat = df_concat.sort_values('_TEMP_DATE', ascending=False)

                    resultado_cruzado = []
                    stats_cruzados = 0
                    stats_so_sol   = 0
                    stats_so_at    = 0

                    for (num_key, nome_key), grupo in df_concat.groupby(['_NUM_KEY','_NOME_KEY'], dropna=False):
                        grupo = grupo.reset_index(drop=True)

                        # Registro mestre = data mais recente (primeiro após sort desc)
                        mestre = grupo.iloc[0].copy()

                        # Herda CONDUTA MÉDICA se o mais recente não tiver
                        campo_conduta = 'PROXIMA'
                        if not mestre.get(campo_conduta,'').strip() or \
                           str(mestre.get(campo_conduta,'')).lower() in ['nan','none','<na>']:
                            for _, row_old in grupo.iloc[1:].iterrows():
                                val = str(row_old.get(campo_conduta,'')).strip()
                                if val and val.lower() not in ['nan','none','<na>','']:
                                    mestre[campo_conduta] = val
                                    break

                        # Herda TELEFONE se o mais recente não tiver
                        # (Atendidos frequentemente vem sem telefone — Solicitados sempre tem)
                        if not mestre.get('TELEFONE','').strip() or \
                           str(mestre.get('TELEFONE','')).lower() in ['nan','none','<na>']:
                            for _, row_old in grupo.iloc[1:].iterrows():
                                tel = str(row_old.get('TELEFONE','')).strip()
                                if tel and tel.lower() not in ['nan','none','<na>','']:
                                    mestre['TELEFONE'] = tel
                                    # Herda TEL.ADIC junto se tiver
                                    tel_adic = str(row_old.get('TEL.ADIC','')).strip()
                                    if tel_adic and tel_adic.lower() not in ['nan','none','<na>',''] \
                                       and not mestre.get('TEL.ADIC','').strip():
                                        mestre['TEL.ADIC'] = tel_adic
                                    break

                        # Estatísticas
                        origens = set(grupo['_ORIGEM'].tolist())
                        if 'SOLICITADO' in origens and 'ATENDIDO' in origens:
                            stats_cruzados += 1
                        elif 'SOLICITADO' in origens:
                            stats_so_sol += 1
                        else:
                            stats_so_at += 1

                        resultado_cruzado.append(mestre)

                    df_final = pd.DataFrame(resultado_cruzado)

                    # Limpa colunas auxiliares e reordena
                    cols_manter = CABECALHO_PADRAO
                    for col in cols_manter:
                        if col not in df_final.columns: df_final[col] = ''
                    df_final = df_final[cols_manter]
                    df_final = df_final.fillna('').astype(str)
                    df_final = df_final.replace(['nan','NaN','None','<NA>'], '')
                    for col in df_final.columns:
                        df_final[col] = df_final[col].apply(
                            lambda x: x.replace('\x00','').strip() if isinstance(x,str) else x
                        )

                    buf_final = io.BytesIO()
                    with pd.ExcelWriter(buf_final, engine='openpyxl') as writer:
                        df_final.to_excel(writer, index=False, sheet_name='BASE_UNIFICADA')
                        ws = writer.sheets['BASE_UNIFICADA']
                        for cell in ws[1]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="1e3d3a", end_color="1e3d3a", fill_type="solid")
                        for col in ws.columns:
                            ws.column_dimensions[col[0].column_letter].width = 22
                        ws.freeze_panes = 'A2'
                    buf_final.seek(0)
                    st.session_state['base_cruzada'] = buf_final.getvalue()
                    st.success(
                        f"✅ Cruzamento concluído! "
                        f"{stats_cruzados} pacientes unificados (estavam nos dois) | "
                        f"{stats_so_sol} só em Solicitados | "
                        f"{stats_so_at} só em Atendidos (sem conduta → VERIFICAR na Tab 12) | "
                        f"Total: {len(df_final)} pacientes únicos."
                    )
            except Exception as e:
                st.error(f"Erro ao cruzar as bases: {e}")
    if 'base_cruzada' in st.session_state:
        st.download_button("📥 BAIXAR BASE UNIFICADA (EXCEL)", st.session_state['base_cruzada'], "HOVA_Solicitados_x_Atendidos.xlsx", mime="application/x-xlsx", type="primary", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

with tab12:
    st.markdown('<div class="master-card">', unsafe_allow_html=True)
    st.markdown("""<div class="premium-title">📋 CONDUTA — ORGANIZAÇÃO POR PROCEDIMENTO</div>""", unsafe_allow_html=True)
    st.info("💡 **Fluxo:** Doctors → Solicitados x Atendidos → aqui. Suba a planilha unificada e o sistema classifica cada paciente na aba certa.")

    # ─── DICIONÁRIO DE ABAS ──────────────────────────────────────────
    ABAS_CONDUTA = {
        'CONSULTAS ESPECIALIZADAS': {
            'cor': '0F6E56', 'cor_linha': 'E1F5EE',
            'palavras': ['CONSULTA ANUAL','CONSULTA SEMESTRAL','CONSULTA TRIMESTRAL',
                         'CONSULTA BIMESTRAL','CONSULTA EM ','NOVA CONSULTA',
                         'SEGUIMENTO ANUAL','SEGUIMENTO SEMESTRAL','CONTROLE ANUAL',
                         'SEGUIMENTO EM ','ACOMPANHAMENTO ANUAL','ACOMPANHAMENTO EM',
                         'CONTROLE SEMESTRAL','CONTROLE TRIMESTRAL','CONTROLE BIMESTRAL',
                         'CONTROLE EM ','CONSULTA DE PIO','CONSULTA  45','CONSULTA 2 MESES'],
            'conduta_exata': ['C','C/MR','MR/C','C/MR/US','US/MR/C','C/US','C/US/MR'],
            'siglas_puras': [],
        },
        'EXAMES DIAGNOSE': {
            'cor': '185FA5', 'cor_linha': 'E6F1FB',
            'palavras': ['RETINOGRAFIA DIGITAL','MAPEAMENTO DE RETINA','PAQUIMETRIA',
                         'TOPOGRAFIA','MICROSCOPIA ESPECULAR','BIOMETRIA ULTRASSONICA',
                         'ULTRASSONOGRAFIA','GONIOSCOPIA','CURVA DIARIA','PENTACAM',
                         'CHECK-UP DE GLAUCOMA','CHECK UP DE GLAUCOMA','CHECKUP DE GLAUCOMA',
                         'TOMOGRAFIA DE COERENCIA','TOMOGRAFIA DO NERVO',
                         'ANGIOFLUORESCEINOGRAFIA','ECOBIOMETRIA','ECOGRAFIA',
                         'CAMPIMETRIA','CAMPO VISUAL'],
            'siglas_puras': ['AGF','OCT','CVC','RD','MR','PAQ','TOPO','MICRO','BIO','US',
                             'GONIO','CDPO','PTC','AV','AGF/OCT','CVC/PAQ/RD/OCT',
                             'CVC/PAQ/RD','RD/OCT','CVC/RD/PAQ','OCT/CVC/RD/PAQ',
                             'RD/PAQ/OCT/CVC','TOPO/MICRO/PAQ','TOPO/PAQ','TOPO/PAQ/MICRO',
                             'PAQ/RD/OCT','FOTOTRAB','GONIO/MR/R','MR/R','R/MR',
                             'BIO/TOPO/MICRO','BIO/TOPO/PAQ','OCT/RD/TOPO',
                             'CVC/RD/OCT/PAQ','PAQ/TOPO/MICRO'],
            'conduta_exata': [],
        },
        'LUZ PULSADA': {
            'cor': '9B6B00', 'cor_linha': 'FFF3CD',
            'palavras': ['LUZ PULSADA','LUZ INTENSA PULSADA','REALIZADO LUZ'],
            'siglas_puras': ['LP'], 'conduta_exata': [],
        },
        'MEIBOMIOGRAFIA': {
            'cor': 'B35A00', 'cor_linha': 'FCE4D6',
            'palavras': ['MEIBOMIOGRAFIA','MEIBOGRAFIA','TESTE DE SCHIRMER','TESTE SCHIRMER'],
            'siglas_puras': ['MBG','MEIBO','TSC'], 'conduta_exata': [],
        },
        'LENTES DE CONTATO': {
            'cor': '6B3FA0', 'cor_linha': 'EDE7F6',
            'palavras': ['BUSCA DE LENTE','COMPRA LENTE','TESTE DE LENTE',
                         'LENTE GELATINOSA','LENTE RIGIDA','LENTE TÓRICA',
                         'LENTE TERAPEUTICA','LENTE MULTIFOCAL','LENTE ESCLERAL',
                         'REVISAO DE LENTE','REVISÃO DE LENTE','RETORNO DE LENTE'],
            'siglas_puras': ['BL','TL','CLGN','CLGP','CLMF','CLR','CLT','TLESCL',
                             'TLR','CLTERAPEUTICA','RL','RO'],
            'conduta_exata': [],
        },
        'FACO — CATARATA': {
            'cor': '8B1A00', 'cor_linha': 'FFEBEE',
            'palavras': ['FACO COM IMPLANTE DE LIO','FACO + LIO','FACO+LIO',
                         'FACOEMULSIFICAÇÃO','CIRURGIA DE CATARATA',
                         'AO SETOR DE CATARATA','IMPLANTE DE LENTE DE CONTATO INTRAOCULAR',
                         'INDICO FACO','AVALIAÇÃO DE FACO','INDICAR FACO'],
            'siglas_puras': ['ILIO'], 'conduta_exata': [],
            'exclusoes': ['LASIK','PRK','IC_REFRATIVA','ORÇ_REFRATIVA','FACORREFRATIVA'],
        },
        'FACO REFRATIVA': {
            'cor': 'A0290E', 'cor_linha': 'FFF0EC',
            'palavras': ['LASIK','PRK','IC_REFRATIVA','ORÇ_REFRATIVA','CIRURGIA REFRATIVA',
                         'FACORREFRATIVA','PRK COM MITOMICINA'],
            'siglas_puras': [], 'conduta_exata': [],
        },
        'RETINA': {
            'cor': '0D5C2E', 'cor_linha': 'E8F5E9',
            'palavras': ['VVPP','VITRECTOMIA','INJEÇÃO INTRAVITREA','INJECAO INTRAVITREA',
                         'IIV','ANTI-VEGF','ANTIVEGF','ANTI VEGF','EYLIA','AVASTIM',
                         'FOTOCOAGULAÇÃO','FOTOCOAGULACAO','PANFOTO','FC LASER',
                         'IC_VITRE','ORÇ_RETINA','INDICO VVPP','INDICO IIV',
                         'INDICO 1 CICLO DE IIV','INDICO NOVA IIV','INDICO 1 IIV',
                         'ANTIANGIOGÊNICO','ANTIANGIOGENICO'],
            'siglas_puras': ['VVPP','FOTO'], 'conduta_exata': [],
        },
        'GLAUCOMA CIRÚRGICO': {
            'cor': '1A5276', 'cor_linha': 'D6EAF8',
            'palavras': ['CICLOFOTOCOAGULAÇÃO','CICLOFOTOCOAGULACAO','CICLOFOTO SLOW',
                         'CICLO SLOW','FOTOTRAB','SLT AO','SLT OD','SLT OE',
                         'TRABECULECTOMIA','TREC OD','TREC OE','IRIDECTOMIA',
                         'IC_CICLO','ORÇ_CICLO','ORÇ_GLAUCOMA','CIRURGIA DE GLAUCOMA',
                         'CIRURGIA FISTULIZANTE','AGULHAMENTO','INDICO SLT','REINDICO SLT',
                         'INDICO CICLO','INDICO CICLOFOTO','INDICO TREC',
                         'INDICO FOTOTRAB','INDICO IRI'],
            'siglas_puras': ['CICLOFOTO','SLT','IRI','IRIDO'], 'conduta_exata': [],
        },
        'CERATOCONE — ANEL': {
            'cor': '145A32', 'cor_linha': 'D5F5E3',
            'palavras': ['IMPLANTE DE ANEL','ANEL OD','ANEL OE','ANEL AO','IC_ANEL',
                         'ORÇ_ANEL','EXPLANTE DE ANEL','CROSS-LINK','CROSSLINK','CXL',
                         'INDICO ANEL'],
            'siglas_puras': [], 'conduta_exata': [],
        },
        'TRANSPLANTE DE CÓRNEA': {
            'cor': '1A5C32', 'cor_linha': 'D0F0E0',
            'palavras': ['TRANSPLANTE DE CORNEA','TRANSPLANTE DE CÓRNEA',
                         'TRANSPLANTE ENDOTELIAL','CERATOPLASTIA'],
            'siglas_puras': [], 'conduta_exata': [],
        },
        'YAG LASER': {
            'cor': '7D6608', 'cor_linha': 'FEF9E7',
            'palavras': ['CAPSULOTOMIA YAG','CAPSULOTOMIA AO','CAPSULOTOMIA OD',
                         'CAPSULOTOMIA OE','YAG LASER','IRIDOTOMIA A LASER'],
            'siglas_puras': ['YAG'], 'conduta_exata': [],
        },
        
        'PTERÍGIO': {
            'cor': '196F3D', 'cor_linha': 'EAFAF1',
            'palavras': ['PTERIGIO COM TCLC','PTERÍGIO COM TCLC','EXERESE DE PTERÍGIO',
                         'EXERESE DE PTERIGIO','EXERESE PTERÍGIO','EXERESE PTERIGIO',
                         'PINGUECULA COM TCLC','PINGUÉCULA COM TCLC','IC_PTERIGIO',
                         'ORÇ_PTERIGIO','RETIRADA PTERIGEO','PTERIGIOCOMTCLC'],
            'siglas_puras': [], 'conduta_exata': [],
        },
        'CALÁZIO': {
            'cor': '1E6B40', 'cor_linha': 'E9F7EF',
            'palavras': ['CIRURGIA DE CALAZIO','CIRURGIA DE CALÁZIO','EXERESE DE CALAZIO',
                         'EXERESE DE CALÁZIO','EXERESE DE TUMOR DE PALPEBRA',
                         'EXERESE DE TUMOR PALPEBRAL','TUMOR DE PALPEBRA',
                         'TUMOR PALPEBRAL','EXERESE TUMOR CONJUNTIVAL',
                         'EXERESE DE TUMOR CONJUNTIVAL','TU DE PALPEBRA',
                         'CIRURGIA DE TUMOR'],
            'siglas_puras': [], 'conduta_exata': [],
        },
        'OCULOPLÁSTICA': {
            'cor': '283593', 'cor_linha': 'E8EAF6',
            'palavras': ['BLEFAROPLASTIA','OCULOPLASTICA','OCULOPLÁSTICA',
                         'SONDAGEM DE VIAS LACRIMAIS','SONDAGEM DE VIIAS',
                         'TARSAL STRIP','PLASTICA OCULAR','PLÁSTICA OCULAR',
                         'PTOSE PALPEBRAL','AO SETOR DE OCULOPLASTICA',
                         'CIRURGIÃO PLASTICA','AVALIAÇÃO PALPEBRA',
                         'ENCAMINHO PARA OCULOPLASTICA'],
            'siglas_puras': [], 'conduta_exata': [],
        },
        'ESTRABISMO': {
            'cor': '4A148C', 'cor_linha': 'F3E5F5',
            'palavras': ['CIRURGIA DE ESTRABISMO','CIRURGIA PARA CORREÇÃO DO ESTRABISMO',
                         'ORÇ_ESTRABISMO','ORÇ_ EXTRABISMO','EXAME ORTOPTICO',
                         'EXAME ORTÓPTICO','AO SETOR DE ESTRABISMO'],
            'siglas_puras': [], 'conduta_exata': [],
        },
    }

    MAP_PRAZO_CONDUTA = {
        365:'CONSULTA ANUAL', 180:'CONSULTA SEMESTRAL', 90:'CONSULTA TRIMESTRAL',
        60:'CONSULTA BIMESTRAL', 120:'CONSULTA EM 4 MESES', 270:'CONSULTA EM 9 MESES',
        150:'CONSULTA EM 5 MESES', 45:'NOVA CONSULTA EM 45 DIAS',
        30:'NOVA CONSULTA EM 30 DIAS', 21:'NOVA CONSULTA EM 3 SEMANAS',
        14:'NOVA CONSULTA EM 2 SEMANAS', 10:'NOVA CONSULTA EM 10 DIAS',
        7:'NOVA CONSULTA EM 1 SEMANA',
    }

    COLUNAS_CONDUTA = ['NUM','NOME','TELEFONE','TEL.ADIC','CONV','DATA VISITA','MEDICO',
                       'CONDUTA ORIGINAL','CONDUTA MÉDICA LIMPA','CONDUTA PADRONIZADA',
                       'PRÓXIMA CALCULADA','ALERTA',
                       'MSG2024','MSG2025','MSG2026','PACIENTE GLAUCOMA',
                       'EMAIL','CPF','PROFISSAO','DATA NASC']

    LARGURAS_CONDUTA = {
        'NUM':8,'NOME':32,'TELEFONE':16,'TEL.ADIC':18,'CONV':22,'DATA VISITA':12,
        'MEDICO':16,'CONDUTA ORIGINAL':14,'CONDUTA MÉDICA LIMPA':42,
        'CONDUTA PADRONIZADA':24,'PRÓXIMA CALCULADA':14,'ALERTA':32,
        'MSG2024':22,'MSG2025':22,'MSG2026':22,'PACIENTE GLAUCOMA':10,
        'EMAIL':28,'CPF':16,'PROFISSAO':16,'DATA NASC':12,
    }

    def _norm_c(t):
        return '' if not t or str(t).lower() in ['nan','none',''] else str(t).upper().strip()

    # Regra por médico — quando CONDUTA ORIGINAL é sigla genérica (RC, PO, AC, R...)
    _REGRA_MEDICO = {
        'ALTAIR':   'FACO — CATARATA',
        'DENISE':   'RETINA',
        'GABRIELL': 'RETINA',
        'GABRIEL':  'RETINA',
        'GUSTAVO':  'GLAUCOMA CIRÚRGICO',
        'FELIPE':   'LUZ PULSADA',
        'VERA':     'ESTRABISMO',
        'EXAMES':   'EXAMES DIAGNOSE',
    }
    _SIGLAS_GENERICAS = {'RC','PO','R','AC','CO','RE','RIC','PAG.A','REP.EX',
                         'IC','IC_FACO','IC_VITRE','IC_CICLO','IC_PTERIGIO',
                         'IC_CALAZIO','IC_ ANTIVEG','ILIO','RO','RL',
                         'FOTO','MR','US','OCT','CVC','AGF','RD','PAQ',
                         'TOPO','MICRO','BIO','GONIO','CDPO','PTC','AV',
                         'IOL/MICRO/PTC','BIO/TOPO/MICRO','BIO/MICRO/TOPO',
                         'IOL/MICRO','IOL/TOPO/MICRO','MICRO/TOPO/OCT',
                         'TOPO/MICRO/PAQ','PAQ/RD/OCT','RD/OCT',
                         'CVC/PAQ/RD/OCT','CVC/PAQ/RD','MR/US/C','US/MR/C',
                         'GONIO/R','BL','RL','CLR','CLGN','CLMF','CLT',
                         'CLTERAPEUTICA','TL','TLR','TLESCL','RIC','MD',
                         'CO','EEX','ORÇ','ORÇ_RETINA','ORÇ_GLAUCOMA',
                         'ORÇ_PTERIGIO','CI_ALTAIR','CI_DENISE','CI_GABRIEL',
                         'CI_GABRIELL','CI_GUSTAVO','CI_MARILUCI','CI_FELIPE',
                         'CI_VERA','PAG.A','REP.EX','nan',''}

    
    def _classificar_conduta(conduta_col, proxima_col, medico_col=''):
        c = _norm_c(conduta_col)
        p = _norm_c(proxima_col)
        m = _norm_c(medico_col)
        p_up = p.upper()

        # ── DETECÇÃO DE PROCEDIMENTOS ─────────────────────────────
        def _tem_faco(t): return any(x in t for x in [
            'FACO','LIO','CATARATA','FACOEMULSIFICAÇÃO','FACO COM IMPLANTE',
            'AO SETOR DE CATARATA','INDICO FACO','PROPEDEUTICA DE CATARATA'])

        def _tem_retina(t): return any(x in t for x in [
            'IIV','ANTIVEGF','ANTI-VEGF','ANTI VEGF','VVPP','FOTOCOAGULAÇÃO',
            'FOTOCOAGULACAO','EYLIA','AVASTIM','INTRAVITREA','ANTIANGIOGÊNICO',
            'PANFOTO','FC LASER','VITRECTOMIA','INDICO IIV','INJEÇÃO INTRAVITREA'])

        def _tem_glaucoma(t): return any(x in t for x in [
            'CICLOFOTOCOAGULAÇÃO','CICLOFOTOCOAGULACAO','CICLOFOTO','TRABECULECT',
            'FOTOTRAB','SLT OD','SLT OE','SLT AO','INDICO SLT','REINDICO SLT',
            'IRIDECTOMIA','CIRURGIA DE GLAUCOMA','TREC OD','TREC OE','AGULHAMENTO',
            'INDICO IRIDO','INDICO IRI'])

        def _tem_pterigio(t): return any(x in t for x in [
            'PTERIGIO COM TCLC','EXERESE DE PTERIGIO','EXERESE PTERIGIO',
            'PINGUECULA COM TCLC','PTERÍGIO COM TCLC',
            'AVALIAÇÃO CIRURGICA COM DRA MARILUCI'])

        def _tem_calazio(t): return any(x in t for x in [
            'CALAZIO','CALÁZIO','TUMOR PALPEBRAL','EXERESE DE TUMOR',
            'TU DE PALPEBRA','XANTELASMA'])

        def _tem_yag(t): return any(x in t for x in [
            'YAG LASER','CAPSULOTOMIA YAG','CAPSULOTOMIA OD','CAPSULOTOMIA OE',
            'CAPSULOTOMIA AO','YAG OD','YAG OE','YAG AO',
            'IRIDOTOMIA A LASER','INDICO YAG'])

        def _tem_anel(t): return any(x in t for x in [
            'IMPLANTE DE ANEL','ANEL OD','ANEL OE','ANEL AO','CROSSLINK',
            'CXL','CERATOCONE','CROSS LINK','CROSS-LINK','INDICO ANEL'])

        def _tem_transplante(t): return any(x in t for x in [
            'TRANSPLANTE DE CORNEA','TRANSPLANTE DE CÓRNEA',
            'CERATOPLASTIA','TRANSPLANTE ENDOTELIAL'])

        def _tem_lente(t): return any(x in t for x in [
            'LENTES TORICAS','LENTES TÓRICAS','TESTE DE LENTE','LENTE RIGIDA',
            'LENTE ESCLERAL','LENTE TERAPEUTICA','LENTE MULTIFOCAL','LENTE GELATINOSA',
            'LENTES GELATINOSA','LENTES MULTIFOCAIS','LENTES GELATINOSAS',
            'NOVAS LENTES','BUSCA DE LENTE','COMPRA LENTE','LENTE MILLENNIUM',
            'LENTE HR','MILLENNIUM XC','TESTE LC','LCG','LENTE DE TESTE',
            'LENTE TORICA','LC TORICA','LC GELATINOSA','LENTE RIGIDA'])

        def _tem_lp(t): return any(x in t for x in [
            'LUZ PULSADA','LUZ INTENSA PULSADA'])

        def _tem_oculo(t): return any(x in t for x in [
            'BLEFAROPLASTIA','OCULOPLASTICA','OCULOPLÁSTICA','PTOSE PALPEBRAL',
            'SONDAGEM DE VIAS','TARSAL STRIP','AO SETOR DE OCULOPLASTICA',
            'AO SETOR DE OCULOPLÁSTICA','AVALIAÇÃO CIRURGICA DE PALPEBRA',
            'AVALIAÇÃO PALPEBRAL','ENCAMINHO PARA OCULOPLASTICA'])

        def _tem_estrabismo(t): return any(x in t for x in [
            'ESTRABISMO','ORTOPTICO','ORTÓPTICO','CIRURGIA DE ESTRABISMO'])

        def _tem_meibo(t): return any(x in t for x in [
            'MEIBOMIOGRAFIA','MEIBOGRAFIA','SCHIRMER'])

        # ── EXAMES: siglas E nomes por extenso ───────────────────
        def _tem_exames(t): return any(x in t for x in [
            # check-up / propedêutica
            'CHECK-UP','CHECK UP','CHECKUP','PROPEDEUTICA','PROPEDÊUTICA',
            'SOLICITO PROP','SOLICITO AV GLAUC','AV GLAUC',
            # nomes por extenso
            'CAMPIMETRIA','CAMPO VISUAL','ANGIOFLUORESCEINOGRAFIA','TOMOGRAFIA DE COERENCIA',
            'TOMOGRAFIA DE COERÊNCIA','RETINOGRAFIA DIGITAL','RETINOGRAFIA FLUORESCENTE',
            'MAPEAMENTO DE RETINA','ULTRASSONOGRAFIA','ECOGRAFIA','BIOMETRIA',
            'MICROSCOPIA ESPECULAR','GONIOSCOPIA','PAQUIMETRIA','TOPOGRAFIA',
            'PENTACAM','CURVA DIARIA','CURVA DIÁRIA','TOMOGRAFIA DO NERVO',
            # siglas por extenso
            'OCT MACULA','OCT PAPILA','OCT PAPILAS','OCT DE MACULA',
            'ECO A','TOPO E PAQUI','TOPO + PAQUI','TOPO,PAQUI',
            'RD + CVC','CVC E VER','CVC E OCT','REPETIR CVC',
            'NOVO CVC','NOVO OCT','MR + OCT','MR E OCT',
            # abreviações
            'ACUIDADE VISUAL - PAM'])

        # ── RETORNO na conduta médica → sempre VERIFICAR ─────────
        def _tem_retorno_textual(t): return any(x in t for x in [
            'RETORNO EM','RETORNEO','RETRONEO','RETOTNO','RETORNO DE',
            'RETORNO PARA','RETORNO REFRAÇ','RETORNO REFRAC',
            'RETORNO 15°','RETORNO 30°','RETORNO COM',
            'REVISÃO DE ','REVISAO DE '])

        # ── PRAZO PURO de consulta (sem procedimento) ────────────
        def _tem_prazo(t): return any(x in t for x in [
            'CONSULTA ANUAL','CONSULTA SEMESTRAL','CONSULTA TRIMESTRAL',
            'CONSULTA BIMESTRAL','NOVA CONSULTA','SEGUIMENTO ANUAL',
            'SEGUIMENTO SEMESTRAL','SEGUIMENTO TRIMESTRAL','CONTROLE ANUAL',
            'CONSULTA EM ','SEGUIMENTO EM ','CONTROLE EM ','ALTA',
            ' MESES',' SEMANAS',' SEMANA',' DIAS','VOLTAR EM',
            'NOVA CONSULTA EM '])

        # ── 0. NAN / VAZIO → VERIFICAR ───────────────────────────
        if not c or c == 'NAN':
            return '⚠️ VERIFICAR', '❓ CONDUTA ORIGINAL vazia ou NAN'

        # ── 1. IC_ MAP ────────────────────────────────────────────
        _IC_MAP = {
            'IC_FACO':'FACO — CATARATA','IC FACO':'FACO — CATARATA',
            'IC_VITRE':'RETINA','IC_ANTIVEG':'RETINA','IC_ ANTIVEG':'RETINA',
            'IC_CICLO':'GLAUCOMA CIRÚRGICO',
            'IC_PTERIGIO':'PTERÍGIO','IC_CALAZIO':'CALÁZIO',
            'IC_REFRATIVA':'FACO REFRATIVA',
            # IC_ANEL: se conduta médica tem propedêutica → EXAMES; senão → ANEL
            'IC_ANEL': None,  # tratado abaixo
        }
        if c == 'IC_ANEL':
            if p and 'PROPEDEUTICA' in p_up:
                return 'EXAMES DIAGNOSE', ''
            return 'CERATOCONE — ANEL', ''
        if c in _IC_MAP and _IC_MAP[c] is not None:
            return _IC_MAP[c], ''

        # ── 2. ORÇ_ MAP ───────────────────────────────────────────
        _ORC_MAP = {
            'ORÇ_RETINA':'RETINA','ORÇ_FACO':'FACO — CATARATA',
            'ORÇ_GLAUCOMA':'GLAUCOMA CIRÚRGICO','ORÇ_TRAB':'GLAUCOMA CIRÚRGICO',
            'ORÇ_CICLO':'GLAUCOMA CIRÚRGICO','ORÇ_PTERIGIO':'PTERÍGIO',
            'ORÇ_CALAZIO':'CALÁZIO','ORÇ_ANEL':'CERATOCONE — ANEL',
            'ORÇ_REFRATIVA':'FACO REFRATIVA','ORÇ_VVPP':'RETINA',
        }
        if c in _ORC_MAP:
            return _ORC_MAP[c], ''

        # ORÇ genérico sem sufixo → herda do médico
        if c == 'ORÇ':
            if 'GUSTAVO' in m:  return 'GLAUCOMA CIRÚRGICO', ''
            if 'ALTAIR'  in m:  return 'FACO — CATARATA', ''
            if any(x in m for x in ['DENISE','GABRIELL','GABRIEL']): return 'RETINA', ''
            if 'MARILUCI' in m: return 'PTERÍGIO', ''
            return '⚠️ VERIFICAR', '❓ ORÇ sem sufixo e sem médico conhecido'

        # ── 3. LENTES siglas ──────────────────────────────────────
        _LENTES_SIGLAS = {'BL','TL','CLGN','CLGP','CLMF','CLR','CLT','TLESCL',
                          'TLR','CLTERAPEUTICA','RL','RO','CLT/CLGN','CLR/RPG'}
        if c in _LENTES_SIGLAS:
            return 'LENTES DE CONTATO', ''

        # ── 4. EXAMES siglas puras ────────────────────────────────
        _EXAMES_SIGLAS = {
            'OCT','CVC','AGF','RD','MR','PAQ','TOPO','MICRO','BIO','US','GONIO',
            'CDPO','PTC','AV','AGF/OCT','CVC/PAQ/RD/OCT','RD/OCT','PAQ/RD/OCT',
            'BIO/TOPO/MICRO','BIO/MICRO/TOPO','IOL/MICRO/PTC','IOL/MICRO',
            'IOL/TOPO/MICRO','MICRO/TOPO/OCT','TOPO/MICRO/PAQ','MR/R','MR/US',
            'CVC/PAQ/RD','TOPO/PAQ','ANGIO-OCT','GONIO/R','TOPO/PAQ/MICRO',
            'CVC/RD/PAQ','OCT/CVC/RD/PAQ','BIO/TOPO/PAQ','OCT/RD/TOPO',
            'PAQ/TOPO/MICRO','RD/PAQ/OCT/CVC','MR/US/C','US/MR/C','C/MR/US',
            'C/MR','MR/C','IOL','IOL/PTC','PTC/PAQ/MICRO/O',
        }
        if c in _EXAMES_SIGLAS:
            if _tem_retina(p_up):   return 'RETINA', ''
            if _tem_glaucoma(p_up): return 'GLAUCOMA CIRÚRGICO', ''
            return 'EXAMES DIAGNOSE', ''

        # ── 5. YAG sigla pura ─────────────────────────────────────
        if c == 'YAG':
            return 'YAG LASER', ''

        # ── 6. FOTO → RETINA ──────────────────────────────────────
        if c == 'FOTO' or c.startswith('FOTO/'):
            return 'RETINA', ''

        # ── 7. ILIO → FACO + CONFERIR ─────────────────────────────
        if c == 'ILIO':
            return 'FACO — CATARATA', 'PRÓXIMA: CONFERIR MANUALMENTE'

        # ── 8. FOTOTRAB: se conduta médica tem prazo → CONSULTAS ──
        if c == 'FOTOTRAB':
            if p and _tem_prazo(p_up) and not _tem_glaucoma(p_up):
                return 'CONSULTAS ESPECIALIZADAS', ''
            return 'GLAUCOMA CIRÚRGICO', ''

        # ── 9. IRI: se conduta médica tem RETORNO → VERIFICAR ─────
        if c in {'IRI','IRIDO'}:
            if p and _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ IRI com RETORNO na conduta — confirmar'
            return 'GLAUCOMA CIRÚRGICO', ''

        # ── 10. TFD / siglas administrativas → VERIFICAR ──────────
        if c in {'TFD','CO','MD','EEX','CE','RP','RPG','TX','EXLAB'}:
            return '⚠️ VERIFICAR', f'❓ {c} — sigla administrativa, sem classificação clínica'

        # ── 11. R / RC / PO / AC e similares ──────────────────────
        _SIGLAS_RETORNO = {'R','RC','PO','AC','RE','RIC','PAG.A','REP.EX'}
        if c in _SIGLAS_RETORNO:

            # PO do ALTAIR → sempre FACO
            if c == 'PO' and 'ALTAIR' in m:
                return 'FACO — CATARATA', ''

            # Sem conduta médica → VERIFICAR
            if not p:
                return '⚠️ VERIFICAR', f'❓ {c} sem conduta médica — conferir com médico'

            # Conduta médica com "RETORNO" textual → VERIFICAR
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', f'❓ {c} com RETORNO na conduta — confirmar próxima'

            # Lê procedimentos na conduta médica
            if _tem_faco(p_up):      return 'FACO — CATARATA', ''
            if _tem_retina(p_up):    return 'RETINA', ''
            if _tem_glaucoma(p_up):  return 'GLAUCOMA CIRÚRGICO', ''
            if _tem_pterigio(p_up):  return 'PTERÍGIO', ''
            if _tem_calazio(p_up):   return 'CALÁZIO', ''
            if _tem_yag(p_up):       return 'YAG LASER', ''
            if _tem_anel(p_up):      return 'CERATOCONE — ANEL', ''
            if _tem_lente(p_up):     return 'LENTES DE CONTATO', ''
            if _tem_lp(p_up):        return 'LUZ PULSADA', ''

            # Exames mencionados na conduta médica → EXAMES
            if _tem_exames(p_up):    return 'EXAMES DIAGNOSE', ''

            # Prazo puro → aba do médico (acompanhamento pós-procedimento)
            if _tem_prazo(p_up):
                if any(x in m for x in ['DENISE','GABRIELL','GABRIEL']): return 'CONSULTAS ESPECIALIZADAS', ''
                if 'GUSTAVO'  in m: return 'CONSULTAS ESPECIALIZADAS', ''
                if 'ALTAIR'   in m: return 'CONSULTAS ESPECIALIZADAS', ''
                if 'MARILUCI' in m: return 'CONSULTAS ESPECIALIZADAS', ''
                if 'FELIPE'   in m: return 'CONSULTAS ESPECIALIZADAS', ''
                return 'CONSULTAS ESPECIALIZADAS', ''

            return '⚠️ VERIFICAR', f'❓ {c} sem padrão identificado'

        # ── 12. C → lê conduta médica ─────────────────────────────
        if c == 'C':

            # Sem conduta médica → depende do médico
            if not p:
                # Mariluci sem conduta = VERIFICAR (não assume pterígio)
                if 'MARILUCI' in m:
                    return '⚠️ VERIFICAR', '❓ MARILUCI sem conduta médica'
                # Outros médicos conhecidos sem conduta = CONSULTAS
                if any(x in m for x in ['ALTAIR','GUSTAVO','DENISE','GABRIELL',
                                         'GABRIEL','FELIPE','VERA']):
                    return 'CONSULTAS ESPECIALIZADAS', ''
                return '⚠️ VERIFICAR', '❓ C sem conduta médica e médico desconhecido'

            # Com conduta médica — RETORNO textual → VERIFICAR
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ C com RETORNO na conduta — confirmar próxima'

            # Lê procedimentos (ordem de prioridade)
            if _tem_faco(p_up):        return 'FACO — CATARATA', ''
            if _tem_retina(p_up):      return 'RETINA', ''
            if _tem_glaucoma(p_up):    return 'GLAUCOMA CIRÚRGICO', ''
            if _tem_pterigio(p_up):    return 'PTERÍGIO', ''
            if _tem_calazio(p_up):     return 'CALÁZIO', ''
            if _tem_yag(p_up):         return 'YAG LASER', ''
            if _tem_anel(p_up):        return 'CERATOCONE — ANEL', ''
            if _tem_transplante(p_up): return 'TRANSPLANTE DE CÓRNEA', ''
            if _tem_oculo(p_up):       return 'OCULOPLÁSTICA', ''
            if _tem_estrabismo(p_up):  return 'ESTRABISMO', ''
            if _tem_meibo(p_up):       return 'MEIBOMIOGRAFIA', ''
            if _tem_lente(p_up):       return 'LENTES DE CONTATO', ''
            if _tem_lp(p_up):          return 'LUZ PULSADA', ''

            # Exames por extenso ou sigla → EXAMES DIAGNOSE
            if _tem_exames(p_up):      return 'EXAMES DIAGNOSE', ''

            # Glaucoma check-up
            if any(x in p_up for x in ['AV GLAUC','CHECK-UP DE GLAUCOMA',
                                        'CHECKUP DE GLAUCOMA','CHECK UP DE GLAUCOMA',
                                        'PROPEDEUTICA DE GLAUCOMA']):
                return 'GLAUCOMA CIRÚRGICO', ''

            # Conduta médica vaga / só medicação / não identificada → VERIFICAR
            palavras_vagas = [
                'MANTER','ACRESCENTO','RETIRO','BLEPHAGEL','OCUPRESS','GLAUB',
                'ORIENTO','PRESCREVO','COMPRESSA','CORCUNDA','RIGIDO',
                'CONTROLE METABOLICO','DESCOLAMENTO','ENCAMINHO PACIENTE A BH'
            ]
            if any(x in p_up for x in palavras_vagas):
                return '⚠️ VERIFICAR', '❓ Conduta médica vaga — conferir manualmente'

            # Prazo puro sem procedimento identificado → CONSULTAS
            if _tem_prazo(p_up):
                return 'CONSULTAS ESPECIALIZADAS', ''

            # C com conduta que não se encaixou → VERIFICAR
            return '⚠️ VERIFICAR', '❓ C com conduta não identificada — conferir'

        # ── 13. MARILUCI ──────────────────────────────────────────
        if 'MARILUCI' in m:
            # Sem conduta → VERIFICAR (nunca assumir pterígio sem info)
            if not p:
                return '⚠️ VERIFICAR', '❓ MARILUCI sem conduta médica'
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ MARILUCI com RETORNO — confirmar'
            if _tem_calazio(p_up):  return 'CALÁZIO', ''
            if _tem_pterigio(p_up) or 'PTERIGIO' in p_up or 'PTERÍGIO' in p_up:
                return 'PTERÍGIO', ''
            if _tem_exames(p_up):   return 'EXAMES DIAGNOSE', ''
            if _tem_lente(p_up):    return 'LENTES DE CONTATO', ''
            if _tem_prazo(p_up):    return 'CONSULTAS ESPECIALIZADAS', ''
            return '⚠️ VERIFICAR', '❓ MARILUCI sem padrão identificado'

        # ── 14. ENFERMAG ──────────────────────────────────────────
        if 'ENFERMAG' in m:
            if _tem_faco(p_up):     return 'FACO — CATARATA', ''
            if _tem_retina(p_up):   return 'RETINA', ''
            if _tem_glaucoma(p_up): return 'GLAUCOMA CIRÚRGICO', ''
            if _tem_pterigio(p_up): return 'PTERÍGIO', ''
            if _tem_calazio(p_up):  return 'CALÁZIO', ''
            if _tem_anel(p_up):     return 'CERATOCONE — ANEL', ''
            if _tem_lp(p_up):       return 'LUZ PULSADA', ''
            if _tem_exames(p_up):   return 'EXAMES DIAGNOSE', ''
            return '⚠️ VERIFICAR', '❓ ENFERMAGEM sem padrão identificado'

        # ── 15. VERA ──────────────────────────────────────────────
        if 'VERA' in m:
            return 'ESTRABISMO', ''

        # ── 16. FELIPE ────────────────────────────────────────────
        if 'FELIPE' in m:
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ FELIPE com RETORNO — confirmar'
            if _tem_exames(p_up): return 'EXAMES DIAGNOSE', ''
            if _tem_lente(p_up):  return 'LENTES DE CONTATO', ''
            return 'LUZ PULSADA', ''

        # ── 17. GUSTAVO ───────────────────────────────────────────
        if 'GUSTAVO' in m:
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ GUSTAVO com RETORNO — confirmar'
            if _tem_faco(p_up):    return 'FACO — CATARATA', ''
            if _tem_retina(p_up):  return 'RETINA', ''
            if _tem_glaucoma(p_up): return 'GLAUCOMA CIRÚRGICO', ''
            if _tem_exames(p_up):  return 'EXAMES DIAGNOSE', ''
            if _tem_lente(p_up):   return 'LENTES DE CONTATO', ''
            if _tem_prazo(p_up):   return 'CONSULTAS ESPECIALIZADAS', ''
            # Sem info identificável → VERIFICAR (não assume glaucoma)
            return '⚠️ VERIFICAR', '❓ GUSTAVO sem padrão identificado'

        # ── 18. ALTAIR ────────────────────────────────────────────
        if 'ALTAIR' in m:
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ ALTAIR com RETORNO — confirmar'
            if _tem_retina(p_up):    return 'RETINA', ''
            if _tem_glaucoma(p_up):  return 'GLAUCOMA CIRÚRGICO', ''
            if _tem_lente(p_up):     return 'LENTES DE CONTATO', ''
            if _tem_yag(p_up):       return 'YAG LASER', ''
            if _tem_anel(p_up):      return 'CERATOCONE — ANEL', ''
            if _tem_oculo(p_up):     return 'OCULOPLÁSTICA', ''
            if _tem_exames(p_up):    return 'EXAMES DIAGNOSE', ''
            if _tem_prazo(p_up):     return 'CONSULTAS ESPECIALIZADAS', ''
            if not p:                return 'CONSULTAS ESPECIALIZADAS', ''
            return 'FACO — CATARATA', ''

        # ── 19. DENISE / GABRIELL / GABRIEL ───────────────────────
        if any(x in m for x in ['DENISE','GABRIELL','GABRIEL']):
            if _tem_retorno_textual(p_up):
                return '⚠️ VERIFICAR', '❓ RETINA com RETORNO — confirmar'
            if _tem_faco(p_up):     return 'FACO — CATARATA', ''
            if _tem_glaucoma(p_up): return 'GLAUCOMA CIRÚRGICO', ''
            if _tem_pterigio(p_up): return 'PTERÍGIO', ''
            if _tem_lente(p_up):    return 'LENTES DE CONTATO', ''
            if _tem_yag(p_up):      return 'YAG LASER', ''
            if _tem_oculo(p_up):    return 'OCULOPLÁSTICA', ''
            if _tem_exames(p_up):   return 'EXAMES DIAGNOSE', ''
            if _tem_prazo(p_up):    return 'CONSULTAS ESPECIALIZADAS', ''
            if not p:               return 'CONSULTAS ESPECIALIZADAS', ''
            return 'RETINA', ''

        # ── 20. Sem médico conhecido ───────────────────────────────
        return '⚠️ VERIFICAR', f'❓ SEM REGRA: médico={m[:20]}'
    
    def _prazo_conduta(p_raw, c_raw):
        p = _norm_c(p_raw); c = _norm_c(c_raw)
        t = (p + ' ' + c).strip()
        for lx in ['RX OCULOS','RX ÓCULOS','MANTER ÓCULOS','ORIENTAÇÕES','LAUDO',
                   'MANTER MEDICAÇÃO','OCULOS A PEDIDO','ORIENTO','ENCAMINHO','MANTENHO']:
            t = t.replace(lx, '')
 
        # Data explícita no formato DD/MM → extrair como próxima direta
        import re as _re
        m_data = _re.search(r'EM\s+(\d{1,2}/\d{2})(?:\s*\()?', t)
        if m_data:
            return -1  # sinal especial: data explícita no texto
 
        M = [(365,['ANUAL','1 ANO','12 MES','365 DIAS']),
             (180,['SEMESTRAL','6 MES','180 DIAS']),
             (90, ['TRIMESTRAL','3 MES','90 DIAS']),
             (60, ['BIMESTRAL','2 MES','60 DIAS']),
             (120,['4 MES','4M ','120 DIAS','CONSULTA EM 4']),
             (270,['9 MES']), (150,['5 MES']),
             (45, ['45 DIAS','CONSULTA  45','CONSULTA 45']),
             (30, ['30 DIAS','1 MES ','1 MÊS ']),
             (21, ['3 SEMANAS','21 DIAS']),
             (14, ['2 SEMANAS','14 DIAS','EM 2 SEMANAS','NOVA FOTO EM 15','FOTO EM OE - 15',
                   'NOVA FOTO EM 10','FOTO EM 15','NOVA FOTO 15','TEROLAC 8/8H']),
             (10, ['10 DIAS']),
             (7,  ['1 SEMANA','7 DIAS','EM 7 DIAS','RETORNO EM 7'])]
        for d, kws in M:
            if any(k in t for k in kws): return d
        m = _re.search(r'EM\s+(\d+)\s*DIAS?', t)
        if m: return int(m.group(1))
        m = _re.search(r'EM\s+(\d+)\s*MES', t)
        if m: return int(m.group(1)) * 30
        if c in ['C','C/MR','MR/C','C/MR/US','US/MR/C'] and not p: return 365
        return None
 

    def _limpar_conduta_medica(texto):
        if not texto or str(texto).lower() in ['nan','none','']: return ''
        LIXO = ['RX OCULOS','RX ÓCULOS','MANTER ÓCULOS','MANTER OCULOS','ORIENTAÇÕES',
                'ORIENTACOES','LAUDO MÉDICO','MANTER MEDICAÇÃO','OCULOS A PEDIDO',
                'ÓCULOS A PEDIDO','LENTES DE CONTATO','MANTER LENTE','ENCAMINHO AO',
                'ENCAMINHO PARA','LACRIFILM','REGENCEL','PRUROK','FLUTINOL','ATROPINA',
                'BACTRIM','DIAMOX','DORZAL','SIMBRINZA','XALATAN','MANTENHO','PRESCREVO',
                'ORIENTO SOBRE','POSIÇÃO DE CABEÇA','CUIDADOS COM OLHO']
        linhas = []
        for l in str(texto).split('\n'):
            l = l.strip()
            if l and not any(lx in l.upper() for lx in LIXO):
                linhas.append(l)
        return '\n'.join(linhas).strip()

    # ─── INTERFACE ──────────────────────────────────────────────────
    f_conduta = st.file_uploader(
        "📂 Suba a planilha unificada (saída de Solicitados x Atendidos)",
        type=["xlsx","xls","csv"], key="f_conduta"
    )

    _MAPA_FILE = os.path.join(os.path.dirname(__file__), 'conduta_mapeamentos.json')
    if 'conduta_mapeamentos_custom' not in st.session_state:
        if os.path.exists(_MAPA_FILE):
            with open(_MAPA_FILE, 'r', encoding='utf-8') as _f:
                st.session_state['conduta_mapeamentos_custom'] = json.load(_f)
        else:
            st.session_state['conduta_mapeamentos_custom'] = {}

    if st.session_state['conduta_mapeamentos_custom']:
        with st.expander(f"📖 Condutas mapeadas por você ({len(st.session_state['conduta_mapeamentos_custom'])})"):
            for conduta_val, aba_val in list(st.session_state['conduta_mapeamentos_custom'].items()):
                col_a, col_b = st.columns([3,1])
                col_a.markdown(f"`{conduta_val}` → **{aba_val}**")
                if col_b.button("🗑️", key=f"del_map_{conduta_val}"):
                    del st.session_state['conduta_mapeamentos_custom'][conduta_val]
                    with open(_MAPA_FILE, 'w', encoding='utf-8') as _f:
                        json.dump(st.session_state['conduta_mapeamentos_custom'], _f, ensure_ascii=False, indent=2)
                    st.rerun()

    if f_conduta:
        if st.button("🚀 CLASSIFICAR E GERAR EXCEL POR PROCEDIMENTO", type="primary"):
            try:
                with st.spinner("Classificando condutas e organizando abas..."):
                    f_conduta.seek(0)
                    if f_conduta.name.lower().endswith('.csv'):
                        try:
                            df_c = pd.read_csv(f_conduta, sep=';', dtype=str)
                            if len(df_c.columns) < 2:
                                f_conduta.seek(0); df_c = pd.read_csv(f_conduta, sep=',', dtype=str)
                        except:
                            f_conduta.seek(0); df_c = pd.read_csv(f_conduta, sep=',', dtype=str)
                    elif f_conduta.name.lower().endswith('.xls'):
                        df_c = pd.read_excel(f_conduta, engine='xlrd', dtype=str, sheet_name=0)
                    else:
                        df_c = pd.read_excel(f_conduta, dtype=str, sheet_name=0)
                    df_c.columns = [str(c).strip().upper() for c in df_c.columns]

                    MAP_COLS = {
                            'NOME':'NOME','PACIENTE':'NOME',
                            'NUM':'NUM','NÚM':'NUM','NÚMERO':'NUM','NUMERO':'NUM',
                            'N°':'NUM','Nº':'NUM','NÚMERO DO PRONTUÁRIO':'NUM',
                            'TELEFONE':'TELEFONE','TEL.ADIC':'TEL.ADIC',
                            'TEL. ADICIONAL':'TEL.ADIC','FONE ADICIONAL':'TEL.ADIC',
                            'CONV':'CONV','CONVENIO':'CONV','CONVÊNIO':'CONV','CATEGORIA':'CONV',
                            'DATA VISITA':'DATA VISITA','DT. VISITA':'DATA VISITA','DATA':'DATA VISITA',
                            'MEDICO':'MEDICO','MÉDICO':'MEDICO',
                            'CONDUTA':'CONDUTA ORIGINAL','ATENDIMENTO':'CONDUTA ORIGINAL',
                            'PROXIMA':'PROXIMA_MED','PRÓXIMA':'PROXIMA_MED',
                            'CONDUTA MÉDICA':'PROXIMA_MED','CONDUTA MÉDICA LIMPA':'PROXIMA_MED',
                            'MSG2024':'MSG2024','MSG2025':'MSG2025','MSG2026':'MSG2026',
                            'PACIENTE GLAUCOMA':'PACIENTE GLAUCOMA',
                            'EMAIL':'EMAIL','E_MAIL':'EMAIL','E-MAIL':'EMAIL','MAIL':'EMAIL',
                            'CPF':'CPF','PROFISSAO':'PROFISSAO','PROFISSÃO':'PROFISSAO',
                            'DATA NASC':'DATA NASC','DATA NASCIMENTO':'DATA NASC',
                            'DT. NASCIMENTO':'DATA NASC','DT.NASCIMENTO':'DATA NASC',
                        }
                    
                    df_c = df_c.rename(columns={c: MAP_COLS.get(c, c) for c in df_c.columns})
                    for col in ['CONDUTA ORIGINAL','PROXIMA_MED']:
                        if col not in df_c.columns: df_c[col] = ''

                    # ── DEDUPLICAÇÃO por NUM+NOME antes de classificar ──────
                    # Mesma lógica da Tab 11: data mais recente vence,
                    # conduta médica é herdada se o mais recente não tiver
                    if 'NUM' in df_c.columns and 'NOME' in df_c.columns:
                        df_c['_num_key']  = df_c['NUM'].str.strip().str.upper().str.replace(r'\.0$','',regex=True)
                        df_c['_nome_key'] = df_c['NOME'].str.strip().str.upper().apply(
                            lambda n: re.sub(r'[^A-Z0-9]','',str(n))
                        )
                        df_c['_dt_key'] = pd.to_datetime(
                            df_c.get('DATA VISITA',''), format='%d/%m/%Y',
                            dayfirst=True, errors='coerce'
                        )
                        df_c = df_c.sort_values('_dt_key', ascending=False)
                        dedup = []
                        df_c['_orig_key'] = df_c['CONDUTA ORIGINAL'].str.strip().str.upper()
                        df_c['_med_key']  = df_c['MEDICO'].str.strip().str.upper()
                        for (nk, nmk, ok, mk), grp in df_c.groupby(
                            ['_num_key','_nome_key','_orig_key','_med_key'], dropna=False
                        ):
                            mestre = grp.iloc[0].copy()
                            # Herda CONDUTA MÉDICA se o mais recente não tiver
                            if not str(mestre.get('PROXIMA_MED','')).strip() or \
                               str(mestre.get('PROXIMA_MED','')).lower() in ['nan','none','']:
                                for _, r2 in grp.iloc[1:].iterrows():
                                    v = str(r2.get('PROXIMA_MED','')).strip()
                                    if v and v.lower() not in ['nan','none','']:
                                        mestre['PROXIMA_MED'] = v
                                        break

                            # Herda TELEFONE se o mais recente não tiver
                            if not str(mestre.get('TELEFONE','')).strip() or \
                               str(mestre.get('TELEFONE','')).lower() in ['nan','none','']:
                                for _, r2 in grp.iloc[1:].iterrows():
                                    tel = str(r2.get('TELEFONE','')).strip()
                                    if tel and tel.lower() not in ['nan','none','']:
                                        mestre['TELEFONE'] = tel
                                        tel_adic = str(r2.get('TEL.ADIC','')).strip()
                                        if tel_adic and tel_adic.lower() not in ['nan','none',''] \
                                           and not str(mestre.get('TEL.ADIC','')).strip():
                                            mestre['TEL.ADIC'] = tel_adic
                                        break
                            dedup.append(mestre)
                        n_antes = len(df_c)
                        df_c = pd.DataFrame(dedup)
                        n_depois = len(df_c)
                        df_c = df_c.drop(columns=['_num_key','_nome_key','_dt_key','_orig_key','_med_key'], errors='ignore')
                        if n_antes != n_depois:
                            st.info(f"🔁 {n_antes - n_depois} duplicados removidos antes de classificar ({n_antes} → {n_depois} pacientes únicos).")
                    # ────────────────────────────────────────────────────────

                    result_abas = {nome: [] for nome in ABAS_CONDUTA.keys()}
                    result_abas['⚠️ VERIFICAR'] = []
                    nao_identificados = {}

                    for _, row in df_c.iterrows():
                        c_raw = str(row.get('CONDUTA ORIGINAL','')).strip()
                        p_raw = str(row.get('PROXIMA_MED','')).strip()
                        if p_raw.lower() in ['nan','none','']: p_raw = ''

                        c_norm = _norm_c(c_raw)
                        p_norm = _norm_c(p_raw)
                        aba_custom = None
                        # JSON só casa com CONDUTA ORIGINAL exata — nunca com texto médico
                        for conduta_key, aba_val in st.session_state['conduta_mapeamentos_custom'].items():
                            if conduta_key.upper() == c_norm:
                                aba_custom = aba_val
                                break

                        medico_raw = str(row.get('MEDICO','')).strip().upper()
                        if aba_custom:
                            aba_dest, flag = aba_custom, ''
                        else:
                            aba_dest, flag = _classificar_conduta(c_raw, p_raw, medico_raw)

                        dias = _prazo_conduta(p_raw, c_raw)
                        flag_ilio = 'CONFERIR MANUALMENTE' in flag
 
                        if aba_dest == '⚠️ VERIFICAR':
                            conduta_std = '⚠️ VERIFICAR'
                            prox_calc = ''
                        elif flag_ilio:
                            # ILIO → FACO com próxima manual
                            conduta_std = 'FACO — CATARATA'
                            prox_calc = '📋 CONFERIR MANUALMENTE'
                        elif dias == -1:
                            # Data explícita no texto (ex: RETORNO EM 24/04)
                            import re as _re2
                            data_vis = formatar_brasileiro_sem_hora(str(row.get('DATA VISITA','')))
                            txt_data = _norm_c(p_raw) + ' ' + _norm_c(c_raw)
                            m_d = _re2.search(r'EM\s+(\d{1,2}/\d{2})(?:\s*\()?', txt_data)
                            if m_d:
                                ano = data_vis.split('/')[-1] if data_vis else '2026'
                                conduta_std = f'RETORNO EM {m_d.group(1)}/{ano}'
                                prox_calc = f'{m_d.group(1)}/{ano}'
                            else:
                                conduta_std = 'CONSULTAS ESPECIALIZADAS'
                                prox_calc = '📋 CONFERIR MANUALMENTE'

                        else:
                            # Abas de procedimento cirúrgico/exame SEM prazo definido
                            # → próxima = CONFERIR MANUALMENTE
                            abas_sem_prazo_obrigatorio = {
                                'FACO — CATARATA','FACO REFRATIVA','RETINA',
                                'GLAUCOMA CIRÚRGICO','CERATOCONE — ANEL',
                                'TRANSPLANTE DE CÓRNEA','YAG LASER','PTERÍGIO',
                                'CALÁZIO','OCULOPLÁSTICA','ESTRABISMO',
                                'EXAMES DIAGNOSE','LUZ PULSADA','MEIBOMIOGRAFIA',
                                'LENTES DE CONTATO',
                            }
                            if dias:
                                conduta_std = MAP_PRAZO_CONDUTA.get(dias,
                                    f'NOVA CONSULTA EM {dias} DIAS')
                                prox_calc = calcular_proxima_data(
                                    formatar_brasileiro_sem_hora(str(row.get('DATA VISITA',''))),
                                    conduta_std
                                )
                            elif aba_dest in abas_sem_prazo_obrigatorio:
                                # Procedimento sem prazo informado → sinalizar
                                conduta_std = aba_dest.upper()
                                prox_calc = '📋 CONFERIR MANUALMENTE'
                            else:
                                # CONSULTAS sem prazo identificável → anual por padrão
                                conduta_std = 'CONSULTA ANUAL'
                                prox_calc = calcular_proxima_data(
                                    formatar_brasileiro_sem_hora(str(row.get('DATA VISITA',''))),
                                    'CONSULTA ANUAL'
                                )

                        reg = {
                            'NUM': str(row.get('NUM', row.get('NÚM', row.get('NÚMERO', row.get('NUMERO', ''))))).replace('nan','').replace('.0','').strip(),
                            'NOME': str(row.get('NOME','')).upper().strip().replace('nan',''),
                            'TELEFONE': str(row.get('TELEFONE','')).replace('nan','').strip(),
                            'TEL.ADIC': str(row.get('TEL.ADIC','')).replace('nan','').strip(),
                            'CONV': str(row.get('CONV','')).replace('nan','').strip(),
                            'DATA VISITA': formatar_brasileiro_sem_hora(str(row.get('DATA VISITA',''))),
                            'MEDICO': str(row.get('MEDICO','')).replace('nan','').strip(),
                            'CONDUTA ORIGINAL': c_raw,
                            'CONDUTA MÉDICA LIMPA': _limpar_conduta_medica(p_raw),
                            'CONDUTA PADRONIZADA': conduta_std,
                            'PRÓXIMA CALCULADA': prox_calc,
                            'ALERTA': flag if flag else '',
                            'MSG2024': str(row.get('MSG2024','')).replace('nan',''),
                            'MSG2025': str(row.get('MSG2025','')).replace('nan',''),
                            'MSG2026': str(row.get('MSG2026','')).replace('nan',''),
                            'PACIENTE GLAUCOMA': str(row.get('PACIENTE GLAUCOMA','')).replace('nan',''),
                            'EMAIL': str(row.get('EMAIL','')).replace('nan',''),
                            'CPF': str(row.get('CPF','')).replace('nan','').strip(),
                            'PROFISSAO': str(row.get('PROFISSAO','')).replace('nan','').strip(),
                            'DATA NASC': formatar_brasileiro_sem_hora(str(row.get('DATA NASC',''))),
                        }

                        if aba_dest is None:
                            chave = c_raw if c_raw else p_raw[:60]
                            if chave not in nao_identificados:
                                nao_identificados[chave] = []
                            nao_identificados[chave].append(reg)
                        elif aba_dest == '⚠️ VERIFICAR':
                            reg['ALERTA'] = flag
                            result_abas['⚠️ VERIFICAR'].append(reg)
                        else:
                            result_abas[aba_dest].append(reg)

                    # ── CLASSIFICAÇÃO POR IA para casos não identificados ──
                    if nao_identificados:
                        total_nao_id = sum(len(v) for v in nao_identificados.values())
                        with st.spinner(f"🤖 IA classificando {total_nao_id} casos não identificados..."):
                            try:
                                # Monta lista de casos para a IA
                                casos_ia = []
                                for chave, regs in nao_identificados.items():
                                    for reg in regs:
                                        casos_ia.append({
                                            'id': len(casos_ia),
                                            'chave': chave,
                                            'reg': reg,
                                            'medico': reg.get('MEDICO',''),
                                            'conduta_original': reg.get('CONDUTA ORIGINAL',''),
                                            'conduta_medica': reg.get('CONDUTA MÉDICA LIMPA',''),
                                        })

                                # Monta o prompt em lote
                                lista_abas = '\n'.join([f'- {a}' for a in list(ABAS_CONDUTA.keys()) + ['⚠️ VERIFICAR']])
                                casos_txt = '\n'.join([
                                    f'[{c["id"]}] MEDICO={c["medico"]} | CONDUTA={c["conduta_original"]} | TEXTO_MEDICO={str(c["conduta_medica"])[:150]}'
                                    for c in casos_ia
                                ])

                                prompt = f"""Você é um assistente especializado em oftalmologia do Hospital de Olhos Vale do Aço (HOVA).

Sua tarefa é classificar cada paciente na aba correta de procedimento, baseado no médico, conduta original e texto da conduta médica.

ABAS DISPONÍVEIS:
{lista_abas}

REGRAS IMPORTANTES:
- ALTAIR: especialista em catarata → RC/PO/AC/R sem info específica = FACO — CATARATA
- DENISE: retinologista → RC/PO/AC/R/FOTO = RETINA
- GABRIELL/GABRIEL: retina → RC/PO/R/nan = RETINA (exceto se texto diz outro procedimento)
- GUSTAVO: glaucoma → RC/PO/R/FOTOTRAB/IRI = GLAUCOMA CIRÚRGICO
- MARILUCI: pterígio/calázio → RC/PO/AC/R = PTERÍGIO (CALÁZIO se texto menciona calázio/tumor)
- FELIPE: luz pulsada → mas se texto médico lista exames = EXAMES DIAGNOSE
- VERA: estrabismo → sempre ESTRABISMO
- EXAMES: sempre EXAMES DIAGNOSE
- LENTES: sempre LENTES DE CONTATO
- ENFERMAG: depende do texto médico (FACO/IIV/CICLO/PTERÍGIO etc)
- Se conduta médica menciona FACO/LIO/CATARATA = FACO — CATARATA
- Se conduta médica menciona IIV/ANTIVEGF/VVPP/FOTOCOAGULAÇÃO = RETINA
- Se conduta médica menciona CICLOFOTO/TRABECULECT/SLT/FOTOTRAB = GLAUCOMA CIRÚRGICO
- Se conduta médica menciona PTERÍGIO/PTERIGIO = PTERÍGIO
- Se conduta médica menciona CALÁZIO/TUMOR PALPEBRAL = CALÁZIO
- Se conduta médica menciona YAG/CAPSULOTOMIA = YAG LASER
- Se conduta médica lista exames (OCT, CAMPIMETRIA, RETINOGRAFIA etc) = EXAMES DIAGNOSE
- Se não há informação suficiente = ⚠️ VERIFICAR

CASOS PARA CLASSIFICAR:
{casos_txt}

Responda SOMENTE em JSON válido, sem texto adicional, sem markdown:
{{"classificacoes": [{{"id": 0, "aba": "NOME DA ABA"}}, ...]}}"""

                                import json as _json
                                response = __import__('urllib.request', fromlist=['urlopen']).urlopen(
                                    __import__('urllib.request', fromlist=['Request']).Request(
                                        'https://api.anthropic.com/v1/messages',
                                        data=_json.dumps({
                                            'model': 'claude-haiku-4-5-20251001',
                                            'max_tokens': 8192,
                                            'messages': [{'role': 'user', 'content': prompt}]
                                        }).encode(),
                                       headers={
                                            'Content-Type': 'application/json',f'[{c["id"]}] MEDICO={c["medico"]} | CONDUTA={c["conduta_original"]} | TEXTO_MEDICO={str(c["conduta_medica"])[:300]}'
                                            'anthropic-version': '2023-06-01',
                                            'x-api-key': st.secrets.get('ANTHROPIC_API_KEY', '')
                                        },
                                        method='POST'
                                    )
                                )
                                resultado = _json.loads(response.read().decode())
                                texto_resposta = resultado['content'][0]['text'].strip()
                                # Limpa possíveis marcadores markdown
                                if texto_resposta.startswith('```'):
                                    texto_resposta = '\n'.join(texto_resposta.split('\n')[1:])
                                if texto_resposta.endswith('```'):
                                    texto_resposta = '\n'.join(texto_resposta.split('\n')[:-1])
                                classificacoes = _json.loads(texto_resposta)['classificacoes']
                                mapa_ia = {c['id']: c['aba'] for c in classificacoes}

                               # Aplica classificações da IA em lotes de 200
                                import math as _math
                                mapa_ia = {}
                                tamanho_lote = 200
                                total_lotes = _math.ceil(len(casos_ia) / tamanho_lote)
                                for n_lote in range(total_lotes):
                                    lote = casos_ia[n_lote*tamanho_lote:(n_lote+1)*tamanho_lote]
                                    casos_txt_lote = '\n'.join([
                                        f'[{caso_item["id"]}] MEDICO={caso_item["medico"]} | CONDUTA={caso_item["conduta_original"]} | TEXTO_MEDICO={str(caso_item["conduta_medica"])[:300]}'
                                        for caso_item in lote
                                    ])
                                    prompt_lote = prompt.replace(casos_txt, casos_txt_lote)
                                    resp_lote = __import__('urllib.request', fromlist=['urlopen']).urlopen(
                                        __import__('urllib.request', fromlist=['Request']).Request(
                                            'https://api.anthropic.com/v1/messages',
                                            data=_json.dumps({
                                                'model': 'claude-haiku-4-5-20251001',
                                                'max_tokens': 8192,
                                                'messages': [{'role': 'user', 'content': prompt_lote}]
                                            }).encode(),
                                            headers={
                                                'Content-Type': 'application/json',
                                                'anthropic-version': '2023-06-01',
                                                'x-api-key': st.secrets.get('ANTHROPIC_API_KEY', '')
                                            },
                                            method='POST'
                                        )
                                    )
                                    res_lote = _json.loads(resp_lote.read().decode())
                                    txt_lote = res_lote['content'][0]['text'].strip()
                                    if txt_lote.startswith('```'): txt_lote = '\n'.join(txt_lote.split('\n')[1:])
                                    if txt_lote.endswith('```'): txt_lote = '\n'.join(txt_lote.split('\n')[:-1])
                                    for cls_item in _json.loads(txt_lote)['classificacoes']:
                                        mapa_ia[cls_item['id']] = cls_item['aba']

                                ia_contadores = {}
                                for caso in casos_ia:
                                    aba_ia = mapa_ia.get(caso['id'], '⚠️ VERIFICAR')
                                    # Valida que a aba existe
                                    abas_validas = list(ABAS_CONDUTA.keys()) + ['⚠️ VERIFICAR']
                                    if aba_ia not in abas_validas:
                                        aba_ia = '⚠️ VERIFICAR'
                                    caso['reg']['ALERTA'] = f'🤖 IA' if aba_ia != '⚠️ VERIFICAR' else '❓ IA: sem padrão'
                                    result_abas[aba_ia].append(caso['reg'])
                                    ia_contadores[aba_ia] = ia_contadores.get(aba_ia, 0) + 1

                                ia_resumo = ' | '.join([f'{a}: {n}' for a, n in sorted(ia_contadores.items())])
                                st.info(f"🤖 IA classificou {total_nao_id} casos → {ia_resumo}")

                            except Exception as e_ia:
                                st.warning(f"⚠️ IA indisponível ({str(e_ia)[:60]}). Casos enviados para VERIFICAR.")
                                for chave, regs in nao_identificados.items():
                                    for r in regs:
                                        r['ALERTA'] = f'❓ CONDUTA NÃO MAPEADA: {chave[:50]}'
                                        result_abas['⚠️ VERIFICAR'].append(r)

                    # ─── GERAR EXCEL ──────────────────────────────────────
                    from openpyxl.utils import get_column_letter
                    wb_out = openpyxl.Workbook()
                    wb_out.remove(wb_out.active)

                    ws_res = wb_out.create_sheet('RESUMO', 0)
                    ws_res.sheet_view.showGridLines = False
                    for i, h in enumerate(['ABA','TOTAL PACIENTES'], 1):
                        c = ws_res.cell(row=1, column=i, value=h)
                        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
                        c.fill = PatternFill(start_color='1e3d3a', end_color='1e3d3a', fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center')
                    ws_res.row_dimensions[1].height = 28
                    ws_res.column_dimensions['A'].width = 34
                    ws_res.column_dimensions['B'].width = 18

                    todas_abas_ordem = list(ABAS_CONDUTA.keys()) + ['⚠️ VERIFICAR']
                    cores_resumo = {**{n: ABAS_CONDUTA[n]['cor'] for n in ABAS_CONDUTA}, '⚠️ VERIFICAR': '922B21'}
                    cores_linha  = {**{n: ABAS_CONDUTA[n]['cor_linha'] for n in ABAS_CONDUTA}, '⚠️ VERIFICAR': 'FDEDEC'}

                    row_r = 2; total_geral = 0
                    for nome_aba in todas_abas_ordem:
                        regs = result_abas.get(nome_aba, [])
                        if not regs: continue
                        c1 = ws_res.cell(row=row_r, column=1, value=nome_aba)
                        c1.font = Font(name='Arial', size=10, bold=True, color=cores_resumo[nome_aba])
                        c1.alignment = Alignment(vertical='center')
                        c2 = ws_res.cell(row=row_r, column=2, value=len(regs))
                        c2.font = Font(name='Arial', size=10, bold=True)
                        c2.alignment = Alignment(horizontal='center', vertical='center')
                        if row_r % 2 == 0:
                            for ci in [c1, c2]:
                                ci.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        total_geral += len(regs); row_r += 1

                    for i, val in enumerate(['TOTAL GERAL', total_geral], 1):
                        c = ws_res.cell(row=row_r+1, column=i, value=val)
                        c.font = Font(name='Arial', bold=True, size=11, color='1e3d3a')
                        c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                    for nome_aba in todas_abas_ordem:
                        regs = result_abas.get(nome_aba, [])
                        if not regs: continue
                        cor_h = cores_resumo[nome_aba]; cor_l = cores_linha[nome_aba]
                        ws = wb_out.create_sheet(nome_aba[:31])
                        for ci, col_name in enumerate(COLUNAS_CONDUTA, 1):
                            cell = ws.cell(row=1, column=ci, value=col_name)
                            cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
                            cell.fill = PatternFill(start_color=cor_h, end_color=cor_h, fill_type='solid')
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            ws.column_dimensions[get_column_letter(ci)].width = LARGURAS_CONDUTA.get(col_name, 14)
                        ws.row_dimensions[1].height = 28
                        for ri, reg in enumerate(regs, 2):
                            fill = PatternFill(start_color=cor_l, end_color=cor_l, fill_type='solid') if ri % 2 == 0 else None
                            for ci, col_name in enumerate(COLUNAS_CONDUTA, 1):
                                val = reg.get(col_name, '')
                                cell = ws.cell(row=ri, column=ci, value=val if val else None)
                                cell.font = Font(name='Arial', size=9)
                                cell.alignment = Alignment(vertical='center', wrap_text=True)
                                if fill: cell.fill = fill
                                if col_name == 'ALERTA' and val:
                                    cell.font = Font(name='Arial', size=9, bold=True, color='9C0006')
                                if col_name == 'CONDUTA PADRONIZADA' and val and '⚠️' in str(val):
                                    cell.font = Font(name='Arial', size=9, bold=True, color='FF0000')
                        ws.freeze_panes = 'A2'
                        ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUNAS_CONDUTA))}1'
                        ws.cell(row=len(regs)+3, column=1, value=f'Total: {len(regs)} pacientes').font = Font(bold=True, italic=True, size=9)

                    buf_out = io.BytesIO()
                    wb_out.save(buf_out); buf_out.seek(0)
                    st.session_state['conduta_excel'] = buf_out.getvalue()
                    st.session_state['conduta_nao_identificados'] = nao_identificados
                    st.session_state['conduta_result_abas'] = {k: len(v) for k, v in result_abas.items() if v}
                    st.success(f"✅ {total_geral} pacientes classificados em {len([n for n in todas_abas_ordem if result_abas.get(n)])} abas!")

            except Exception as e:
                st.error(f"Erro ao processar: {e}")
                import traceback; st.code(traceback.format_exc())

        if 'conduta_result_abas' in st.session_state:
            col_res1, col_res2 = st.columns([2, 1])
            with col_res1:
                st.markdown("**Resultado por aba:**")
                rows = []
                for aba_n, cnt in st.session_state['conduta_result_abas'].items():
                    rows.append({"Aba": aba_n, "Pacientes": cnt})
                if rows:
                    df_res = pd.DataFrame(rows)

                    def _colorir_resultado(row):
                        if "VERIFICAR" in str(row["Aba"]):
                            return ['background-color:#FFF5F5; color:#8B1A1A; font-weight:500',
                                    'background-color:#FFF5F5; color:#8B1A1A; font-weight:600']
                        return ['color:#0f2421', 'color:#0f2421; font-weight:600']

                    st.dataframe(
                        df_res.style.apply(_colorir_resultado, axis=1),
                        use_container_width=True,
                        hide_index=True,
                        height=min(40 * len(rows) + 40, 520),
                    )

            nao_id = st.session_state.get('conduta_nao_identificados', {})
            if nao_id:
                st.markdown("---")
                st.warning(f"⚠️ **{len(nao_id)} condutas não reconhecidas.** Me diga para qual aba cada uma deve ir:")
                for conduta_val in list(nao_id.keys()):
                    if conduta_val in st.session_state['conduta_mapeamentos_custom']:
                        continue
                    qtd = len(nao_id[conduta_val])
                    col_q1, col_q2, col_q3 = st.columns([3, 2, 1])
                    col_q1.markdown(f"**`{conduta_val[:60]}`** ({qtd} paciente{'s' if qtd > 1 else ''})")
                    opcoes_abas = ['— Selecione —'] + list(ABAS_CONDUTA.keys()) + ['⚠️ VERIFICAR', 'IGNORAR / DESCARTAR']
                    aba_escolhida = col_q2.selectbox('→', opcoes_abas, key=f"mapeia_{conduta_val}")
                    if col_q3.button('✅', key=f"confirma_{conduta_val}"):
                        if aba_escolhida != '— Selecione —':
                            st.session_state['conduta_mapeamentos_custom'][conduta_val] = aba_escolhida
                            with open(_MAPA_FILE, 'w', encoding='utf-8') as _f:
                                json.dump(st.session_state['conduta_mapeamentos_custom'], _f, ensure_ascii=False, indent=2)
                            st.success("Mapeado! Clique em Processar novamente para aplicar.")
                            st.rerun()

        if 'conduta_excel' in st.session_state:
            st.download_button(
                "📥 BAIXAR EXCEL ORGANIZADO POR PROCEDIMENTO",
                st.session_state['conduta_excel'],
                "HOVA_Conduta_Organizada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True
            )

    st.markdown('</div>', unsafe_allow_html=True)
    
st.markdown('<div class="footer-master">HOVA MASTER INTELLIGENCE — UNIDADE IPATINGA</div>', unsafe_allow_html=True)
